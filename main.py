import os  # Работа с операционной системой и путями к файлам
import sys  # Доступ к системным переменным и параметрам интерпретатора
import math  # Выполнение инженерных математических расчетов
import openpyxl # Библиотека для чтения данных из таблиц Excel (.xlsx)

# --- ПУТИ К ФАЙЛАМ (Оптимизировано для Android 13-14) --- # 
# Мы НЕ используем os.chdir, так как это вызывает вылет на новых Android.
BASE_DIR = os.path.dirname(os.path.abspath(__file__)) # Определение папки, где лежит main.py
BACKGROUND_PATH = os.path.join(BASE_DIR, 'background.png') # Путь к фоновой картинке / заставке
ICON_PATH = os.path.join(BASE_DIR, 'icon.png') # Путь к значку приложения
LOGO_PATH = os.path.join(BASE_DIR, 'logo.png') # Путь к логотипу для интерфейса

# Пути к твоим базам данных Excel
EXCEL_PATH = os.path.join(BASE_DIR, 'data', 'nkt_base.xlsx') # База данных по трубам НКТ
CA320_PATH = os.path.join(BASE_DIR, 'data', 'ca320_stats.xlsx') # Технические характеристики агрегата ЦА-320
VZD_PATH = os.path.join(BASE_DIR, 'data', 'vzd_catalog.xlsx') # Справочник винтовых забойных двигателей

# --- БАЗОВЫЕ ИМПОРТЫ KIVY ---
from kivy.metrics import dp, sp # Функции адаптивного масштабирования интерфейса (dp) и текста (sp)
from kivy.uix.screenmanager import ScreenManager, Screen, FadeTransition # Менеджер экранов, класс экрана и эффект плавного перехода
from kivy.uix.widget import Widget # Базовый пустой элемент интерфейса
from kivy.uix.scrollview import ScrollView # Контейнер для прокрутки содержимого (скролл)
from kivy.uix.floatlayout import FloatLayout # Макет со свободным (абсолютным) позиционированием элементов
from kivy.core.window import Window # Управление окном приложения (размеры, клавиатура)
from kivy.core.image import Image as CoreImage # Низкоуровневая загрузка графических текстур в память
from kivy.graphics import Color, Line, Rectangle, Ellipse # Инструменты для рисования фигур на холсте (Canvas)
from kivy.clock import Clock # Таймер для отложенных функций и анимаций
from kivy.properties import ObjectProperty, NumericProperty, StringProperty # Специальные динамические свойства Kivy для связывания данных
from kivy.storage.jsonstore import JsonStore # Инструмент для локального сохранения настроек в файл .json
from kivy.storage.jsonstore import JsonStore # Дублирующийся импорт JsonStore
from kivy.lang import Builder # Добавили импорт Builder! Загрузчик разметки интерфейса
from kivymd.uix.boxlayout import MDBoxLayout # Вертикальный или горизонтальный контейнер Material Design
from kivymd.uix.floatlayout import MDFloatLayout # Добавили эту строчку! Свободный макет Material Design
from kivy.utils import platform # Добавили эту строчку! Определение ОС (Android, Windows, iOS)

# --- КОМПОНЕНТЫ KIVYMD ---
from kivymd.app import MDApp # Главный класс для создания Material Design приложения
from kivymd.uix.screen import MDScreen # Класс экрана, адаптированный под Material Design
from kivymd.uix.boxlayout import MDBoxLayout # Повторный импорт контейнера MDBoxLayout
from kivymd.uix.gridlayout import MDGridLayout # Табличный макет Material Design (сетка)
from kivymd.uix.anchorlayout import MDAnchorLayout # Макет с привязкой элементов к краям или центру
from kivymd.uix.menu import MDDropdownMenu # Выпадающие списки выбора
from kivymd.uix.textfield import MDTextField # Поля для ввода текста пользователем
from kivymd.uix.label import MDLabel # Текстовые метки Material Design
from kivymd.uix.fitimage import FitImage # Изображение, автоматически заполняющее выделенную область
from kivymd.uix.button import (
    MDRaisedButton, # Кнопка с заливкой цветом и тенью
    MDFillRoundFlatIconButton, # Закругленная кнопка с заливкой и иконкой
    MDFillRoundFlatButton, # Закругленная кнопка с заливкой
    MDIconButton, # Простая кнопка-иконка без текста
    MDFlatButton # Плоская кнопка без заливки и тени (для диалогов)
)
from kivymd.uix.progressbar import MDProgressBar # Индикатор прогресса загрузки/вычислений
from kivymd.uix.snackbar import Snackbar # Всплывающие информационные полоски внизу экрана
from kivymd.uix.card import MDCard # Информационные карточки-блоки
from kivymd.uix.dialog import MDDialog # Всплывающие диалоговые окна (окна действий)
from kivymd.uix.list import TwoLineListItem

# Обычные Kivy компоненты для верстки
from kivy.uix.label import Label # Базовый класс текстовой метки Kivy
from kivy.uix.image import Image # Базовый класс изображения Kivy
from kivy.core.image import Image as CoreImage # Повторный низкоуровневый импорт картинки

# Фиксация масштаба шрифта, чтобы системные настройки Android не ломали меню
from kivy.metrics import Metrics
Metrics.font_scale = 1.0 # Блокировка масштаба текста на значении 100% для сохранения верстки

import webbrowser # Модуль для открытия интернет-ссылок в системном браузере


class BaseScreen(MDScreen): # Родоначальник всех экранов. Создает адаптивную подложку и управляет фокусом полей ввода.
    # Фундамент. Отвечает за «умный» фон, который подстраивается под экран без искажений.
    def __init__(self, **kw):
        super().__init__(**kw) # Инициализируем стандартные свойства родительского класса MDScreen
        app = MDApp.get_running_app() # Связываемся с запущенным главным инстансом приложения
        
        # Пытаемся получить текстуру фона, загруженную в главном классе приложения
        self.bg_texture = getattr(app, 'global_texture', None) # Вытаскиваем общую текстуру фона
        
        with self.canvas.before: # Начинаем отрисовку графического контекста ПОД всеми кнопками и текстами экрана
            # Слой 1: Сама картинка фона
            Color(1, 1, 1, 1) # Сбросываем цветовой фильтр в белый
            self.bg_rect = Rectangle(pos=self.pos, size=self.size) # Рисуем прямоугольник под размеры текущего экрана
            if self.bg_texture:
                self.bg_rect.texture = self.bg_texture # Если картинка загрузилась, натягиваем её текстуру
                
            # Слой 2: Затемнение (эффект тонировки)
            # Черный цвет с прозрачностью 70%
            Color(0, 0, 0, 0.7) # Задаем черный цвет и альфа-канал 0.7 для читаемости белого текста поверх картинки
            self.fade_rect = Rectangle(pos=self.pos, size=self.size) # Рисуем полупрозрачную вуаль поверх фона
            
        # Обновляем размеры при изменении окна (поворот экрана)
        self.bind(size=self._update_bg, pos=self._update_bg) # Привязываем автоматический перерасчет фона к геометрии
        
        # КЛАВИАТУРНЫЙ ФИКС: Привязываем отслеживание высоты выезжающей клавиатуры к методу _on_keyboard_height
        Window.bind(on_keyboard_height=self._on_keyboard_height)

    def _update_bg(self, instance, value):
        # Метод перерисовывает слои фона, если экран телефона повернулся или изменился
        self.fade_rect.pos = self.bg_rect.pos = instance.pos 
        self.fade_rect.size = self.bg_rect.size = instance.size 
        
        if self.bg_texture:
            win_w, win_h = instance.size 
            img_w, img_h = self.bg_texture.size 
            win_ratio = win_w / max(win_h, 1) 
            img_ratio = img_w / max(img_h, 1) 
            
            # Алгоритм обрезки лишнего, чтобы сохранить пропорции (Aspect Ratio)
            if win_ratio > img_ratio:
                h_diff = (1 - img_ratio / win_ratio) / 2
                coords = (0, 1 - h_diff, 1, 1 - h_diff, 1, h_diff, 0, h_diff) 
            else:
                w_diff = (1 - win_ratio / img_ratio) / 2
                coords = (w_diff, 1, 1 - w_diff, 1, 1 - w_diff, 0, w_diff, 0) 
                
            self.bg_rect.tex_coords = coords 

    def _on_keyboard_height(self, window, height):
        """Фоновый планировщик: срабатывает на Android при изменении высоты клавиатуры"""
        if not self.manager or self.manager.current != self.name:
            return # Страховка: обрабатываем событие только на том экране, который сейчас активен перед глазами инженера
            
        # Пытаемся автоматически найти зону прокрутки ScrollView на текущем активном экране
        scroll_view = None
        for child in self.walk(): # Метод walk() глубоко сканирует всю иерархию виджетов экрана
            if isinstance(child, ScrollView):
                scroll_view = child
                break
                
        if not scroll_view:
            return # Если на экране нет скролла (например, главное меню), то двигать нечего, выходим
            
        if height > 0:
            # СЦЕНАРИЙ: Клавиатура выехала
            # Ищем, какое именно текстовое поле (MDTextField) сейчас находится в фокусе (нажато инженером)
            focused_widget = None
            for child in self.walk():
                if hasattr(child, 'focus') and child.focus:
                    focused_widget = child
                    break
                    
            if focused_widget:
                # Находим относительные координаты нажатого поля ввода внутри общего ScrollView
                widget_pos = focused_widget.to_window(*focused_widget.pos)
                widget_y = widget_pos[1] # Фиксируем вертикальную позицию поля по оси Y
                
                # Если клавиатура перекрывает поле ввода (координата Y поля ниже, чем высота выехавшей клавиатуры)
                if widget_y < height + dp(40):
                    # Вычисляем необходимую величину плавного докрута скролла вверх
                    extra_scroll = (height + dp(60) - widget_y) / max(scroll_view.height, 1)
                    # Корректируем положение ползунка скролла (в Kivy это доля от 0.0 до 1.0)
                    new_scroll_y = max(0.0, min(1.0, scroll_view.scroll_y + extra_scroll))
                    
                    # Плавная анимация докрутки: сдвигаем поле в зону видимости за 0.15 секунды
                    from kivy.animation import Animation
                    Animation(scroll_y=new_scroll_y, d=0.15, t='out_quad').start(scroll_view)
        else:
            # СЦЕНАРИЙ: Клавиатуру закрыли (height == 0)
            # Если скролл был сильно перекручен, плавно возвращаем его к стандартным границам
            if scroll_view.scroll_y < 0 or scroll_view.scroll_y > 1:
                from kivy.animation import Animation
                target_y = max(0.0, min(1.0, scroll_view.scroll_y))
                Animation(scroll_y=target_y, d=0.15, t='out_quad').start(scroll_view)


class WellSchematic(Widget):               #Движок отрисовки. Рисует </kivy.uix.image.Image object at 0x00000174CD1C50F0>конструкцию скважины (колонну, хвостовик, ГНО) и анимирует движение раствора.
    def __init__(self, hz=2000, hv=1500, gno_type="Воронка", l_hvost=0, method="ПРЯМАЯ", **kwargs):
        super().__init__(**kwargs)
        try:
            self.hz = float(hz) if hz else 2000.0
            self.hv = float(hv) if hv else 1500.0
            self.l_hvost = float(l_hvost) if l_hvost else 0.0
        except:
            self.hz, self.hv, self.l_hvost = 2000.0, 1500.0, 0.0
            
        self.gno_type = str(gno_type).upper()
        self.method = str(method).upper()
        self.progress = 0
        Clock.schedule_interval(self.animate, 1/60)
        self.bind(pos=self.update_canvas, size=self.update_canvas)

    def animate(self, dt):
        if self.progress < 2.0:
            self.progress += 0.01
            self.update_canvas()
            return True
        return False

    def update_canvas(self, *args):
        self.canvas.clear()
        cx, cy = self.center_x, self.y
        w, h = self.size
        m = dp(90) # Верхний отступ для ФА
        s = (h - m * 2) / self.hz if self.hz > 0 else 1
        
        y_t = cy + h - m
        y_v = y_t - (self.hv * s)
        y_z = cy + m
        y_h = y_t - ((self.hz - self.l_hvost) * s)

        with self.canvas:
            # --- 1. ВОЗВРАЩАЕМ РОДНОЙ ФОН ПРИЛОЖЕНИЯ ИЗ ГЛАВНОГО ЯДРА ---
            from kivymd.app import MDApp
            app = MDApp.get_running_app()
            
            if hasattr(app, 'global_texture') and app.global_texture:
                # Если текстура фона загружена, накатываем её на весь экран схемы
                Color(1, 1, 1, 0.45) # opacity=0.45, как на главном экране
                Rectangle(pos=self.pos, size=self.size, texture=app.global_texture)
            else:
                # На случай, если картинка не нашлась — строгий темный цвет, чтобы не было пустоты
                Color(0.08, 0.09, 0.12, 1) 
                Rectangle(pos=self.pos, size=self.size)

            # --- 2. ВНУТРЕННИЙ ТЕМНЫЙ ФОН СКВАЖИНЫ (СТРОГО ПОД ЗЕМЛЕЙ) ---
            Color(0.1, 0.1, 0.15, 1) 
            Rectangle(pos=(cx - dp(30), y_z - dp(10)), size=(dp(60), y_t - y_z + dp(10)))

            # --- 3. КОЛОННА (Металлические стенки скважины 168мм) ---
            Color(0.4, 0.4, 0.45, 1) # Тень
            Rectangle(pos=(cx - dp(26), y_h), size=(dp(3), y_t - y_h))
            Rectangle(pos=(cx + dp(23), y_h), size=(dp(3), y_t - y_h))
            Color(0.8, 0.8, 0.85, 1) # Блик
            Rectangle(pos=(cx - dp(25), y_h), size=(dp(1.5), y_t - y_h))
            Rectangle(pos=(cx + dp(24), y_h), size=(dp(1.5), y_t - y_h))

            # Хвостовик 114
            if self.l_hvost > 0:
                Color(0.3, 0.3, 0.35, 1)
                Rectangle(pos=(cx - dp(19), y_z), size=(dp(2), y_h - y_z))
                Rectangle(pos=(cx + dp(17), y_z), size=(dp(2), y_h - y_z))
                Line(points=[cx - dp(25), y_h, cx - dp(18), y_h, cx - dp(18), y_z, cx + dp(18), y_z, cx + dp(18), y_h, cx + dp(25), y_h], width=dp(1.5))
            else:
                Line(points=[cx - dp(25), y_z, cx + dp(25), y_z], width=dp(2)) # Забой

            # --- 4. АНИМАЦИЯ РАСТВОРА (Твоя оригинальная логика) ---
            p = self.progress
            
            # ЭТАП 1: Заполнение НКТ
            p1 = min(p / 0.6, 1.0)
            h1 = p1 * (self.hv * s)
            Color(0.1, 0.7, 0.2, 0.9) 
            Rectangle(pos=(cx - dp(6), y_t - h1), size=(dp(12), h1))
            Color(1, 1, 1, 0.2) 
            Rectangle(pos=(cx - dp(1), y_t - h1), size=(dp(2), h1))

            # ЭТАП 2: Забойное пространство
            if p > 0.6:
                p2 = min((p - 0.6) / 0.6, 1.0)
                h_total_pod = (self.hz - self.hv) * s
                fill_h2 = p2 * h_total_pod
                
                Color(0.05, 0.5, 0.1, 0.8) 
                if y_v > y_h:
                    h_col = min(fill_h2, (y_v - y_h))
                    Rectangle(pos=(cx - dp(23), y_v - h_col), size=(dp(46), h_col))
                    if fill_h2 > (y_v - y_h):
                        h_z = min(fill_h2 - (y_v - y_h), (y_h - y_z))
                        Rectangle(pos=(cx - dp(17), y_h - h_z), size=(dp(34), h_z))
                else:
                    h_z = min(fill_h2, (y_v - y_z))
                    Rectangle(pos=(cx - dp(17), y_v - h_z), size=(dp(34), h_z))

            # ЭТАП 3: Подъем по затрубу
            if p > 1.2 and "УСТ." not in self.gno_type:
                p3 = min((p - 1.2) / 0.8, 1.0)
                h3 = p3 * (self.hv * s)
                Color(0.1, 0.6, 0.2, 0.8)
                Rectangle(pos=(cx - dp(23), y_v), size=(dp(17), h3))
                Rectangle(pos=(cx + dp(6), y_v), size=(dp(17), h3))

            # --- 5. НКТ ---
            Color(0.9, 0.9, 0.9, 0.6)
            Line(points=[cx - dp(7), y_t, cx - dp(7), y_v], width=dp(1))
            Line(points=[cx + dp(7), y_t, cx + dp(7), y_v], width=dp(1))

            # --- 6. ОБОРУДОВАНИЕ (ГНО) ---
            if "ПАКЕР" in self.gno_type:
                Color(0.2, 0.2, 0.2, 1) 
                Rectangle(pos=(cx - dp(24), y_v), size=(dp(17), dp(16)))
                Rectangle(pos=(cx + dp(7), y_v), size=(dp(17), dp(16)))
                Color(1, 0.7, 0, 1) 
                Rectangle(pos=(cx - dp(24), y_v + dp(6)), size=(dp(17), dp(3)))
                Rectangle(pos=(cx + dp(7), y_v + dp(6)), size=(dp(17), dp(3)))
            
            elif "ЭЦН" in self.gno_type:
                Color(0.6, 0.6, 0.65, 1)
                Rectangle(pos=(cx - dp(6), y_v - dp(50)), size=(dp(12), dp(50)))
                Color(0.8, 0.1, 0.1, 1)
                Line(points=[cx + dp(6), y_t, cx + dp(6), y_v - dp(25)], width=dp(1))
                Color(0, 0, 0, 0.5)
                for i in range(5):
                    Line(points=[cx - dp(6), y_v - dp(35 + i * 3), cx + dp(6), y_v - dp(35 + i * 3)], width=dp(1))
            
            else: 
                Color(1, 0.8, 0, 1)
                Line(points=[cx - dp(7), y_v, cx - dp(16), y_v - dp(12)], width=dp(1.5))
                Line(points=[cx + dp(7), y_v, cx + dp(16), y_v - dp(12)], width=dp(1.5))
                
            # --- 7. ЗОНА ПЕРФОРАЦИИ: ТОЧКАМИ ПО СПИРАЛИ ---
            Color(1, 1, 1, 0.6) 
            perf_start_y = y_z + dp(15)
            Ellipse(pos=(cx - dp(20), perf_start_y), size=(dp(3.5), dp(3.5)))
            Ellipse(pos=(cx - dp(10), perf_start_y + dp(10)), size=(dp(3.5), dp(3.5)))
            Ellipse(pos=(cx, perf_start_y + dp(20)), size=(dp(3.5), dp(3.5)))
            Ellipse(pos=(cx + dp(10), perf_start_y + dp(30)), size=(dp(3.5), dp(3.5)))
            Ellipse(pos=(cx + dp(17), perf_start_y + dp(40)), size=(dp(3.5), dp(3.5)))
            Ellipse(pos=(cx - dp(15), perf_start_y + dp(50)), size=(dp(3.5), dp(3.5)))
            Ellipse(pos=(cx - dp(5), perf_start_y + dp(60)), size=(dp(3.5), dp(3.5)))
            Ellipse(pos=(cx + dp(5), perf_start_y + dp(70)), size=(dp(3.5), dp(3.5)))
            Ellipse(pos=(cx + dp(15), perf_start_y + dp(80)), size=(dp(3.5), dp(3.5)))

            # =========================================================================
            # --- 8. ФОНТАННАЯ АРМАТУРА (ЧИСТО СИДИТ НА НАШЕМ РОДНОМ ФОНЕ ПРИЛОЖЕНИЯ) ---
            # =========================================================================
            ground_y = y_t - dp(5)
            Color(0.4, 0.4, 0.45, 1)
            Line(points=[cx - dp(100), ground_y, cx + dp(100), ground_y], width=dp(2))

            Color(0.15, 0.45, 0.75, 1) 
            Rectangle(pos=(cx - dp(32), ground_y), size=(dp(64), dp(8)))
            Rectangle(pos=(cx - dp(16), ground_y + dp(8)), size=(dp(32), dp(16)))
            Rectangle(pos=(cx - dp(52), ground_y + dp(10)), size=(dp(36), dp(12)))
            Rectangle(pos=(cx + dp(16), ground_y + dp(10)), size=(dp(36), dp(12)))
            Rectangle(pos=(cx - dp(14), ground_y + dp(24)), size=(dp(28), dp(16)))
            Rectangle(pos=(cx - dp(14), ground_y + dp(40)), size=(dp(28), dp(20)))
            Rectangle(pos=(cx - dp(56), ground_y + dp(44)), size=(dp(42), dp(12)))
            Rectangle(pos=(cx + dp(14), ground_y + dp(44)), size=(dp(42), dp(12)))
            Rectangle(pos=(cx - dp(12), ground_y + dp(60)), size=(dp(24), dp(14)))

            # Манометры
            Color(1, 1, 1, 1)
            Ellipse(pos=(cx - dp(72), ground_y + dp(43)), size=(dp(14), dp(14))) 
            Ellipse(pos=(cx - dp(7), ground_y + dp(74)), size=(dp(14), dp(14)))  
            Color(0.8, 0.1, 0.1, 1)
            Line(points=[cx - dp(65), ground_y + dp(50), cx - dp(69), ground_y + dp(46)], width=dp(1.5))
            Line(points=[cx, ground_y + dp(74), cx + dp(2), ground_y + dp(84)], width=dp(1.5))

class WellViewScreen(BaseScreen):   # Экран визуализации. «Сцена», на которой запускается анимация закачки раствора.
    def __init__(self, **kw):
        super().__init__(**kw)
        # 1. Создаем основной макет
        self.layout = MDBoxLayout(orientation='vertical')
        
        # 2. Шапка (Твой стандарт)
        self.toolbar = MDBoxLayout(
            orientation='horizontal', size_hint_y=None, height=dp(45),
            md_bg_color=(0.1, 0.4, 0.8, 1), padding=[dp(15), 0, dp(10), 0]
        )
        from kivymd.uix.button import MDIconButton
        btn_back = MDIconButton(
            icon="arrow-left", theme_text_color="Custom", text_color=(1, 1, 1, 1),
            on_release=lambda x: self.go_back()
        )
        self.toolbar.add_widget(btn_back)
        self.toolbar.add_widget(MDLabel(
            text="СИМУЛЯЦИЯ ГЛУШЕНИЯ", theme_text_color="Custom", 
            text_color=(1,1,1,1), font_style="H6"
        ))
        self.layout.add_widget(self.toolbar)

        # 3. ТА САМАЯ ОБЛАСТЬ, где будет рисоваться скважина
        self.draw_area = MDBoxLayout()
        self.layout.add_widget(self.draw_area)
        
        # 4. Кнопка повтора (Зеленая, как на производстве)
        btn_repeat = MDFillRoundFlatButton(
            text="ПОВТОРИТЬ ЗАКАЧКУ", size_hint_x=1, height=dp(50),
            md_bg_color=(0.1, 0.6, 0.2, 1),
            on_release=lambda x: self.on_enter() # Перезапуск метода отрисовки
        )
        self.layout.add_widget(btn_repeat)
        
        self.add_widget(self.layout)

    def on_enter(self):
        """Срабатывает каждый раз при переходе на этот экран"""
        self.draw_area.clear_widgets()
        try:
            # Получаем доступ к данным из основного калькулятора
            calc = self.manager.get_screen('calc')
            
            # Собираем данные из полей ввода
            hz_val = float(calc.f_ckod.text or 2000)
            lh_val = float(calc.f_l_hvost.text or 0)
            
            # Логика глубины воронки/насоса
            if "ЭЦН" in calc.selected_gno.upper():
                # Для ЭЦН глубина — это сумма секций НКТ
                hv_val = float(calc.f_l73.text or 0) + float(calc.f_l60.text or 0)
            else:
                hv_val = float(calc.f_h_v.text or 1500)
            
            # Создаем динамический виджет схемы
            schematic = WellSchematic(
                hz=hz_val, 
                hv=hv_val, 
                gno_type=calc.selected_gno, 
                l_hvost=lh_val, 
                # Если пакер снят или просто воронка — метод считается обраткой/прямой
                method="ОБРАТКА" if ("ВОРОНКА" in calc.selected_gno.upper() or "СНЯТ" in calc.selected_gno.upper()) else "ПРЯМАЯ"
            )
            # Вставляем схему в область отрисовки
            self.draw_area.add_widget(schematic)
        except Exception as e:
            print(f"Ошибка при инициализации симуляции: {e}")

    def on_leave(self):
        """Очищаем память при уходе с экрана"""
        self.draw_area.clear_widgets()

    def go_back(self):
        self.manager.current = 'kill_res'

class LoadingScreen(MDScreen): # Класс экрана заставки (Splash Screen), который видит пользователь при старте приложения.
    def __init__(self, **kw):
        super().__init__(**kw) # Инициализируем базовые параметры экрана KivyMD
        
        # Строгий чистый черный фон без картинок
        with self.canvas.before: # Открываем графический контекст для отрисовки подложки экрана
            Color(0, 0, 0, 1) # Задаем сплошной черный цвет (К=0, З=0, С=0, Альфа=1)
            self.bg_rect = Rectangle(pos=self.pos, size=self.size) # Заливаем прямоугольник во весь экран
            
        self.bind(size=self._update_bg, pos=self._update_bg) # Привязываем обновление черного фона к изменению размеров экрана
        
        layout = MDBoxLayout(orientation='vertical', padding=dp(50)) # Создаем главный вертикальный контейнер с отступами по краям в 50 единиц
        layout.add_widget(MDBoxLayout(size_hint_y=0.45)) # Добавляем пустой блок-распорку сверху, занимающий 45% вертикального пространства
        
        # Тонкая линия загрузки (ниточка)
        self.progress = MDProgressBar(
            value=0, # Начальное положение ползунка загрузки
            max=100, # Максимальное значение шкалы (100%)
            color=(0.1, 0.4, 0.8, 1), # Красивый сине-голубой цвет индикатора
            size_hint_x=0.5, # Ширина полосы занимает ровно половину ширины экрана
            size_hint_y=None, # Отключаем автоматическое масштабирование по высоте
            height=dp(2), # Задаем фиксированную толщину линии в 2 пикселя (эффект ниточки)
            pos_hint={'center_x': 0.5} # Центрируем полосу загрузки горизонтально по оси X
        )
        layout.add_widget(self.progress) # Помещаем индикатор прогресса внутрь главного контейнера
        
        # Твоё название приложения OIL MATE вместо фамилии
        layout.add_widget(MDLabel(
            text="WELL CONTROL", # Отображаемый текст названия твоего инженерного приложения
            halign="center", # Выравнивание текста строго по центру
            font_style="Button", # Используем системный стиль шрифта, предназначенный для кнопок (крупный, четкий)
            theme_text_color="Custom", # Включаем ручную настройку палитры для текста
            text_color=(1, 1, 1, 0.6), # Белый цвет с легкой прозрачностью 60% (приглушенный матовый оттенок)
            size_hint_y=None, # Отключаем авто-высоту для текстового блока
            height=dp(40) # Задаем жесткую высоту текстового поля в 40 единиц
        ))
        
        layout.add_widget(MDBoxLayout(size_hint_y=0.45)) # Добавляем нижнюю пустую распорку на 45%, зажимая элементы по центру
        self.add_widget(layout) # Физически выводим скомпонованный макет на экран смартфона

    def _update_bg(self, instance, value):
        # Метод перерисовывает черный фон во всю ширь, если экран телефона повернулся или изменился
        self.bg_rect.pos = instance.pos
        self.bg_rect.size = instance.size

    def on_enter(self):
        """Автоматическое событие Kivy: срабатывает в момент, когда черный экран загрузки появился перед глазами"""
        self.progress.value = 0 # Жестко обнуляем показатель шкалы перед стартом инициализации
        # Запускаем циклический таймер, вызывающий планировщик каждые 0.02 секунды для плавной прокрутки
        Clock.schedule_interval(self.update_progress, 0.02)

    def update_progress(self, dt):
        """Основной планировщик: плавно двигает линию и дозированно подгружает экраны в фоне"""
        app = MDApp.get_running_app() # Получаем доступ к центральному ядру приложения KrsBookApp
        
        # Двигаем полосу вперед на 2% на каждом тике таймера
        if self.progress.value < 100:
            self.progress.value += 2
            
            # --- СЕКЦИЯ ПОШАГОВОЙ АСИНХРОННОЙ ЗАГРУЗКИ РЕСУРСОВ ---
            
            # Шаг 1 (на 10% прогресса): Безопасно декодируем и кэшируем тяжелый фоновый рисунок вышки
            if self.progress.value == 10:
                if os.path.exists(BACKGROUND_PATH):
                    try:
                        app.global_texture = CoreImage(BACKGROUND_PATH).texture
                    except Exception as e:
                        print(f"Ошибка отложенной загрузки фона: {e}")
                        
            # Шаг 2 (на 30% прогресса): Инициализируем и добавляем в менеджер первую партию базовых экранов
            elif self.progress.value == 30:
                app.sm.add_widget(MainMenuScreen(name='menu'))
                app.sm.add_widget(CalcScreen(name='calc'))
                app.sm.add_widget(KillResultScreen(name='kill_res'))
                app.sm.add_widget(SpoScreen(name='spo'))
                
            # Шаг 3 (на 50% прогресса): Создаем и добавляем вторую партию расчетных справочников
            elif self.progress.value == 50:
                app.sm.add_widget(VzdScreen(name='vzd'))
                app.sm.add_widget(ThreadsScreen(name='threads'))
                app.sm.add_widget(CementAcidScreen(name='cmoist'))
                app.sm.add_widget(WellViewScreen(name='well_view'))
                app.sm.add_widget(FishingScreen(name='fishing'))
                app.sm.add_widget(CalculatorHubScreen(name='calc_hub'))
                app.sm.add_widget(SettingsScreen(name='settings'))
                
            # Шаг 4 (на 70% прогресса): Финишируем — создаем и регистрируем экраны Блокнота и Справочника персонала КРС
            elif self.progress.value == 70:
                app.sm.add_widget(NotebookMainMenu(name='notebook_main_menu'))
                app.sm.add_widget(CreateWellScreen(name='create_well_screen'))
                app.sm.add_widget(OpenWellScreen(name='open_well_screen'))
                app.sm.add_widget(ViewWellScreen(name='view_well_screen'))
                app.sm.add_widget(EditWellScreen(name='edit_well_screen'))
                
                # ТЕХНОЛОГИЧЕСКИЙ ФИКС: Добавляем новые полноценные экраны Справочника персонала
                app.sm.add_widget(OpenBrigadesScreen(name='open_brigades_screen'))
                app.sm.add_widget(CreateBrigadeScreen(name='create_brigade_screen'))
                
            return True # ИСПРАВЛЕНО: Строка возврата теперь строго тут, таймер будет крутиться до 100% без остановок!
            
        else:
            # --- ВЫХОД ИЗ КАТАЛОГА ЗАГРУЗКИ ---
            self.manager.current = 'menu' # Переключаем экран на полностью готовое главное меню
            # Запускаем показ окна дисклеймера через 0.3 секунды после открытия меню
            Clock.schedule_once(lambda x: app.show_disclaimer(), 0.3)
            return False # Полностью останавливаем и уничтожаем этот фоновый таймер


class CalcScreen(BaseScreen): # Экран технологических расчетов. Наследует BaseScreen, получая его адаптивный фон и фикс клавиатуры.
    def __init__(self, **kw):
        super().__init__(**kw) # Инициализируем базовые параметры экрана Kivy
        
        # --- ТЕХНОЛОГИЧЕСКИЕ ПЕРЕМЕННЫЕ И НАСТРОЙКИ ПО УМОЛЧАНИЮ ---
        self.selected_gno = "Воронка" # Тип глубинно-насосного оборудования на конце колонны (компоновка низа)
        self.tech_nkt_name = "" # Текстовое имя выбранной НКТ для вывода на экран или в отчеты
        self.v_pipe = None # Внутренний удельный объем трубы для расчетов объема глушения (измеряется в литрах на метр)
        
        # Параметры насосного агрегата ЦА-320 по умолчанию: диаметр втулки (мм), передача КПП, обороты двигателя в минуту
        self.vtulka, self.gear, self.rpm = "115", "3", 1000 
        
        # Категория опасности (по умолчанию 3-я) с критериями газоопасности Ростехнадзора
        self.well_category = "3 КАТЕГОРИЯ" # Начальный уровень газоопасности скважины при КРС
        
        # Метод глушения по умолчанию и словарь человеческих названий
        self.kill_method = "STAND" # Внутренний технический идентификатор выбранной технологии глушения
        self.method_names = { # Словарь-маппинг для перевода кодов в понятные инженерам названия на экране
            "STAND": "ПЛАНОВОЕ ГЛУШЕНИЕ",
            "DRILLER": "МЕТОД БУРИЛЬЩИКА (2 ЦИКЛА)",
            "WAIT": "ОЖИДАНИЕ И УТЯЖЕЛЕНИЕ (1 ЦИКЛ)",
            "BULL": "ЗАДАВКА В ПЛАСТ "
        }
        
        # Коэффициенты подачи насоса агрегата ЦА-320 (в литрах за один ход поршня) 
        # в зависимости от диаметра втулки (100, 115, 127 мм) и включенной передачи (со 2-й по 5-ю)
        self.q_coeffs = { 
            "100": {"2": 0.00171, "3": 0.00342, "4": 0.00513, "5": 0.00777},
            "115": {"2": 0.00227, "3": 0.00453, "4": 0.00680, "5": 0.01020},
            "127": {"2": 0.00277, "3": 0.00555, "4": 0.00832, "5": 0.01247}
        }
        
        # --- ВЕРСТКА ИНТЕРФЕЙСА ЭКРАНА ---
        # Главный вертикальный контейнер, занимающий 100% ширины и высоты доступного пространства экрана
        main_layout = MDBoxLayout(orientation='vertical', size_hint=(1, 1)) 
        
        # 1. ШАПКА (Верхняя панель / Тулбар)
        self.toolbar = MDBoxLayout(
            orientation='horizontal', # Элементы внутри тулбара идут слева направо
            size_hint_y=None, # Отключаем автоматическую высоту
            height=dp(45), # Задаем фиксированную высоту панели в 45 адаптивных пикселей
            md_bg_color=(0.1, 0.4, 0.8, 1), # Фирменный сине-голубой цвет шапки WELL CONTROL
            padding=[dp(15), 0, dp(10), 0] # Отступы внутри панели: слева 15dp, справа 10dp
        )
        
        # Кнопка «Назад» (стрелочка) на верхней панели
        btn_back = MDIconButton(
            icon="arrow-left", # Имя встроенной системной иконки Material Design
            pos_hint={'center_y': .5}, # Центрируем кнопку строго посередине высоты тулбара
            theme_text_color="Custom", # Разрешаем ручной выбор цвета для иконки
            text_color=(1, 1, 1, 1), # Белый цвет стрелочки
            on_release=lambda x: self.go_back() # При нажатии вызываем метод возврата на предыдущий экран
        )
        self.toolbar.add_widget(btn_back) # Помещаем кнопку назад в левую часть тулбара
        
        # Заголовок текущего расчетного экрана на тулбаре
        self.toolbar.add_widget(MDLabel(
            text="ВЫБОР ТЕХНОЛОГИИ ГЛУШЕНИЯ", # Текст заголовка
            theme_text_color="Custom", # Разрешаем свой цвет
            text_color=(1,1,1,1), # Чистый белый цвет текста
            font_style="H6" # Крупный и четкий стиль шрифта под заголовки
        ))
        main_layout.add_widget(self.toolbar) # Добавляем готовую шапку в самый верх главного макета
        
        # Зона прокрутки (скролл) для вмещения всех полей ввода параметров скважины
        scroll = ScrollView() 
        # Вертикальный контейнер внутри скролла, автоматически растягивающийся по высоте под количество элементов
        self.container = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(10), padding=dp(20))
        # 2. КНОПКА КАТЕГОРИИ ОПАСНОСТИ (ЯРКО-КРАСНАЯ) 
        # Подпись-подсказка над кнопкой категории опасности
        self.container.add_widget(MDLabel(text="КАТЕГОРИЯ СКВАЖИНЫ ПО ПЛАНУ РАБОТ:", font_style="Caption", theme_text_color="Hint")) 
        
        # Специальный контейнер для центрирования и жесткого контроля размеров кнопки
        anchor_cat = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(45)) 
        
        # Сама кнопка выбора категории газоопасности скважины
        self.btn_category = MDFillRoundFlatButton( 
            text=f"ОПАСНОСТЬ: {self.well_category}", # Выводит текущую категорию
            size_hint=(1, 1), # Занимает все пространство внутри своего анкора
            md_bg_color=(0.85, 0.15, 0.15, 1), # Ярко-красный сигнальный цвет (внимание, опасность)
            on_release=self.menu_category # При нажатии открывает выпадающее меню выбора категорий
        ) 
        self.btn_category.size_hint_max_x = dp(2000) # Ограничитель максимальной ширины кнопки для широких экранов
        anchor_cat.add_widget(self.btn_category) # Кладем кнопку в центровщик
        self.container.add_widget(anchor_cat) # Добавляем этот блок в вертикальный список расчетов

        # 3. РЯД КНОПОК ГНО И НКТ 
        # Горизонтальный контейнер для верхнего ряда кнопок выбора компоновок
        self.top_btns_row = MDBoxLayout(orientation='horizontal', spacing=dp(5), size_hint_y=None, height=dp(45), size_hint_x=1) 
        
        # Кнопка для выбора типа глубинно-насосного оборудования (Воронка, ЭЦН и т.д.)
        self.btn_gno = MDFillRoundFlatButton(text=f"ГНО: {self.selected_gno.upper()}", size_hint=(1, 1), md_bg_color=(0.1, 0.3, 0.6, 1), on_release=self.menu_gno) 
        
        # Обертка-контейнер для кнопки выбора НКТ, позволяющая динамически скрывать/показывать её
        self.nkt_box_wrap = MDBoxLayout(size_hint=(1, 1)) 
        # Сама кнопка выбора труб НКТ из базы данных
        self.btn_nkt = MDFillRoundFlatButton(text="ВЫБРАТЬ НКТ", size_hint=(1, 1), md_bg_color=(0.1, 0.3, 0.6, 1), on_release=self.open_nkt_menu) 
        
        self.nkt_box_wrap.add_widget(self.btn_nkt) # Вкладываем кнопку НКТ в её обертку
        self.top_btns_row.add_widget(self.btn_gno) # Добавляем кнопку ГНО в общий горизонтальный ряд
        self.top_btns_row.add_widget(self.nkt_box_wrap) # Добавляем обертку кнопки НКТ в горизонтальный ряд
        self.container.add_widget(self.top_btns_row) # Помещаем получившийся ряд кнопок в главный вертикальный список

        # 4. ПОЛЯ ВВОДА (ГЕОМЕТРИЯ) 
        # Создаем компактные текстовые поля для ввода геометрических параметров скважины со значениями по умолчанию
        self.f_d_col = self.create_small_field("Ø экспл. колонны (мм)", "168") # Диаметр эксплуатационной колонны
        self.f_ckod = self.create_small_field("ЦКОД / Иск. забой (м)", "3200") # Глубина искусственного забоя (цементного кольца)
        self.f_l_hvost = self.create_small_field("Длина хвостовика 114 (м)", "0") # Длина хвостовика, если он имеется (маппинг-индекс 9)
        self.f_h_tvd = self.create_small_field("Глубина по вертикали (м)", "2200") # Истинная глубина скважины по вертикали (для гидростатики)
        self.f_h_perf = self.create_small_field("Интервал перфорации (кровля, м)", "2150") # Верхняя граница вскрытого пласта
        
        # Выравниваем отступы: строка добавления должна быть ВНУТРИ цикла (сдвинута вправо)
        # Циклически добавляем все созданные геометрические поля на экран Журнала расчетов
        for f in [self.f_d_col, self.f_ckod, self.f_l_hvost, self.f_h_perf, self.f_h_tvd]: 
            self.container.add_widget(f) 
            
        # Динамический блок для полей ЭЦН или Воронки 
        # Контейнер, содержимое которого меняется на лету в зависимости от того, выбран ли тип компоновки ЭЦН или обычная Воронка
        self.gno_fields_box = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(10)) 
        self.container.add_widget(self.gno_fields_box) 

        # 5. ДАННЫЕ ДАВЛЕНИЙ 
        # Метка подзаголовка для блока параметров давлений и плотностей технологической жидкости
        self.container.add_widget(MDLabel(text="ПАРАМЕТРЫ ДАВЛЕНИЙ И РАСТВОРА:", font_style="Caption", theme_text_color="Hint")) 
        
        self.f_p_pl = self.create_small_field("P пласт (атм) — ЕСЛИ ИЗВЕСТНО", "") # Текущее пластовое давление
        self.f_p_ust = self.create_small_field("P трубное (атм)", "0") # Давление на устье в трубном пространстве 
        self.f_p_zatr = self.create_small_field("P затрубное (атм)", "0") # Давление в затрубном пространстве скважины
        self.f_rho = self.create_small_field("Текущая плотность раствора в скв. (г/см³)", "1.10") # Удельный вес глушильной жидкости в стволе
        
        # Циклически выводим поля параметров давлений на экран в порядке очереди
        for f in [self.f_p_pl, self.f_p_ust, self.f_p_zatr, self.f_rho]: 
            self.container.add_widget(f) 
        # 6. ВЫБОР МЕТОДА ГЛУШЕНИЯ 
        # Подпись над выбором технологического регламента глушения скважины
        self.container.add_widget(MDLabel(text="МЕТОД ГЛУШЕНИЯ:", font_style="Caption", theme_text_color="Hint")) 
        
        # Разметка с привязкой по центру для кнопки выбора метода глушения
        anchor_method = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(50)) 
        self.btn_method = MDFillRoundFlatButton( 
            text="", # Текст установится автоматически функцией refresh_ui_elements
            size_hint=(1, 1), 
            md_bg_color=(0.1, 0.3, 0.6, 1), 
            on_release=self.menu_kill_methods # По нажатию выпадает список (Плановый, Бурильщика, Задавка и т.д.)
        ) 
        self.btn_method.size_hint_max_x = dp(2000) # Ограничитель ширины кнопки
        anchor_method.add_widget(self.btn_method) 
        self.container.add_widget(anchor_method) 

        # 7. СЕКЦИЯ АГРЕГАТА ЦА-320 
        # Заголовок настроек режима работы промывочного насосного агрегата
        self.container.add_widget(MDLabel(text="РЕЖИМ РАБОТЫ ЦА-320", font_style="Subtitle2", bold=True, theme_text_color="Secondary", halign="center")) 
        
        # Сетка из 3 колонок для компактного размещения параметров ЦА-320 (Втулка, Передача, Обороты)
        pump_grid = MDGridLayout(cols=3, spacing=dp(5), size_hint_y=None, height=dp(45), size_hint_x=1) 
        
        def create_p_btn(text, func): 
            # Внутренняя функция для быстрой генерации кнопок управления параметрами насоса агрегата
            wrap = MDBoxLayout(size_hint_x=1) 
            btn = MDFillRoundFlatButton(text=text, size_hint=(1,1), md_bg_color=(0.1, 0.4, 0.8, 1), on_release=func) 
            wrap.add_widget(btn) 
            return btn, wrap 
            
        # Создаем 3 управляющие кнопки на основе технологической матрицы Q_coeffs
        self.btn_v, w1 = create_p_btn(f"ВТ:{self.vtulka}", self.menu_v) # Диаметр поршневой втулки
        self.btn_g, w2 = create_p_btn(f"СКОР:{self.gear}", self.menu_g) # Номер передачи КПП агрегата
        self.btn_r, w3 = create_p_btn(f"ОБ:{self.rpm}", self.menu_r) # Скорость вращения вала (об/мин)
        
        # Добавляем обертки кнопок в созданную трехколоночную сетку
        for w in [w1, w2, w3]: 
            pump_grid.add_widget(w) 
        self.container.add_widget(pump_grid) 

        # 8. КНОПКА РАСЧЕТА / ИНДИКАТОР ПРЕДУПРЕЖДЕНИЯ (Ширина привязана к якорю) 
        # Якорный контейнер под финишную кнопку запуска вычислений параметров глушения
        self.calc_anchor = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(60)) 
        self.container.add_widget(self.calc_anchor) 
        
        scroll.add_widget(self.container) # Вкладываем заполненный контейнер со всеми полями внутрь зоны прокрутки
        main_layout.add_widget(scroll) # Вкладываем скролл под верхний тулбар
        self.add_widget(main_layout) # Выводим весь сформированный макет на экран
        
        self.refresh_ui_elements() # Первичный вызов синхронизации интерфейса для заполнения текстов на кнопках

    def create_small_field(self, hint, val=""): 
        # Вспомогательный метод для генерации однотипных компактных текстовых полей Material Design
        f = MDTextField(
            hint_text=hint, # Текст подсказки/названия параметра внутри поля
            text=val, # Предустановленное дефолтное значение
            mode="fill", # Стиль поля со сплошным фоном-заливкой
            font_size="15sp", # Размер шрифта, адаптированный под системный масштаб
            size_hint_y=None, 
            height=dp(34), # Компактная фиксированная высота поля ввода
            input_filter='float' # Блокировка клавиатуры: разрешен ввод только цифр и точек (дробные числа)
        )
        # При получении фокуса (нажатии на поле) автоматически выделяем весь текст для быстрой перезаписи за 0.1 секунды
        f.bind(focus=lambda ins, val: Clock.schedule_once(lambda dt: ins.select_all(), 0.1) if val else None) 
        return f 

    def refresh_ui_elements(self): 
        """Ключевой метод динамического обновления интерфейса при изменении выбранных параметров"""
        self.gno_fields_box.clear_widgets() # Полностью очищаем блок динамических полей ГНО перед перерисовкой
        
        # Синхронизация текста кнопки выбора НКТ с учетом выбранного типоразмера и объема
        if self.tech_nkt_name: 
            self.btn_nkt.text = f"НКТ {self.tech_nkt_name} ({self.v_pipe:.2f} л/м)" 
        else: 
            self.btn_nkt.text = "ВЫБРАТЬ НКТ" 
            
        self.btn_gno.text = f"ГНО: {self.selected_gno.upper()}" # Обновляем текст на кнопке ГНО
        
        # --- ТЕХНОЛОГИЧЕСКАЯ АВТОМАТИЗАЦИЯ ДЛЯ ПАКЕРА И МЕТОДОВ ГЛУШЕНИЯ ---
        if "ПАКЕР (УСТ.)" in self.selected_gno.upper():
            # Если пакер установлен, циркуляция невозможна. Принудительно включаем метод Задавки в пласт (BULL)
            self.kill_method = "BULL"
            # Блокируем кнопку выбора методов глушения, чтобы избежать ошибочных циркуляционных расчетов
            if hasattr(self, 'btn_method') and self.btn_method:
                self.btn_method.disabled = True
        else:
            # Для остальных типов ГНО (Воронка, ЭЦН, Снятый пакер) разблокируем кнопку выбора методов
            if hasattr(self, 'btn_method') and self.btn_method:
                self.btn_method.disabled = False

        self.btn_method.text = self.method_names[self.kill_method] # Обновляем название выбранного метода глушения из словаря
        
        # Логика изменения интерфейса: если выбрана компоновка ЭЦН (электроцентробежный насос)
        if "ЭЦН" in self.selected_gno.upper(): 
            # Прячем и отключаем кнопку выбора НКТ, так как для ЭЦН идет комбинированная стандартная подвеска
            self.nkt_box_wrap.size_hint_x, self.nkt_box_wrap.opacity, self.nkt_box_wrap.disabled = 0, 0, True 
            self.btn_gno.size_hint_x = 1 
            # Генерируем два специфических поля ввода для глубин спуска ЭЦН (длины секций труб 73 и 60 мм)
            self.f_l73 = self.create_small_field("Длина НКТ 73 (м)", "1000") 
            self.f_l60 = self.create_small_field("Длина НКТ 60 (м)", "500") 
            self.gno_fields_box.add_widget(self.f_l73) 
            self.gno_fields_box.add_widget(self.f_l60) 
        else: 
            # Если выбрана обычная Воронка или иные пакерные компоновки
            # Возвращаем видимость кнопке выбора НКТ на экране
            self.nkt_box_wrap.size_hint_x, self.nkt_box_wrap.opacity, self.nkt_box_wrap.disabled = 1, 1, False 
            self.btn_gno.size_hint_x = 1 
            # Генерируем одно стандартное поле ввода глубины спуска воронки НКТ или пакера
            self.f_h_v = self.create_small_field("Глубина подвески ГНО / Воронки (м)", "2480") 
            self.gno_fields_box.add_widget(self.f_h_v) 

        # ФИКС КРАША ПРИ СОХРАНЕНИИ: Сюда передаются ТОЛЬКО валидные свойства MDCard 
        # Полностью очищаем якорный контейнер перед перерисовкой нижней панели (кнопка или предупреждение)
        self.calc_anchor.clear_widgets() 
        
        # Условие проверки: если это НЕ ЭЦН (требуется ручной выбор труб) и при этом пользователь еще НЕ выбрал НКТ из базы
        if "ЭЦН" not in self.selected_gno.upper() and self.v_pipe is None: 
            # Карточка БЕЗ аргумента text. Создаем предупреждающую оранжевую плашку на базе MDCard
            attention_table = MDCard( 
                orientation="vertical", padding=[dp(15), dp(12)], size_hint=(1, 1), radius=[dp(10)],
                md_bg_color=(0.85, 0.4, 0.0, 1), ripple_behavior=False
            ) 
            attention_table.size_hint_max_x = dp(2000) 
            
            # Текст строго внутри MDLabel
            attention_text = MDLabel( 
                text="[b]ВНИМАНИЕ![/b] СНАЧАЛА ВЫБЕРИТЕ ТИП НКТ ИЗ БАЗЫ EXCEL!", 
                halign="center", valign="middle", markup=True, theme_text_color="Custom",
                text_color=(1, 1, 1, 1), font_style="Button" 
            ) 
            attention_table.add_widget(attention_text) 
            self.calc_anchor.add_widget(attention_table) 
        else: 
            # Если все условия соблюдены (или выбран ЭЦН, или НКТ успешно загружен из Excel)
            # Создаем большую зеленую кнопку для запуска математического ядра расчетов
            self.btn_calc = MDFillRoundFlatIconButton( 
                text="ВЫПОЛНИТЬ РАСЧЕТ ТЕХНОЛОГИИ", icon="calculator", size_hint=(1, 1),
                md_bg_color=(0.1, 0.6, 0.2, 1), on_release=self.do_calc 
            ) 
            self.btn_calc.size_hint_max_x = dp(2000) 
            self.calc_anchor.add_widget(self.btn_calc) 

    def do_calc(self, *args): 
        """Основное математическое ядро. Считает объемы глушения, плотности, гидростатику и время закачки"""
        # Дополнительная страховка безопасности: блокируем расчет, если трубы не выбраны
        if "ЭЦН" not in self.selected_gno.upper() and self.v_pipe is None: 
            return 
            
        try: 
            # Внутренняя мини-функция для безопасного извлечения и конвертации текста из полей ввода в дробные числа float
            def val(f): 
                return float(f.text) if (hasattr(f, 'text') and f.text) else 0.0 
                
            # Считываем и переводим в числа геометрические параметры скважины
            D, L_ckod, L_hvost, H_tvd = val(self.f_d_col), val(self.f_ckod), val(self.f_l_hvost), val(self.f_h_tvd) 
            H_perf = val(self.f_h_perf) # Кровля перфорации
            
            # Считываем параметры давлений и текущей плотности глушильного раствора
            p_pl, p_ust, p_zatr, rho_c = val(self.f_p_pl), val(self.f_p_ust), val(self.f_p_zatr), val(self.f_rho) 
            
            # Вычисляем теоретическую производительность насоса ЦА-320 (л/мин) 
            # Умножаем коэффициент объема одного хода (из матрицы) на обороты в минуту и на 60 минут
            Q = self.q_coeffs[self.vtulka][self.gear] * self.rpm * 60 
            
            # Давление и гидростатика рассчитываются по зоне перфорации. Выбираем опорную глубину (h_ref)
            # Приоритет: Кровля перфорации -> Глубина по вертикали -> Искусственный забой
            h_ref = H_perf if H_perf > 0 else (H_tvd if H_tvd > 0 else L_ckod) 
            if h_ref == 0: 
                return # Защита от деления на ноль, если все ключевые глубины скважины равны нулю
                
            # Авторасчет пластового давления, если поле оставлено пустым (инженер не заполнил P_pl)
            if p_pl == 0: 
                p_izb_calc = p_ust if p_ust > 0 else p_zatr # Берем наибольшее избыточное давление на устье (трубное или затрубное)
                # Формула: Гидростатика текущего столба жидкости ((Глубина * Плотность) / 10) + Избыточное давление на устье
                p_pl = round(((h_ref * rho_c) / 10) + p_izb_calc, 1) 
                calculated_p_pl_flag = True # Ставим внутренний флаг, что пластовое давление было рассчитано автоматически
            else: 
                calculated_p_pl_flag = False # Пластовое давление введено инженером вручную
                
            # Выбор коэффициента запаса Кз согласно требованиям Ростехнадзора и регламентам ПАО "Сургутнефтегаз"
            if "1 КАТЕГОРИЯ" in self.well_category: 
                k_zapas = 1.10 # Для 1-й категории (газовые, нефтяные с ГФ > 200) запас 10%
            elif "2 КАТЕГОРИЯ" in self.well_category: 
                k_zapas = 1.07 # Для 2-й категории запас 7%
            else: 
                k_zapas = 1.05 # Для 3-й категории (депрессия менее 10%) запас 5%
                
            # Расчет необходимой плотности нового раствора для глушения скважины (rho)
            if self.kill_method == "STAND": 
                # Плановый режим (заглушить гидростатикой с запасом): Плотность = (P_пластовое * Коэффициент перевода) / Глубину * К_запаса
                rho = (p_pl * 10.197 / h_ref * k_zapas) 
            else: 
                # Другие методы (Бурильщика и т.д.): берем избыток на устье
                p_izb = p_ust if p_ust > 0 else p_zatr 
                # Наращиваем плотность на величину устьевого давления плюс технологический избыток в 0.05 г/см³
                rho = round(rho_c + (p_izb * 10.197 / h_ref) + 0.05, 2) 
                
            # Расчет внутренних удельных объемов пространств скважины (литры на погонный метр)
            # Формула площади круга через диаметр: (D - толщина_стенок)^2 * 0.0007854
            v_col_m = ((D - 15)**2) * 0.0007854 # Удельный объем эксплуатационной колонны
            v_hvost_m = ((114 - 15)**2) * 0.0007854 # Удельный объем хвостовика диаметром 114 мм
            
            # Динамический расчет внутреннего объема труб в зависимости от выбранного типа ГНО
            if "ЭЦН" in self.selected_gno.upper(): 
                l73, l60 = val(self.f_l73), val(self.f_l60) # Считываем длины комбинированной подвески труб
                Hv = l73 + l60 # Общая глубина спуска насоса ЭЦН равна сумме длин труб
                # Суммируем внутренние объемы: 3.02 л/м для трубы 73 мм и 1.99 л/м для трубы 60 мм, переводим в кубические метры
                v_trub = ((3.02 * l73) + (1.99 * l60)) / 1000 
            else: 
                Hv = val(self.f_h_v) # Глубина спуска воронки НКТ
                v_trub = (self.v_pipe * Hv) / 1000 # Внутренний объем одиночной колонны НКТ (куб. метры)
                
            # Расчет объема затрубного пространства до глубины подвески ГНО (с учетом коэффициента вытеснения металла труб ~1.35)
            v_zatrub = (v_col_m * Hv / 1000) - (v_trub * 1.35) 
            
            # Расчет объема подпакерного (подвороночного) пространства скважины строго до искусственного забоя L_ckod
            # Учитывает, есть ли внизу сужение в виде хвостовика 114 мм (переменная L_hvost)
            v_pod = (v_col_m * ((L_ckod - L_hvost) - Hv) / 1000) + (v_hvost_m * L_hvost / 1000) if L_hvost > 0 else (v_col_m * (L_ckod - Hv) / 1000) 
            
            # Логика сборки полного объема скважины на основе текущего состояния ГНО (установлено или снимается)
            gno_low = self.selected_gno.lower() 
            if "уст." in gno_low: 
                v_total = v_trub # Если ГНО только устанавливается, заполняем внутренний объем труб
            elif "снят" in gno_low: 
                v_total = v_zatrub + v_pod # Если ГНО извлекается, считаем только внешнее кольцо ствола и поднасосную зону
            else: 
                v_total = v_trub + v_zatrub + v_pod # Стандартная полная схема: трубы + затруб + подпакерное пространство
                
            # Сургутский нормативный аварийный избыток. По регламенту КРС к объему скважины жестко прибавляется +3.0 м³ на поглощение
            v_total_with_zap = round(v_total + 3.0, 2) 
            
            # Вычисляем чистое расчетное время полной закачки раствора технологическим агрегатом в минутах
            t_min = (v_total_with_zap * 1000) / Q if Q > 0 else 0
            # --- УЧЕБНО-ТЕХНОЛОГИЧЕСКИЙ РЕГЛАМЕНТ ГЛУШЕНИЯ СКВАЖИН (ИНСТРУКЦИИ КРС) ---
            
            # Вспомогательный блок: если установлена компоновка ЭЦН, принудительно внедряем обязательный ЭТАП 0
            ecn_prefix = ""
            if "ЭЦН" in self.selected_gno.upper():
                ecn_prefix = (
                    f"[color=#FFCC00][b]ЭТАП 0: СБИТИЕ СЛИВНОГО КЛАПАНА ЭЦН (ОБЯЗАТЕЛЬНО)[/b][/color]\n"
                    f"• Перед началом любых технологических операций глушения сбросьте металлический ломик (бросавку) в трубное пространство НКТ.\n"
                    f"• Произойдет разрушение срезного байпасного клапана, что обеспечит сообщение между НКТ и затрубным пространством выше насоса.\n"
                    f"• Критерий успешности: Падение уровня жидкости в НКТ и выравнивание устьевых давлений. Без сбития клапана циркуляция невозможна!\n\n"
                )

            # ВЕТВЛЕНИЕ ПО МЕТОДАМ И ТИПАМ ГЛУБИННОГО ОБОРУДОВАНИЯ (ПАКЕРЫ / ГНО)
            if self.kill_method == "STAND":
                # Метод 1: ПЛАНОВОЕ ГЛУШЕНИЕ (ШТАТНЫЙ РЕЖИМ ЗАМЕЩЕНИЯ)
                if "ПАКЕР (СНЯТ)" in self.selected_gno.upper():
                    # Специфика для сорванного пакера — глушение ОБРАТНОЙ циркуляцией через затруб
                    instr = ecn_prefix + (
                        f"[color=#33FF33][b]ПЛАНОВОЕ ОБРАТНОЕ ГЛУШЕНИЕ (ПАКЕР СНЯТ)[/b][/color]\n"
                        f"• [b]Суть метода:[/b] Замещение скважинной жидкости новым раствором глушения плотностью [b]{rho:.2f} г/см³[/b] в объеме [b]{v_total_with_zap:.2f} м³[/b].\n"
                        f"• [b]Технология закачки:[/b] Подачу технологической жидкости от агрегата ЦА-320 осуществлять строго в [color=#FFFF33]ЗАТРУБНОЕ ПРОСТРАНСТВО[/color].\n"
                        f"• Вытесняемый флюид и старый раствор плотностью [b]{rho_c:.2f} г/см³[/b] выходят на поверхность через трубное пространство (НКТ) в мерную емкость.\n"
                        f"• [b]Контроль параметров:[/b] Каждые 15 минут контролируйте объемы долива и затруба по мерникам ЦС. При разнице затруба и долива > 0.5 м³ — немедленно СТОП СПО и герметизация устья превентором!\n"
                        f"• [b]Критерий завершения:[/b] Появление чистого утяжеленного раствора из НКТ, выравнивание плотности на входе и выходе. Расчетное время циркуляции: [b]{int(t_min)} мин[/b]."
                    )
                else:
                    # Стандартная прямая циркуляция для Воронки или ЭЦН со сбитым клапаном
                    gno_text = "через НКТ" if "эцн" not in self.selected_gno.lower() else "в затруб (клапан сбит)"
                    instr = ecn_prefix + (
                        f"[color=#33FF33][b]ПЛАНОВОЕ ГЛУШЕНИЕ СКВАЖИНЫ ПРЯМОЙ ЦИРКУЛЯЦИЕЙ[/b][/color]\n"
                        f"• [b]Суть метода:[/b] Плановое технологическое вытеснение объема скважины утяжеленным раствором плотностью [b]{rho:.2f} г/см³[/b]. Общий проектный объем: [b]{v_total_with_zap:.2f} м³[/b].\n"
                        f"• [b]Технология закачки:[/b] Закачка ведется {gno_low}. Обеспечьте плавный пуск насоса ЦА-320 на {self.gear}-й передаче при {self.rpm} об/мин.\n"
                        f"• Контролируйте давление на манометре насосного агрегата. Оно не должно превышать давление опрессовки эксплуатационной колонны.\n"
                        f"• [b]Правило промышленной безопасности:[/b] Ведите непрерывный учет баланса долива. При обнаружении поглощения технологической жидкости снизьте подачу насоса.\n"
                        f"• Расчетное время полной промывки ствола составляет: [b]{int(t_min)} мин[/b]."
                    )

            elif self.kill_method == "DRILLER":
                # Метод 2: МЕТОД БУРИЛЬЩИКА (2 ЦИКЛА ЦИРКУЛЯЦИИ ПО IWCF)
                instr = ecn_prefix + (
                    f"[color=#FFCC00][b]МЕТОД БУРИЛЬЩИКА — ДВУХЦИКЛОВОЕ ЛИКВИДИРОВАНИЕ ГНВП[/b][/color]\n\n"
                    f"[color=#33FF33][b]ПЕРВЫЙ ЦИКЛ: ВЫМЫВАНИЕ ГАЗОВОЙ ПАЧКИ (ПРИТОКА)[/b][/color]\n"
                    f"• Цель: Полное удаление пластового флюида из ствола скважины с сохранением старого раствора плотностью [b]{rho_c:.2f} г/см³[/b].\n"
                    f"• Запустите ЦА-320 и начните вымыв пачки флюида на поверхность через дроссель.\n"
                    f"• Регулированием штуцера поддерживайте давление в бурильных трубах (НКТ) ПОСТОЯННЫМ. Пачка пластового газа покажется на устье примерно через [b]{v_trub:.2f} м³[/b] закачки.\n\n"
                    f"[color=#33FF33][b]ВТОРОЙ ЦИКЛ: ЗАКАЧКА ТЯЖЕЛОГО РАСТВОРА ГЛУШЕНИЯ[/b][/color]\n"
                    f"• Цель: Закачка проектного утяжеленного раствора плотностью [color=#FF5555][b]{rho:.2f} г/см³[/b][/color] для уравновешивания пластового давления.\n"
                    f"• По мере продвижения тяжелого раствора от устья к забою, плавно СНИЖАЙТЕ давление в НКТ согласно графику глушения.\n"
                    f"• Как только тяжелый раствор дойдет до забоя, удерживайте давление затруба постоянным до выхода раствора на поверхность.\n\n"
                    f"[color=#888888]----------------------------------------\n"
                    f"[b]ПЛЮСЫ:[/b] Вымыв притока можно начинать немедленно, не дожидаясь заготовки утяжеленного раствора.\n"
                    f"[b]МИНУСЫ:[/b] Требует минимум двух полных циклов циркуляции. Создает повышенные давления на устье и стенки эксплуатационной колонны.[/color]"
                )

            elif self.kill_method == "WAIT":
                # Метод 3: МЕТОД ОЖИДАНИЯ И УТЯЖЕЛЕНИЯ (1 ЦИКЛ ЦИРКУЛЯЦИИ ПО IWCF)
                instr = ecn_prefix + (
                    f"[color=#33FF33][b]МЕТОД ОЖИДАНИЯ И УТЯЖЕЛЕНИЯ (WAIT & WEIGHT)[/b][/color]\n"
                    f"• [b]Основная суть:[/b] Герметизация устья превенторами, расчет и приготовление утяжеленного раствора, а затем вымывание газового притока и одновременное глушение за ОДИН полный цикл циркуляции.\n\n"
                    f"[b]ЭТАПЫ ПРОВЕДЕНИЯ СОГЛАСНО РЕГЛАМЕНТУ IWCF:[/b]\n"
                    f"1. [b]Ожидание и приготовление:[/b] Скважина находится в закрытом состоянии под давлением. В это время на буровой готовится утяжеленный раствор проектной плотности [b]{rho:.2f} г/см³[/b] в объеме [b]{v_total_with_zap:.2f} м³[/b].\n"
                    f"2. [b]Закачка и вымывание за 1 цикл:[/b] Тяжелый раствор закачивается в скважину. Давление в НКТ плавно снижается от начального до конечного давления циркуляции строго по графику по мере заполнения труб.\n"
                    f"3. [b]Завершение операции:[/b] ФЛЮИД полностью вымывается через манифольд дросселирования, а затрубное пространство заполняется тяжелой жидкостью, полностью блокируя пласт.\n\n"
                    f"[color=#888888][b]ПЛЮСЫ:[/b] Скважина глушится всего за один цикл циркуляции. Давление на устье и обсадную колонну снижается до минимума.\n"
                    f"[b]МИНУСЫ:[/b] Требуется длительное время простоя скважины под давлением на период заготовки раствора.[/color]"
                )

            else:
                # Метод 4: ЗАДАВКА В ПЛАСТ (BULLHEADING — ОБЯЗАТЕЛЬНО ПРИ УСТАНОВЛЕННОМ ПАКЕРЕ)
                if "ПАКЕР (УСТ.)" in self.selected_gno.upper():
                    instr = (
                        f"[color=#FF5555][b]ТЕХНОЛОГИЧЕСКАЯ КАРТА ЗАДАВКИ В ПЛАСТ (ПАКЕР УСТАНОВЛЕН)[/b][/color]\n"
                        f"• [b]Обоснование метода:[/b] Так как пакер установлен и затрубное пространство изолировано, циркуляция невозможна. Глушение проводится методом продавливания флюида обратно в пласт.\n"
                        f"• [b]Технология ведения работ:[/b] Подключите манифольд агрегата ЦА-320 к трубному пространству (НКТ). Начните прямую закачку раствора плотностью [color=#FFCC00][b]{rho:.2f} г/см³[/b][/color].\n"
                        f"• [b]Критическое требование безопасности:[/b] Непрерывно контролируйте давление на устье! Давление закачки ЦА-320 НЕ ДОЛЖНО превышать давление гидроразрыва пласта (Р_гр) и предел прочности обсадной колонны на смятие.\n"
                        f"• Качайте проектный объем [b]{v_total_with_zap:.2f} м³[/b] непрерывно. После окончания закачки закройте задвижку НКТ и оставьте скважину под давлением на ожидание спада давлений (ОСД) минимум на 2 часа."
                    )
                else:
                    instr = f"ЗАДАВКА (BULLHEADING):\nГлушение давлением без циркуляции. Качай Ро {rho:.2f} напрямую в затрубное пространство. Внимание: контроль давления гидроразрыва пласта!"

            # --- ПЕРЕДАЧА ДАННЫХ НА ЭКРАН РЕЗУЛЬТАТОВ ГЛУШЕНИЯ ---

            res_scr = self.manager.get_screen('kill_res') # Находим объект результирующего экрана с именем 'kill_res'
            res_scr.selected_gno = self.selected_gno # Передаем текущий тип ГНО (Воронка/ЭЦН/Пакер) для вывода на экран результатов
            
            # Формируем строковое значение пластового давления: если оно рассчитывалось по устью, добавляем желтый маркер
            p_pl_lbl = f"{p_pl} атм [color=#FFCC00](Расчет по устью)[/color]" if calculated_p_pl_flag else f"{p_pl} атм" 
            
            # Заполняем главное текстовое поле экрана результатов итоговой сводкой по всем вычисленным параметрам
            res_scr.result_text.text = ( 
                f"[size=22sp][color=#33FF33]{self.selected_gno.upper()}[/color][/size]\n" # Тип компоновки крупным шрифтом
                f"[size=15sp][color=#FF5555]Уровень опасности: {self.well_category}[/color][/size]\n\n" # Уровень опасности скважины
                f"• Плотность жидкости глушения (с коэффициентом): [color=#FF5555][b]{rho:.2f} г/см³[/b][/color]\n" # Расчетная плотность
                f"• Пластовое давление: [b]{p_pl_lbl}[/b]\n" # Пластовое давление с признаком источника
                f"• Объем с запасом регламента (+3м³): [b]{v_total_with_zap:.2f} м³[/b]\n" # Необходимый объем тяжелого раствора
                f"• Время закачки агрегатом: [color=#FFFF33][b]{int(t_min)} мин[/b][/color]\n" # Поминутное время работы насоса ЦА-320
                # Распределение объемов: внутренний объем НКТ и объем внешнего кольцевого пространства ствола
                f"• Инструмент НКТ: [b]{v_trub:.2f} м³[/b] | Кольцевое: [b]{v_zatrub:.2f} м³[/b]" 
            ) 
            
            res_scr.instruction_text.text = instr # Передаем сгенерированную пошаговую инструкцию IWCF на экран результатов
            self.manager.current = 'kill_res' # Физически переключаем ScreenManager на отображение экрана результатов
            
        except Exception as e: 
            # Отлавливаем любые непредвиденные математические или системные сбои внутри монолитной функции вычислений
            print(f"Error Monolith Calc: {e}") # Выводим лог ошибки в консоль/терминал, защищая приложение от падения

    # --- МЕНЮ ВЫПАДАЮЩИХ СПИСКОВ С КРИТЕРИЯМИ КАТЕГОРИЙ СУРГУТНЕФТЕГАЗА --- 
    def menu_category(self, b): 
        """Метод формирует массив элементов для выпадающего меню выбора категорий газоопасности скважины"""
        items = [ 
            # 1 Категория: фонтанирующие или газовые скважины, высокий газовый фактор (>200 м3/т), АВПД или сероводород
            {
                "viewclass": "OneLineListItem", 
                "text": "1 КАТЕГОРИЯ: Газ; ГФ >= 200; АВПД; H2S", 
                "theme_text_color": "Custom", 
                "text_color": (1,1,1,1), 
                "on_release": lambda: self.set_category_val("Скважина 1 категории!")
            }, 
            # 2 Категория: нефтяные скважины, где пластовое давление выше гидростатического, сероводород в пределах ПДК
            {
                "viewclass": "OneLineListItem", 
                "text": "2 КАТЕГОРИЯ: Нефть; Рпл > Ргидр; H2S в ПДК", 
                "theme_text_color": "Custom", 
                "text_color": (1,1,1,1), 
                "on_release": lambda: self.set_category_val("Скважина 2 категории!")
            }, 
            # 3 Категория: стабильные условия КРС, пластовое давление ниже гидростатического столба, сероводорода нет
            {
                "viewclass": "OneLineListItem", 
                "text": "3 КАТЕГОРИЯ: Рпл <= Ргидр; H2S нет", 
                "theme_text_color": "Custom", 
                "text_color": (1,1,1,1), 
                "on_release": lambda: self.set_category_val("Скважина 3 категории!")
            } 
        ]

        
        # ТЕХНОЛОГИЧЕСКИЙ ФИКС: Оптимизировали ширину и зафиксировали позицию меню строго под кнопкой 
        # Инициализируем всплывающее меню категорий газоопасности
        self.m_cat = MDDropdownMenu( 
            caller=b, # Привязываем выпадающее окно к кнопке b, вызвавшей этот метод
            items=items, # Передаем сформированный массив категорий скважины
            width_mult=4.5, # Уменьшили ширину, чтобы меню гарантированно не вылезало за границы экрана смартфона
            position="bottom", # Меню откроется ровно под кнопкой, предотвращая перекрытие шапки
            md_bg_color=(0.1, 0.1, 0.15, 1) # Темный графитовый цвет подложки списка
        ) 
        self.m_cat.open() # Выводим меню выбора категорий на экран

    def set_category_val(self, val): 
        """Метод-обработчик: фиксирует выбранную инженером категорию скважины"""
        self.well_category = val # Записываем полное текстовое описание критериев в системную переменную
        
        # Определяем короткое имя для аккуратного вывода на кнопку (чтобы длинный текст не ломал верстку)
        short_cat = "1 КАТЕГОРИЯ" 
        if "2" in val: 
            short_cat = "2 КАТЕГОРИЯ" 
        elif "3" in val: 
            short_cat = "3 КАТЕГОРИЯ" 
            
        # ЧИСТЫЙ ФИКС: Обновляем текст напрямую через родную переменную кнопки 
        # Проверяем наличие свойства btn_category у класса во избежание сбоя AttributeError
        if hasattr(self, 'btn_category') and self.btn_category: 
            self.btn_category.text = f"ОПАСНОСТЬ: {short_cat}" # Изменяем надпись на кнопке
            
        # Проверяем существование объекта меню и закрываем (скрываем) его после клика пользователя
        if hasattr(self, 'm_cat') and self.m_cat: 
            self.m_cat.dismiss() 
            
        self.refresh_ui_elements() # Перерисовываем интерфейс для пересчета коэффициентов запаса под новую категорию

    def menu_kill_methods(self, b): 
        """Метод генерирует и выводит меню выбора технологических регламентов глушения скважины"""
        items = [ 
            # Массив элементов: при нажатии передают условный строковый маркер технологии (STAND, DRILLER, WAIT, BULL)
            {"viewclass": "OneLineListItem", "text": "ПЛАНОВЫЙ РЕЖИМ ГЛУШЕНИЯ", "theme_text_color": "Custom", "text_color": (1,1,1,1), "on_release": lambda: self.set_kill_method_from_menu("STAND")}, 
            {"viewclass": "OneLineListItem", "text": "МЕТОД БУРИЛЬЩИКА (2 ЦИКЛА)", "theme_text_color": "Custom", "text_color": (1,1,1,1), "on_release": lambda: self.set_kill_method_from_menu("DRILLER")}, 
            {"viewclass": "OneLineListItem", "text": "ОЖИДАНИЕ И УТЯЖЕЛЕНИЕ (1 ЦИКЛ)", "theme_text_color": "Custom", "text_color": (1,1,1,1), "on_release": lambda: self.set_kill_method_from_menu("WAIT")}, 
            {"viewclass": "OneLineListItem", "text": "ЗАДАВКА В ПЛАСТ (BULLHEADING)", "theme_text_color": "Custom", "text_color": (1,1,1,1), "on_release": lambda: self.set_kill_method_from_menu("BULL")} 
        ] 
        # Создаем выпадающий список методов глушения с шириной в 5 блоков
        self.m_kill = MDDropdownMenu(caller=b, items=items, width_mult=5, md_bg_color=(0.1, 0.1, 0.15, 1)) 
        self.m_kill.open() # Открываем меню методов глушения

    def set_kill_method_from_menu(self, mode): 
        """Фиксирует выбранный технологический режим глушения и убирает меню с экрана"""
        self.kill_method = mode # Перезаписываем маркер текущего метода
        if hasattr(self, 'm_kill') and self.m_kill: 
            self.m_kill.dismiss() # Закрываем всплывающее окно методов
        self.refresh_ui_elements() # Синхронизируем надписи на кнопках через центральный UI-обновлятор

    def menu_gno(self, b): 
        """Генерирует компактный список подземного оборудования (ГНО) или текущего состояния пакеров"""
        # Inline-цикл собирает 4 стандартных варианта компоновок низа скважины
        items = [{"viewclass":"OneLineListItem","text":i,"theme_text_color":"Custom","text_color":(1,1,1,1),"on_release":lambda x=i:self.set_gno(x)} for i in ["Воронка", "ЭЦН", "Пакер (уст.)", "Пакер (снят)"]] 
        # Инициализируем меню выбора ГНО с шириной в 4 блока
        self.m_gno = MDDropdownMenu(caller=b, items=items, width_mult=4, md_bg_color=(0.1, 0.1, 0.15, 1)); 
        self.m_gno.open() # Выводим меню выбора ГНО на экран

    def set_gno(self, v): 
        """Сохраняет выбранный тип ГНО и инициирует автоматическую перестройку полей геометрии на экране"""
        self.selected_gno = v # Запоминаем тип компоновки (например, "Воронка" или "ЭЦН")
        if hasattr(self, 'm_gno') and self.m_gno: 
            self.m_gno.dismiss() # Прячем меню выбора ГНО с экрана
        self.refresh_ui_elements() # Критический вызов: метод скроет/покажет поля труб 73/60 или глубины воронки


    def open_nkt_menu(self, b): 
        """Динамически считывает базу данных НКТ из Excel и открывает выпадающий список для выбора труб инженером"""
        try: 
            import openpyxl # Локальный импорт библиотеки работы с Excel-таблицами
            
            # Загружаем книгу Excel по пути EXCEL_PATH, извлекая только вычисленные значения ячеек (без формул)
            wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True); 
            ws = wb.active # Активируем первый (основной) лист рабочей книги
            
            # Извлекаем названия колонок (хедер таблицы) из первой строки файла
            head = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))] 
            
            # Конвертируем строки данных (начиная со 2-й) в список удобных Python-словарей с привязкой к заголовкам колонок
            rows = [dict(zip(head, r)) for r in ws.iter_rows(min_row=2, values_only=True) if r and r] 
            
            # Формируем массив элементов выпадающего меню KivyMD на базе прочитанных типов НКТ
            items = [{"viewclass": "OneLineListItem", "text": f"НКТ {r['Тип_НКТ']}", "theme_text_color":"Custom","text_color":(1,1,1,1), "on_release": lambda x=r: self.set_nkt(x)} for r in rows] 
            
            # Создаем и открываем окно меню выбора труб шириной 4 блока
            self.m_nkt = MDDropdownMenu(caller=b, items=items, width_mult=4, md_bg_color=(0.1, 0.1, 0.15, 1)); 
            self.m_nkt.open() 
        except Exception as e: 
            # Отлавливаем сбои чтения файла (например, если файл поврежден, занят другой программой или отсутствует в data/)
            print(f"Excel Err: {e}") 

    def set_nkt(self, r): 
        """Метод-приемщик: извлекает характеристики выбранной трубы и рассчитывает удельную вместимость"""
        try: 
            # Достаем из словаря внутренний диаметр трубы, если ячейка пустая — подставляем стандартные 62 мм
            d_vnutr = float(r.get('Внутр_d_мм', 62) or 62) 
            
            # Математическая формула площади круга для расчета внутреннего объема 1 погонного метра трубы (в литрах на метр)
            # Умножение квадрата внутреннего диаметра на коэффициент площади круга 0.0007854
            self.v_pipe = (d_vnutr ** 2) * 0.0007854 
            
            # Запоминаем текстовое наименование маркировки выбранной трубы
            self.tech_nkt_name = str(r.get('Тип_НКТ', '')) 
            
            if hasattr(self, 'm_nkt') and self.m_nkt: 
                self.m_nkt.dismiss() # Закрываем меню выбора НКТ
                
            self.refresh_ui_elements() # Синхронизируем интерфейс калькулятора, чтобы убрать оранжевую плашку-предупреждение
        except Exception as e: 
            print(f"Set Nkt Err: {e}") 

    def menu_v(self, b): 
        """Открывает компактное выпадающее меню для выбора диаметра поршневой втулки насоса ЦА-320"""
        # Доступные отраслевые стандарты втулок агрегата: 100 мм, 115 мм, 127 мм
        self.m_v = MDDropdownMenu(caller=b, items=[{"viewclass":"OneLineListItem","text":i,"theme_text_color":"Custom","text_color":(1,1,1,1),"on_release":lambda x=i:self.set_v(x)} for i in ["100", "115", "127"]], width_mult=2, md_bg_color=(0.1, 0.1, 0.15, 1)); 
        self.m_v.open() 

    def set_v(self, v): 
        """Фиксирует выбранный диаметр втулки поршня ЦА-320 и обновляет надпись на кнопке"""
        self.vtulka = v; 
        self.btn_v.text = f"ВТ:{v}"; # Записываем новое сокращенное название на кнопку
        self.m_v.dismiss() # Закрываем меню выбора втулок

    def menu_g(self, b): 
        """Открывает меню для выбора рабочей передачи (скорости) коробки передач цементировочного агрегата"""
        # Силовые передачи ЦА-320, используемые при промывках и продавливании: со 2-й по 5-ю
        items = [{"viewclass":"OneLineListItem","text":f"СК:{i}","theme_text_color":"Custom","text_color":(1,1,1,1),"on_release":lambda x=i:self.set_g(x)} for i in ["2","3","4","5"]] 
        self.m_g = MDDropdownMenu(caller=b, items=items, width_mult=2, md_bg_color=(0.1, 0.1, 0.15, 1)) 
        self.m_g.open() 

    def set_g(self, v): 
        """Фиксирует выбранную передачу КПП насосного агрегата и скрывает список"""
        self.gear = v 
        self.btn_g.text = f"СКОР:{v}" # Визуализируем номер передачи на кнопке
        if hasattr(self, 'm_g') and self.m_g: 
            self.m_g.dismiss() 

    def menu_r(self, b): 
        """Генерирует и открывает список доступных оборотов приводного двигателя (вала насоса) в минуту"""
        # Используем шаг в 100 об/мин: генерируем диапазон рабочих скоростей мотора от 600 до 2000 об/мин
        items = [{"viewclass":"OneLineListItem","text":str(i),"theme_text_color":"Custom","text_color":(1,1,1,1),"on_release":lambda x=i:self.set_r(x)} for i in range(600, 2100, 100)] 
        self.m_r = MDDropdownMenu(caller=b, items=items, width_mult=2, md_bg_color=(0.1, 0.1, 0.15, 1)) 
        self.m_r.open() 

    def set_r(self, v): 
        """Записывает числовое значение текущих оборотов двигателя для расчета поминутного расхода Q"""
        self.rpm = v 
        self.btn_r.text = f"ОБ:{v}" # Отображаем установленные обороты на панели режима ЦА-320
        if hasattr(self, 'm_r') and self.m_r: 
            self.m_r.dismiss() 

    def go_back(self): 
        """Метод навигации: возвращает пользователя из технологического калькулятора на главный экран меню"""
        self.manager.current = 'menu' 


class SpoScreen(BaseScreen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.v_metalla = 0
        self.v_vnutr = 0
        self.siphon_mode = False
        
        main_layout = MDBoxLayout(orientation='vertical', size_hint=(1, 1))

        # 1. ШАПКА
        self.toolbar = MDBoxLayout(
            orientation='horizontal', size_hint_y=None, height=dp(50),
            md_bg_color=(0.1, 0.4, 0.8, 1), padding=[dp(15), 0, dp(10), 0]
        )
        btn_back = MDIconButton(
            icon="arrow-left", pos_hint={'center_y': .5},
            theme_text_color="Custom", text_color=(1, 1, 1, 1),
            on_release=lambda x: self.go_back()
        )
        self.toolbar.add_widget(btn_back)
        self.toolbar.add_widget(MDLabel(
            text="КОНТРОЛЬ ДОЛИВА", theme_text_color="Custom",
            text_color=(1, 1, 1, 1), font_style="H6"
        ))
        main_layout.add_widget(self.toolbar)

        # 2. СКРОЛЛ
        scroll = ScrollView(size_hint=(1, 1))
        self.container = MDBoxLayout(
            orientation='vertical', spacing=dp(15), 
            padding=[dp(20), dp(20), dp(20), dp(20)], 
            size_hint_y=None, adaptive_height=True
        )

        # ФУНКЦИЯ ДЛЯ СОЗДАНИЯ ШИРОКИХ КНОПОК
        def create_wide_btn(text, color, func):
            anchor = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(52))
            btn = MDFillRoundFlatButton(
                text=text, size_hint=(1, 1), md_bg_color=color, on_release=func
            )
            # Снимаем ограничение ширины
            btn.size_hint_max_x = dp(2000)
            anchor.add_widget(btn)
            return anchor, btn

        # КНОПКА НКТ
        anchor_nkt, self.nkt_btn = create_wide_btn("ВЫБРАТЬ ТИП НКТ", (0.1, 0.3, 0.6, 1), self.open_nkt_menu)

        # КНОПКА СИФОНА
        anchor_siph, self.siphon_btn = create_wide_btn("БЕЗ СИФОНА (ТОЛЬКО МЕТАЛЛ)", (0.1, 0.6, 0.2, 1), self.toggle_siphon)

        # ПОЛЯ ВВОДА
        self.count = MDTextField(hint_text="Кол-во трубок / свечей (шт)", mode="fill", input_filter="float", size_hint_x=1)
        self.len_one = MDTextField(hint_text="Длина одной ед. (м)", text="10.1", mode="fill", input_filter="float", size_hint_x=1)

        # КНОПКА РАСЧЕТА
        anchor_calc = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(55))
        self.calc_btn = MDFillRoundFlatIconButton(
            text="РАССЧИТАТЬ ОБЪЕМ", icon="calculator", size_hint=(1, 1),
            md_bg_color=(0.1, 0.4, 0.8, 1), on_release=self.calculate_doliv
        )
        self.calc_btn.size_hint_max_x = dp(2000)
        anchor_calc.add_widget(self.calc_btn)

        # РЕЗУЛЬТАТ
        self.res_label = MDLabel(
            text="[color=#AAAAAA]Выберите НКТ[/color]", 
            theme_text_color="Custom", text_color=(1, 1, 1, 1),
            font_style="H5", size_hint_y=None, halign="center", markup=True
        )

        # КРУПНОЕ ПРЕДУПРЕЖДЕНИЕ ВНИЗУ
        self.warning_label = MDLabel(
            text="[color=#FFCC00][b]ВАЖНО:[/b][/color]\nПеред СПО убедись в наличии солевого раствора не менее [b]4,5 м³[/b]",
            theme_text_color="Custom", text_color=(1, 1, 1, 1),
            font_style="Subtitle1", halign="center", markup=True,
            size_hint_y=None, line_height=1.2
        )
        self.warning_label.bind(texture_size=lambda ins, sz: setattr(ins, 'height', sz[1] + dp(30)))

        # Сборка
        for w in [anchor_nkt, anchor_siph, self.count, self.len_one, anchor_calc, self.res_label, self.warning_label]:
            self.container.add_widget(w)

        scroll.add_widget(self.container)
        main_layout.add_widget(scroll)
        self.add_widget(main_layout)

    def toggle_siphon(self, btn):
        self.siphon_mode = not self.siphon_mode
        btn.text = "С СИФОНОМ (МЕТАЛЛ + ВНУТР.)" if self.siphon_mode else "БЕЗ СИФОНА (ТОЛЬКО МЕТАЛЛ)"
        btn.md_bg_color = (0.8, 0.1, 0.1, 1) if self.siphon_mode else (0.1, 0.6, 0.2, 1)
        if self.v_metalla > 0: self.calculate_doliv()

    def open_nkt_menu(self, button):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
            ws = wb.active
            head = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = [dict(zip(head, r)) for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
            items = [{"viewclass": "OneLineListItem", "text": f"НКТ {r['Тип_НКТ']}", "on_release": lambda x=r: self.set_item(x)} for r in rows]
            self.menu = MDDropdownMenu(caller=button, items=items, width_mult=4)
            self.menu.open()
        except: self.res_label.text = "[color=#FF5555]Ошибка базы[/color]"

    def set_item(self, row):
        self.nkt_btn.text = f"НКТ {row['Тип_НКТ']}"
        self.v_metalla = float(row.get('V_metalla_m3', 0))
        self.v_vnutr = float(row.get('V_vnutr_m3', 0))
        self.menu.dismiss()
        self.calculate_doliv()

    def calculate_doliv(self, *args):
        try:
            if self.v_metalla == 0: return
            cnt = float(self.count.text or 0)
            l_one = float(self.len_one.text or 0)
            v_factor = (self.v_metalla + self.v_vnutr) if self.siphon_mode else self.v_metalla
            vol = (cnt * l_one) * v_factor * 1000
            self.res_label.text = f"[size=24sp][color=#33FF33]ДОЛИВ: {vol:.1f} л[/color][/size]"
        except: self.res_label.text = "Ошибка"

    def go_back(self):
        self.manager.current = 'menu'

class VzdScreen(BaseScreen):    # ВЗД и Режимы. Расчет оборотов винтового двигателя и справочник по долотам/фрезам
    def __init__(self, **kw):
        super().__init__(**kw)
        self.vtulka = "115"
        self.gear = "3"
        self.rpm_ca = 1000
        
        # Коэффициенты подачи Q для ЦА-320 (л/об)
        self.q_coeffs = {
            "100": {"2": 0.171, "3": 0.342, "4": 0.513, "5": 0.777},
            "115": {"2": 0.227, "3": 0.453, "4": 0.680, "5": 1.020},
            "127": {"2": 0.277, "3": 0.555, "4": 0.832, "5": 1.247}
        }
        
        # РАСШИРЕННАЯ БАЗА ВЗД (из техпаспортов)
        self.vzd_data = {
            "Д-76": {"flow": (4, 8), "ratio": 1.25, "torque": 850, "wob": 4.5, "press": 40},
            "Д-85": {"flow": (6, 12), "ratio": 0.88, "torque": 1300, "wob": 6.0, "press": 50},
            "Д-106": {"flow": (8, 16), "ratio": 0.65, "torque": 2200, "wob": 8.0, "press": 65},
            "Д-121": {"flow": (10, 22), "ratio": 0.52, "torque": 2900, "wob": 10.0, "press": 70}
        }
        self.selected_vzd = "Д-85"

        main_layout = MDBoxLayout(orientation='vertical', size_hint=(1, 1))

        # 1. ШАПКА
        self.toolbar = MDBoxLayout(
            orientation='horizontal', size_hint_y=None, height=dp(50),
            md_bg_color=(0.1, 0.4, 0.8, 1), padding=[dp(15), 0, dp(10), 0]
        )
        from kivymd.uix.button import MDIconButton
        btn_back = MDIconButton(
            icon="arrow-left", theme_text_color="Custom", text_color=(1, 1, 1, 1),
            on_release=lambda x: self.go_back()
        )
        self.toolbar.add_widget(btn_back)
        self.toolbar.add_widget(MDLabel(text="ВЗД И ИНСТРУМЕНТ", theme_text_color="Custom", text_color=(1, 1, 1, 1), font_style="H6"))
        main_layout.add_widget(self.toolbar)

        scroll = ScrollView()
        self.container = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(12), padding=dp(15))

        # 2. ВЫБОР ДВИГАТЕЛЯ
        self.btn_vzd = MDFillRoundFlatButton(
            text=f"ДВИГАТЕЛЬ: {self.selected_vzd}", 
            size_hint=(1, None), height=dp(48), md_bg_color=(0.1, 0.3, 0.6, 1),
            on_release=self.menu_vzd
        )
        self.container.add_widget(self.btn_vzd)

        # 3. РЕЖИМ НАСОСА
        self.container.add_widget(MDLabel(text="РЕЖИМ АГРЕГАТА ЦА-320:", font_style="Button", theme_text_color="Hint", halign="center"))
        pump_grid = MDGridLayout(cols=3, spacing=dp(5), size_hint_y=None, height=dp(45))
        
        self.btn_v = MDFillRoundFlatButton(text=f"ВТ:{self.vtulka}", size_hint=(1,1), md_bg_color=(0.1, 0.4, 0.8, 1), on_release=self.menu_v)
        self.btn_g = MDFillRoundFlatButton(text=f"СКОР:{self.gear}", size_hint=(1,1), md_bg_color=(0.1, 0.4, 0.8, 1), on_release=self.menu_g)
        self.btn_r = MDFillRoundFlatButton(text=f"ОБ:{self.rpm_ca}", size_hint=(1,1), md_bg_color=(0.1, 0.4, 0.8, 1), on_release=self.menu_r)
        
        for b in [self.btn_v, self.btn_g, self.btn_r]: pump_grid.add_widget(b)
        self.container.add_widget(pump_grid)

        # 4. ПЕРЕКЛЮЧАТЕЛЬ ИНФОРМАЦИИ
        type_grid = MDGridLayout(cols=3, spacing=dp(5), size_hint_y=None, height=dp(45))
        self.b1 = MDFillRoundFlatButton(text="РЕЖИМ", size_hint=(1,1), on_release=lambda x: self.switch_info("ВЗД"))
        self.b2 = MDFillRoundFlatButton(text="ДОЛОТА", size_hint=(1,1), on_release=lambda x: self.switch_info("ДОЛОТО"))
        self.b3 = MDFillRoundFlatButton(text="ФРЕЗЫ", size_hint=(1,1), on_release=lambda x: self.switch_info("ФРЕЗ"))
        for b in [self.b1, self.b2, self.b3]: type_grid.add_widget(b)
        self.container.add_widget(type_grid)

        # 5. ОКНО ВЫВОДА (С поддержкой разметки)
        self.res_label = MDLabel(
            text="Загрузка данных...", theme_text_color="Custom", text_color=(1, 1, 1, 1),
            font_style="Body1", size_hint_y=None, markup=True, line_height=1.2
        )
        self.res_label.bind(texture_size=lambda ins, sz: setattr(ins, 'height', sz[1]))
        self.container.add_widget(self.res_label)

        scroll.add_widget(self.container)
        main_layout.add_widget(scroll)
        self.add_widget(main_layout)
        self.switch_info("ВЗД")

    # --- ЛОГИКА ---
    def switch_info(self, mode):
        active, inactive = (0.1, 0.4, 0.8, 1), (0.2, 0.2, 0.2, 1)
        self.b1.md_bg_color = active if mode == "ВЗД" else inactive
        self.b2.md_bg_color = active if mode == "ДОЛОТО" else inactive
        self.b3.md_bg_color = active if mode == "ФРЕЗ" else inactive
        if mode == "ВЗД": self.calc_vzd()
        elif mode == "ДОЛОТО": self.show_bits()
        else: self.show_mills()

    def calc_vzd(self):
        # Расчет подачи в л/с
        Q = (self.q_coeffs[self.vtulka][self.gear] * self.rpm_ca) / 60
        v = self.vzd_data[self.selected_vzd]
        rpm_vzd = int(Q * v["ratio"] * 60)
        
        # Цветовая индикация расхода
        status_color = "#33FF33" if v["flow"][0] <= Q <= v["flow"][1] else "#FF5555"
        status_text = "В НОРМЕ" if status_color == "#33FF33" else "ВНЕ ДИАПАЗОНА!"

        self.res_label.text = (
            f"[size=20sp][color=#33FF33]ТЕХ.КАРТА {self.selected_vzd}:[/color][/size]\n\n"
            f"• Подача ЦА: [b]{Q:.2f} л/с[/b] ({status_text})\n"
            f"• Обороты: [color={status_color}][b]{rpm_vzd} об/мин[/b][/color]\n"
            f"• Допуст. расход: [b]{v['flow'][0]}-{v['flow'][1]} л/с[/b]\n"
            f"• Макс. нагрузка (WOB): [b]{v['wob']} тонн[/b]\n"
            f"• Дифф. давление ($\Delta P$): [b]{v['press']/10:.1f} МПа ({v['press']} атм)[/b]\n"
            f"--------------------------------\n"
            f"[color=#FFCC00][b]ПРОВЕРКА НА УСТЬЕ:[/b][/color]\n"
            f"1. Подать миним. расход [b]{v['flow'][0]} л/с[/b].\n"
            f"2. Мотор должен запуститься плавно, без рывков.\n"
            f"3. Давление ХХ должно быть стабильным.\n"
            f"--------------------------------\n"
            f"[color=#FFCC00][b]ПРИЗНАКИ ПОЛОМКИ:[/b][/color]\n"
            f"• Давление растет, а проходки нет — [b]ЗАКЛИНКА[/b].\n"
            f"• Давление упало, обороты выросли — [b]ПРОМЫВ СТАТОРА[/b].\n"
            f"• Стук в колонне — [b]ИЗНОС КАРДАНА[/b]."
        )

    def show_bits(self):
        self.res_label.text = (
            "[size=20sp][color=#33FF33]ИНСТРУКЦИЯ: ДОЛОТА[/color][/size]\n\n"
            "[b]PDC (Алмазные):[/b]\n"
            "• Применяются на высоких оборотах (>100 об/мин).\n"
            "• [color=#FF5555]Запрещено:[/color] работа по металлу и «прыжки».\n"
            "• Нагрузка: плавная, 1-3 тонны.\n\n"
            "[b]Шарошечные (Т, СТ):[/b]\n"
            "• Эффективны на малых оборотах с высокой нагрузкой.\n"
            "• Нагрузка: до [b]6-8 тонн[/b] (зависит от Ø).\n"
            "• Обязательна приработка подшипников (15 мин)."
        )

    def show_mills(self):
        self.res_label.text = (
            "[size=20sp][color=#33FF33]ИНСТРУКЦИЯ: ФРЕЗЫ[/color][/size]\n\n"
            "[b]Торцевые (ФТ):[/b]\n"
            "• Работа по «голове» аварийного инструмента.\n"
            "• Обязательна интенсивная промывка для выноса стружки.\n\n"
            "[b]Магнитные (ФМ):[/b]\n"
            "• Сбор мелкого металла (зубья, сухари).\n"
            "• Спускать без вращения ВЗД (на прямую промывку)."
        )

    # --- МЕНЮ ВЫБОРА ---
    def menu_vzd(self, b):
        items = [{"viewclass":"OneLineListItem","text":k,"on_release":lambda x=k: self.set_vzd(x)} for k in self.vzd_data.keys()]
        self.m_vzd = MDDropdownMenu(caller=b, items=items, width_mult=4); self.m_vzd.open()
    def set_vzd(self, v): self.selected_vzd = v; self.btn_vzd.text = f"ДВИГАТЕЛЬ: {v}"; self.m_vzd.dismiss(); self.calc_vzd()

    def menu_v(self, b):
        items = [{"viewclass":"OneLineListItem","text":i,"on_release":lambda x=i:self.set_v(x)} for i in ["100", "115", "127"]]
        self.m_v = MDDropdownMenu(caller=b, items=items, width_mult=2); self.m_v.open()
    def set_v(self, v): self.vtulka = v; self.btn_v.text = f"ВТ:{v}"; self.m_v.dismiss(); self.calc_vzd()

    def menu_g(self, b):
        items = [{"viewclass":"OneLineListItem","text":f"СК:{i}","on_release":lambda x=i:self.set_g(x)} for i in ["2","3","4","5"]]
        self.m_g = MDDropdownMenu(caller=b, items=items, width_mult=2); self.m_g.open()
    def set_g(self, v): self.gear = v; self.btn_g.text = f"СКОР:{v}"; self.m_g.dismiss(); self.calc_vzd()

    def menu_r(self, b):
        items = [{"viewclass":"OneLineListItem","text":str(i),"on_release":lambda x=i:self.set_r(x)} for i in range(600, 2100, 100)]
        self.m_r = MDDropdownMenu(caller=b, items=items, width_mult=2); self.m_r.open()
    def set_r(self, v): self.rpm_ca = v; self.btn_r.text = f"ОБ:{v}"; self.m_r.dismiss(); self.calc_vzd()

    def go_back(self): self.manager.current = 'menu'

class ThreadsScreen(BaseScreen):      # Справочник инструмента. Данные по НКТ, СБТ и пакерам (нагрузки, веса, объемы).
    def __init__(self, **kw):
        super().__init__(**kw)
        self.sel_type = "НКТ"
        self.sel_size = "73 x 5.5"
        self.sel_grade = "Д"
        
        # Группы прочности сталей (Предел текучести в МПа)
        self.steel_grades = {"Д": 379, "К": 492, "Е": 552, "Л": 655, "М": 758, "Р": 862}
        
        layout = MDBoxLayout(orientation='vertical')
        
        # 1. ШАПКА
        self.toolbar = MDBoxLayout(
            orientation='horizontal', size_hint_y=None, height=dp(45),
            md_bg_color=(0.1, 0.4, 0.8, 1), padding=[dp(15), 0, dp(10), 0]
        )
        btn_back = MDIconButton(
            icon="arrow-left", pos_hint={'center_y': .5},
            theme_text_color="Custom", text_color=(1, 1, 1, 1),
            on_release=lambda x: setattr(self.manager, 'current', 'menu')
        )
        self.toolbar.add_widget(btn_back)
        self.toolbar.add_widget(MDLabel(text="СПРАВОЧНИК ИНСТРУМЕНТА", theme_text_color="Custom", 
                                       text_color=(1,1,1,1), font_style="H6"))
        layout.add_widget(self.toolbar)

        scroll = ScrollView()
        self.container = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(12), padding=dp(15))

        # 2. ТАБЫ (НКТ / СБТ / ПАКЕР)
        type_grid = MDGridLayout(cols=3, spacing=dp(5), size_hint_y=None, height=dp(45))
        
        self.btn_tab_nkt = MDFillRoundFlatButton(text="НКТ", size_hint=(1, 1), on_release=lambda x: self.switch_tab("НКТ"))
        self.btn_tab_sbt = MDFillRoundFlatButton(text="СБТ", size_hint=(1, 1), on_release=lambda x: self.switch_tab("СБТ"))
        self.btn_tab_pak = MDFillRoundFlatButton(text="ПАКЕР", size_hint=(1, 1), on_release=lambda x: self.switch_tab("ПАКЕР"))
        
        for b in [self.btn_tab_nkt, self.btn_tab_sbt, self.btn_tab_pak]: type_grid.add_widget(b)
        self.container.add_widget(type_grid)

        # 3. КНОПКИ ПАРАМЕТРОВ
        self.anchor_size = MDAnchorLayout(anchor_x='center', size_hint=(1, None), height=dp(50))
        self.btn_size = MDFillRoundFlatButton(text="РАЗМЕР", on_release=self.menu_sizes, size_hint=(1, 1), md_bg_color=(0.1, 0.3, 0.6, 1))
        self.anchor_size.add_widget(self.btn_size)

        self.anchor_steel = MDAnchorLayout(anchor_x='center', size_hint=(1, None), height=dp(50))
        self.btn_steel = MDFillRoundFlatButton(text="ГРУППА ПРОЧНОСТИ", on_release=self.menu_steel, size_hint=(1, 1), md_bg_color=(0.1, 0.3, 0.6, 1))
        self.anchor_steel.add_widget(self.btn_steel)
        
        self.container.add_widget(self.anchor_size)
        self.container.add_widget(self.anchor_steel)

        # 4. ВЫВОД ИНФОРМАЦИИ
        self.res_card = MDLabel(
            text="Загрузка...", theme_text_color="Custom", text_color=(1,1,1,1),
            font_style="Body1", size_hint_y=None, markup=True, line_height=1.3
        )
        self.res_card.bind(texture_size=lambda ins, sz: setattr(ins, 'height', sz[1]))
        self.container.add_widget(self.res_card)
        
        # ТЕХНОЛОГИЧЕСКИЙ ФИКС: Врезали кнопку перехода в блок ВЗД из Справочника
        self.container.add_widget(MDBoxLayout(size_hint_y=None, height=dp(10))) # Небольшой зазор
        self.anchor_vzd_btn = MDAnchorLayout(anchor_x='center', size_hint=(1, None), height=dp(50))
        self.btn_goto_vzd = MDFillRoundFlatButton(
            text="⚙️   ПЕРЕЙТИ В КАТАЛОГ ВЗД",
            on_release=lambda x: setattr(self.manager, 'current', 'vzd'),
            size_hint=(1, 1),
            md_bg_color=(0.1, 0.4, 0.8, 1) # Красивый фирменный синий цвет
        )
        self.anchor_vzd_btn.add_widget(self.btn_goto_vzd)
        self.container.add_widget(self.anchor_vzd_btn)

        scroll.add_widget(self.container)

        layout.add_widget(scroll)
        self.add_widget(layout)
        self.switch_tab("НКТ")

    # --- ЛОГИКА ---
    def switch_tab(self, mode):
        self.sel_type = mode
        active, inactive = (0.1, 0.4, 0.8, 1), (0.2, 0.2, 0.2, 1)
        self.btn_tab_nkt.md_bg_color = active if mode == "НКТ" else inactive
        self.btn_tab_sbt.md_bg_color = active if mode == "СБТ" else inactive
        self.btn_tab_pak.md_bg_color = active if mode == "ПАКЕР" else inactive
        
        if mode == "ПАКЕР":
            self.sel_size = "ПМР-122"
            self.anchor_steel.opacity, self.anchor_steel.disabled = 0, True
        else:
            self.sel_size = "73 x 5.5" if mode == "НКТ" else "СБТ 73"
            self.anchor_steel.opacity, self.anchor_steel.disabled = 1, False
        self.update_info()

    def menu_sizes(self, b):
        if self.sel_type == "НКТ": 
            list_s = ["60 x 5.0", "73 x 5.5", "89 x 6.5", "НКТУ 89 x 7.3"]
        elif self.sel_type == "СБТ": 
            list_s = ["СБТ 60", "СБТ 73", "СБТ 89"]
        else: 
            list_s = ["ПМР-122", "ПМР-140", "ПРО-ЯДЖО-О", "2ПНМ-122"]
        items = [{"viewclass": "OneLineListItem", "text": i, "on_release": lambda x=i: self.set_size(x)} for i in list_s]
        self.m_size = MDDropdownMenu(caller=b, items=items, width_mult=4); self.m_size.open()

    def set_size(self, v):
        self.sel_size = v; self.m_size.dismiss(); self.update_info()

    def menu_steel(self, b):
        items = [{"viewclass": "OneLineListItem", "text": f"Сталь: {k}", "on_release": lambda x=k: self.set_grade(x)} for k in self.steel_grades.keys()]
        self.m_steel = MDDropdownMenu(caller=b, items=items, width_mult=3); self.m_steel.open()

    def set_grade(self, v):
        self.sel_grade = v; self.m_steel.dismiss(); self.update_info()

    def update_info(self):
        self.btn_size.text = f"РАЗМЕР: {self.sel_size}"
        self.btn_steel.text = f"СТАЛЬ: {self.sel_grade}"
        self.res_card.text = self.get_packer_info(self.sel_size) if self.sel_type == "ПАКЕР" else self.get_pipe_info()

    def get_pipe_info(self):
        # Расширенная база: {площадь_мм2, вес_кг, объем_л_м, резьба, момент_нм}
        data = {
            "60 x 5.0": {"area": 864, "weight": 7.1, "vol": 1.96, "thread": "60 ГОСТ 633", "torque": "2500-3500"},
            "73 x 5.5": {"area": 1166, "weight": 9.5, "vol": 3.02, "thread": "73 ГОСТ 633", "torque": "4500-5500"},
            "89 x 6.5": {"area": 1685, "weight": 13.5, "vol": 4.54, "thread": "89 ГОСТ 633", "torque": "7000-8000"},
            "НКТУ 89 x 7.3": {"area": 1870, "weight": 15.2, "vol": 4.35, "thread": "89-У (Усиленная)", "torque": "9000-11000"},
            "СБТ 60": {"area": 1100, "weight": 10.2, "vol": 1.45, "thread": "З-73 (NC26)", "torque": "5500-7500"},
            "СБТ 73": {"area": 1500, "weight": 14.5, "vol": 2.15, "thread": "З-86 (NC31)", "torque": "8000-10000"},
            "СБТ 89": {"area": 2100, "weight": 19.8, "vol": 3.42, "thread": "З-102 (NC38)", "torque": "12000-15000"}
        }
        d = data.get(self.sel_size, data["73 x 5.5"])
        sigma = self.steel_grades[self.sel_grade]
        force = (d["area"] * sigma) / 9810 * 0.8 # Тонны с запасом 0.8
        
        res = f"[size=22sp][color=#33FF33]{self.sel_size}[/color][/size]\n\n"
        res += f"• Разрывное (запас 0.8): [color=#FF5555][b]{force:.1f} тонн[/b][/color]\n"
        res += f"• Резьба: [b]{d['thread']}[/b]\n"
        res += f"• Момент затяжки: [b]{d['torque']} Н*м[/b]\n"
        res += f"• Вес метра: [b]{d['weight']} кг/м[/b] | Объем: [b]{d['vol']} л/м[/b]\n"
        res += f"• Группа прочности: [b]{self.sel_grade}[/b] ({sigma} МПа)\n"
        
        if "НКТУ" in self.sel_size:
            res += "--------------------------------\n[i]Применяется при ГРП и МГРП. Имеет увеличенную толщину стенки и усиленную муфту.[/i]"
        return res

    def get_packer_info(self, name):
        # Подробные инструкции по эксплуатации
        packers = {
            "ПМР-122": (
                "[b]Инструкция ПМР (Механический):[/b]\n"
                "1. Спустить до глубины. Выше цели на 1-2м.\n"
                "2. Приподнять на 0.5м, повернуть влево на 1/4 оборота.\n"
                "3. Разгрузить вес [b]8-10 тонн[/b]. Контроль по ИВЭ.\n"
                "• [color=#FFFF33]Снятие:[/color] Прямая натяжка вверх."
            ),
            "ПМР-140": (
                "[b]Инструкция ПМР-140:[/b]\n"
                "1. Спуск в колонну 168 мм.\n"
                "2. Активация осевым перемещением с поворотом.\n"
                "3. Рекомендуемая нагрузка [b]10-12 тонн[/b].\n"
                "• [color=#FFFF33]Внимание:[/color] Перед снятием выровнять давление."
            ),
            "ПРО-ЯДЖО-О": (
                "[b]Инструкция ПРО-ЯДЖО (Якорный):[/b]\n"
                "1. Спуск до интервала.\n"
                "2. Поворот вправо на 1/4 оборота.\n"
                "3. Натяжка вверх или разгрузка (зависит от исполнения).\n"
                "4. Удерживает перепад до [b]35-50 МПа[/b].\n"
                "• [color=#FFFF33]Снятие:[/color] Осевое перемещение в нейтраль."
            ),
            "2ПНМ-122": (
                "[b]Инструкция 2ПНМ (Гидравлический):[/b]\n"
                "1. Спустить на проектную глубину.\n"
                "2. Бросить шар (если предусмотрено) или создать Р.\n"
                "3. Поднять давление в трубах до [b]120-150 атм[/b].\n"
                "4. Выдержать 5-10 мин для фиксации плашек.\n"
                "• [color=#FFFF33]Снятие:[/color] Натяжка с разрывом срезных винтов."
            )
        }
        return f"[size=22sp][color=#33FF33]{name}[/color][/size]\n\n" + packers.get(name, "Инструкция дополняется...")

class FishingScreen(BaseScreen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.area_cm2 = 0
        
        main_layout = MDBoxLayout(orientation='vertical', size_hint=(1, 1))
        
        # 1. ШАПКА
        self.toolbar = MDBoxLayout(
            orientation='horizontal', size_hint_y=None, height=dp(50),
            md_bg_color=(0.1, 0.4, 0.8, 1), padding=[dp(15), 0, dp(10), 0]
        )
        self.toolbar.add_widget(MDIconButton(
            icon="arrow-left", theme_text_color="Custom", text_color=(1, 1, 1, 1),
            on_release=lambda x: self.go_back()
        ))
        self.toolbar.add_widget(MDLabel(text="ЛОВИЛЬНЫЙ ЦЕНТР", theme_text_color="Custom", 
                                       text_color=(1,1,1,1), font_style="H6"))
        main_layout.add_widget(self.toolbar)

        # 2. ОСНОВНОЙ СКРОЛЛ
        scroll = ScrollView()
        self.container = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(15), padding=dp(20))

        # ТАБЫ (РЕЖИМЫ)
        tab_grid = MDGridLayout(cols=2, spacing=dp(5), size_hint_y=None, height=dp(45))
        self.b1 = MDFillRoundFlatButton(text="РАСЧЕТ ПРИХВАТА", size_hint=(1,1), on_release=lambda x: self.switch_mode("CALC"))
        self.b2 = MDFillRoundFlatButton(text="ИНСТРУМЕНТАРИЙ", size_hint=(1,1), on_release=lambda x: self.switch_mode("TOOL"))
        tab_grid.add_widget(self.b1); tab_grid.add_widget(self.b2)
        self.container.add_widget(tab_grid)

        # КОНТЕНТ
        self.content_box = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(15))
        self.container.add_widget(self.content_box)

        scroll.add_widget(self.container); main_layout.add_widget(scroll); self.add_widget(main_layout)
        self.switch_mode("CALC")

    def switch_mode(self, mode):
        self.content_box.clear_widgets()
        active, inactive = (0.1, 0.4, 0.8, 1), (0.2, 0.2, 0.2, 1)
        self.b1.md_bg_color = active if mode == "CALC" else inactive
        self.b2.md_bg_color = active if mode == "TOOL" else inactive
        if mode == "CALC": self.draw_calc_ui()
        else: self.draw_tool_info()

    def draw_calc_ui(self):
        # Инструкция (Кратко и понятно)
        help_t = (
            "[color=#FFCC00][b]ЗАМЕР ВЫТЯЖКИ:[/b][/color]\n"
            "• Натяни до собственного веса, поставь метку.\n"
            "• Дай натяжку выше веса на [b]P[/b] (10-15 тн).\n"
            "• Замерь вытяжку трубы [b]dL[/b] в сантиметрах."
        )
        self.content_box.add_widget(MDLabel(text=help_t, markup=True, theme_text_color="Secondary", size_hint_y=None, adaptive_height=True))

        # Кнопка НКТ
        anchor_nkt = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(50))
        self.btn_nkt = MDFillRoundFlatButton(text="1. ВЫБРАТЬ НКТ", size_hint=(1, 1), md_bg_color=(0.1, 0.3, 0.6, 1), on_release=self.open_nkt_menu)
        self.btn_nkt.size_hint_max_x = dp(2000); anchor_nkt.add_widget(self.btn_nkt)
        self.content_box.add_widget(anchor_nkt)

        self.f_force = MDTextField(hint_text="Натяжка P (тонн)", mode="fill", input_filter="float")
        self.f_stretch = MDTextField(hint_text="Вытяжка dL (сантиметры)", mode="fill", input_filter="float")
        self.content_box.add_widget(self.f_force); self.content_box.add_widget(self.f_stretch)

        # Кнопка расчета
        anchor_calc = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(55))
        btn_calc = MDFillRoundFlatIconButton(text="РАСЧИТАТЬ ГЛУБИНУ", icon="calculator", size_hint=(1, 1), md_bg_color=(0.1, 0.6, 0.2, 1), on_release=self.calc_stuck)
        btn_calc.size_hint_max_x = dp(2000); anchor_calc.add_widget(btn_calc)
        self.content_box.add_widget(anchor_calc)

        # Табло результата (Карточка)
        self.res_card = MDCard(orientation='vertical', padding=dp(15), size_hint_y=None, adaptive_height=True, md_bg_color=(0,0,0,1))
        self.res_label = MDLabel(text="Верхняя граница прихвата: -- м", halign="center", theme_text_color="Custom", text_color=(0.2, 1, 0.2, 1), bold=True)
        self.res_card.add_widget(self.res_label)
        self.content_box.add_widget(self.res_card)

    def draw_tool_info(self):
        info = (
            "[size=18sp][color=#33FF33]БАЗА ЛОВИЛЬНОГО ИНСТРУМЕНТА:[/color][/size]\n\n"
            "[b]1. ОВЕРШОТЫ (ТИП ОС):[/b]\n"
            "• Назначение: Наружный захват за муфту/тело.\n"
            "• [b]Захват:[/b] Спиральные или цанговые граппы.\n"
            "• [b]Метод:[/b] Накрыть «голову», посадка 1-2т, поворот вправо 1/4, натяжка.\n\n"
            "[b]2. МЕТЧИКИ (МЭС / МЭУ):[/b]\n"
            "• Назначение: Внутренний захват за резьбу или тело.\n"
            "• [b]МЭС:[/b] Специальная резьба для врезки.\n"
            "• [b]Метод:[/b] Заворот вправо 5-8 оборотов. ОСТОРОЖНО: Инструмент хрупкий!\n\n"
            "[b]3. КОЛОКОЛА (К / КГ):[/b]\n"
            "• Назначение: Нарезка наружной резьбы на инструменте.\n"
            "• [b]Метод:[/b] Вращение 15-20 об/мин с осевой нагрузкой 1.5-2.0 тн.\n\n"
            "[b]4. ПЕЧАТИ (П-140 / П-118):[/b]\n"
            "• Определение формы «головы» и её положения в колонне.\n"
            "• [b]Важно:[/b] Посадка [b]не более 2-3 тонн[/b], чтобы не смять корпус."
        )
        self.content_box.add_widget(MDLabel(text=info, markup=True, theme_text_color="Custom", text_color=(1,1,1,1), size_hint_y=None, adaptive_height=True))

    def calc_stuck(self, *args):
        try:
            if self.area_cm2 == 0:
                self.res_label.text = "Сначала выберите НКТ!"; return
            P = float(self.f_force.text or 0) * 1000 # тн -> кг
            L = float(self.f_stretch.text or 0)      # см
            if P <= 0: return
            # Формула: H = (E * F * dL) / P. E=2.1e6
            depth = (2100000 * self.area_cm2 * L) / P / 100 # см -> м
            self.res_label.text = f"Верх прихвата: ~{int(depth)} метров"
        except: self.res_label.text = "Ошибка в данных!"

    def open_nkt_menu(self, b):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True); ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
            items = [{"viewclass": "OneLineListItem", "text": f"НКТ {r['Тип_НКТ']}", "on_release": lambda x=r: self.set_nkt(x)} for r in rows]
            self.menu = MDDropdownMenu(caller=b, items=items, width_mult=4); self.menu.open()
        except: self.res_label.text = "Ошибка базы"

    def set_nkt(self, r):
        # S сечения = V_metalla * 10000 (перевод м2 в см2)
        self.area_cm2 = float(r.get('V_metalla_m3', 0)) * 10000
        self.btn_nkt.text = f"ИНСТРУМЕНТ: {r['Тип_НКТ']}"; self.menu.dismiss()

    def go_back(self): self.manager.current = 'menu'
# ДЕКЛАРАТИВНАЯ РАЗМЕТКА ЭКРАНА СПИСКА БРИГАД (ПО АНАЛОГИИ С ЖУРНАЛОМ СКВАЖИН)
Builder.load_string('''
<OpenBrigadesScreen>:
    MDBoxLayout:
        orientation: "vertical"
        md_bg_color: [0.05, 0.05, 0.05, 1] # Глубокий темный фон приложения WELL CONTROL

        # 1. ВЕРХНЯЯ ШАПКА ЭКРАНА
        MDTopAppBar:
            title: "Справочник бригад КРС"
            anchor_title: "left"
            elevation: 4
            md_bg_color: [0.2, 0.25, 0.3, 1] # Графитовый цвет административного блока
            left_action_items: [["arrow-left", lambda x: root.go_back()]]

        # Main контентная зона
        MDBoxLayout:
            orientation: "vertical"
            padding: ["15dp", "10dp", "15dp", "15dp"]
            spacing: "12dp"

            # 2. МАССИВНАЯ УПРАВЛЯЮЩАЯ КНОПКА ДОБАВЛЕНИЯ ПЕРСОНАЛА
            MDAnchorLayout:
                anchor_x: "center"
                size_hint_y: None
                height: "48dp"
                MDFillRoundFlatIconButton:
                    text: "➕ ДОБАВИТЬ НОВУЮ БРИГАДУ В БАЗУ"
                    icon: "account-plus-outline"
                    size_hint: [1, 1]
                    size_hint_max_x: "2000dp"
                    md_bg_color: [0.1, 0.4, 0.8, 1] # Синий цвет сервисных действий
                    on_release: root.open_create_screen()

            MDLabel:
                text: "ЗАРЕГИСТРИРОВАННЫЙ ПЕРСОНАЛ ТКРС:"
                font_style: "Caption"
                theme_text_color: "Hint"
                size_hint_y: None
                height: "20dp"

            # 3. ЗОНА СВОБОДНОЙ ПРОКРУТКИ КАРТОЧЕК МАСТЕРОВ
            ScrollView:
                do_scroll_x: False
                do_scroll_y: True
                MDBoxLayout:
                    id: brigades_container
                    orientation: "vertical"
                    spacing: "8dp"
                    adaptive_height: True
''')
class OpenBrigadesScreen(MDScreen): # Класс экрана, отображающего весь список персонала бригад КРС из базы SQLite
    def go_back(self):
        """Метод навигации: возвращает пользователя назад на экран глобальных настроек приложения"""
        self.manager.current = "settings"

    def open_create_screen(self):
        """Метод навигации: переключает интерфейс на экран создания новой карточки бригады"""
        self.manager.current = "create_brigade_screen"

    def on_enter(self):
        """Автоматическое событие Kivy: срабатывает при входе на экран и динамически выстраивает список мастеров из БД"""
        from kivymd.uix.list import TwoLineListItem # Локальный импорт Material Design двухстрочного элемента списка
        from kivymd.uix.label import MDLabel # Локальный импорт текстовой метки
        
        # Гарантированно очищаем контейнер, убирая старые элементы, чтобы избежать дублирования при повторном входе
        self.ids.brigades_container.clear_widgets()
        
        conn = get_db_connection() # Устанавливаем официальную сессию соединения с локальным файлом базы SQLite
        cursor = conn.cursor()
        # Выгружаем данные (номер бригады, ФИО мастера, внутренний и мобильный телефоны) с сортировкой от меньшего номера к большему
        cursor.execute("SELECT number, master, phone_int, phone_mob FROM brigades ORDER BY number ASC")
        rows = cursor.fetchall() # Извлекаем все найденные строки из таблицы
        conn.close() # Обязательно закрываем сессию базы данных, освобождая дескриптор
        
        if not rows:
            # Если база данных пустая, выводим информирующую заглушку по центру контентной зоны
            lbl = MDLabel(
                text="Справочник пуст. Нажмите кнопку выше, чтобы добавить первую бригаду.",
                halign="center", # Центрирование текста по горизонтали
                theme_text_color="Hint", # Серый цвет шрифта системных подсказок
                size_hint_y=None, 
                height="48dp",
                font_style="Body2"
            )
            self.ids.brigades_container.add_widget(lbl) # Выводим заглушку на экран
        else:
            # Если записи обнаружены, запускаем цикл генерации строчек персонала
            for row in rows:
                b_num = str(row[0]) # Номер бригады (Индекс 0)
                b_master = str(row[1]) # ФИО мастера бригады (Индекс 1)
                
                # Формируем префиксы связи: если номера существуют, подставляем маркеры связи промысла
                p_int = f"вн. {row[2]}" if row[2] else "" # Внутренний АТС-номер (Индекс 2)
                p_mob = f"сот. {row[3]}" if row[3] else "" # Сотовый / мобильный номер (Индекс 3)
                
                # Склеиваем доступные контакты в одну строку через красивый разделитель «вертикальная черта»
                phones = " | ".join([p for p in [p_int, p_mob] if p])
                
                # Создаем двухстрочный графический элемент списка Material Design
                item = TwoLineListItem(
                    text=f"Бригада №{b_num} — мастер {b_master}", # Верхняя строка: номер и ФИО руководящего ИТР
                    secondary_text=phones if phones else "Контакты не указаны", # Нижняя строка: собранные номера связи
                    theme_text_color="Custom", 
                    text_color=(1, 1, 1, 1), # Чистый белый цвет букв заголовка
                    secondary_theme_text_color="Custom", 
                    secondary_text_color=(0.9, 0.4, 0.0, 1) # Фирменный ярко-оранжевый цвет для выделения телефонов
                )
                
                # Помещаем готовую строчку в наш вертикальный список brigades_container внутри зоны прокрутки
                self.ids.brigades_container.add_widget(item)
# ДЕКЛАРАТИВНАЯ РАЗМЕТКА ЭКРАНА СОЗДАНИЯ КАРТОЧКИ БРИГДЫ (ПО АНАЛОГИИ С ДОБАВЛЕНИЕМ СКВАЖИН)
Builder.load_string('''
<CreateBrigadeScreen>:
    MDBoxLayout:
        orientation: "vertical"
        md_bg_color: [0.05, 0.05, 0.05, 1] # Единый темный фон WELL CONTROL

        # 1. ШАПКА ЭКРАНА (ТУЛБАР)
        MDTopAppBar:
            title: "Новая бригада"
            anchor_title: "left"
            elevation: 4
            md_bg_color: [0.2, 0.25, 0.3, 1] # Графитовый цвет административного блока
            left_action_items: [["arrow-left", lambda x: root.go_back()]]

        # Зона прокрутки формы ввода параметров персонала
        ScrollView:
            do_scroll_x: False
            do_scroll_y: True
            MDBoxLayout:
                orientation: "vertical"
                padding: ["20dp", "15dp", "20dp", "20dp"]
                spacing: "14dp"
                adaptive_height: True

                MDLabel:
                    text: "РЕГИСТРАЦИОННЫЕ ДАННЫЕ ПЕРСОНАЛА:"
                    font_style: "Caption"
                    theme_text_color: "Hint"
                    size_hint_y: None
                    height: "20dp"

                # 2. ПОЛЯ ВВОДА ХАРАКТЕРИСТИК БРИГАДЫ ТКРС
                MDTextField:
                    id: f_br_num
                    hint_text: "№ бригады КРС / ТКРС"
                    mode: "fill"
                    size_hint_y: None
                    height: "44dp"
                    input_type: "number" # Открывает только цифровую клавиатуру на Android

                MDTextField:
                    id: f_br_master
                    hint_text: "ФИО мастера (Руководитель ИТР)"
                    mode: "fill"
                    size_hint_y: None
                    height: "44dp"
                    # Текстовое поле без ограничений — пишите любые фамилии буквами!

                MDTextField:
                    id: f_br_phone_int
                    hint_text: "Внутренний промысловый телефон (АТС)"
                    mode: "fill"
                    size_hint_y: None
                    height: "44dp"
                    input_type: "number"

                MDTextField:
                    id: f_br_phone_mob
                    hint_text: "Сотовый / спутниковый телефон мастера"
                    mode: "fill"
                    size_hint_y: None
                    height: "44dp"
                    input_type: "number"

                MDBoxLayout:
                    size_hint_y: None
                    height: "10dp"

                # 3. КНОПКА ФИКСАЦИИ И СОХРАНЕНИЯ В SQLITE
                MDAnchorLayout:
                    anchor_x: "center"
                    size_hint_y: None
                    height: "50dp"
                    MDFillRoundFlatIconButton:
                        text: "СОХРАНИТЬ КАРТОЧКУ В БАЗУ ДАННЫХ"
                        icon: "content-save-outline"
                        size_hint: (1, 1)
                        size_hint_max_x: "2000dp"
                        md_bg_color: [0.1, 0.6, 0.2, 1] # Насыщенный зеленый цвет успеха
                        on_release: root.save_brigade()
''')
class CreateBrigadeScreen(MDScreen): # Класс экрана создания новой карточки бригады, унаследованный от MDScreen
    def go_back(self):
        """Метод навигации: возвращает пользователя назад к экрану списка бригад КРС"""
        self.manager.current = "open_brigades_screen"

    def save_brigade(self):
        """Бэкенд-метод: считывает текст напрямую из ID-виджетов KV-разметки экрана и производит запись в SQLite"""
        # Считываем текстовые значения из полей ввода, очищая случайные пробелы по краям через strip()
        num_txt = self.ids.f_br_num.text.strip()
        master_txt = self.ids.f_br_master.text.strip()
        int_txt = self.ids.f_br_phone_int.text.strip()
        mob_txt = self.ids.f_br_phone_mob.text.strip()

        # Валидация безопасности: если номер бригады или ФИО мастера оставлены пустыми, блокируем сохранение пустых строк
        if not num_txt or not master_txt:
            return # Досрочно прерываем функцию, защищая SQLite от пустых полей

        try:
            conn = get_db_connection() # Устанавливаем официальную сессию соединения с локальным файлом базы SQLite
            cursor = conn.cursor()
            
            # СТРОГИЙ СУРГУТСКИЙ РЕГЛАМЕНТ: Записываем данные в таблицу бригад. Инструкция REPLACE гарантирует, 
            # что если номер бригады (Unique) уже существовал в системе, старая карточка просто обновится без создания дубликатов
            cursor.execute(
                "INSERT OR REPLACE INTO brigades (number, master, phone_int, phone_mob) VALUES (?, ?, ?, ?)",
                (num_txt, master_txt, int_txt, mob_txt)
            )
            conn.commit() # Жестко фиксируем транзакцию изменений в файле базы на флеш-памяти телефона
            conn.close() # Закрываем соединение, освобождая дескриптор базы
            
            # Очищаем все поля ввода после успешного сохранения карточки, чтобы при следующем входе они были пустыми
            self.ids.f_br_num.text = ""
            self.ids.f_br_master.text = ""
            self.ids.f_br_phone_int.text = ""
            self.ids.f_br_phone_mob.text = ""
            
            # Перенаправляем инженера обратно на экран Справочника, который при входе (на событии on_enter) мгновенно перерисует список уже с новой строкой бригады!
            self.go_back()
            
        except Exception as e:
            # Ловим любые непредвиденные системные сбои (например, если заблокирована таблица brigades)
            print(f"Ошибка сохранения бригады в SQLite: {e}")


class SettingsScreen(BaseScreen): # Экран глобальных настроек приложения WELL CONTROL. Наследует BaseScreen для получения адаптивного фона.
    def __init__(self, **kw):
        super().__init__(**kw) # Инициализируем базовые параметры экрана Kivy
        self.current_version = "1.1.0" # Текущая установленная инженерная версия программного обеспечения
        self.update_url = "https://github.com" # Ссылка на репозиторий для проверки и скачивания свежих сборок APK
        
        # Главный вертикальный контейнер экрана настроек, занимающий 100% ширины и высоты
        layout = MDBoxLayout(orientation='vertical', size_hint=(1, 1)) 
        
        # 1. ШАПКА ЭКРАНА (Верхний Тулбар)
        self.toolbar = MDBoxLayout(
            orientation='horizontal', # Элементы внутри шапки укладываются слева направо
            size_hint_y=None, # Отключаем авто-высоту
            height=dp(45), # Задаем фиксированную высоту панели навигации в 45 пикселей
            md_bg_color=(0.2, 0.25, 0.3, 1), # Спокойный графитово-серый цвет шапки, отличный от расчетных экранов
            padding=[dp(15), 0, dp(10), 0] # Внутренние отступы элементов от краев панели
        )
        
        # Левая управляющая кнопка-иконка «Назад» (стрелочка)
        self.toolbar.add_widget(MDIconButton(
            icon="arrow-left", # Встроенная системная иконка Material Design
            theme_text_color="Custom", # Разрешаем ручную настройку цвета
            text_color=(1, 1, 1, 1), # Белый цвет стрелочки
            pos_hint={'center_y': .5}, # Центрируем иконку строго посередине высоты тулбара
            # При нажатии мгновенно переключаем ScreenManager назад на экран главного меню ('menu')
            on_release=lambda x: setattr(self.manager, 'current', 'menu') 
        ))
        
        # Текстовый заголовок экрана настроек на тулбаре
        self.toolbar.add_widget(MDLabel(
            text="НАСТРОЙКИ ПРИЛОЖЕНИЯ", # Название экрана
            theme_text_color="Custom", 
            text_color=(1, 1, 1, 1), # Чистый белый цвет текста
            font_style="H6" # Крупный и четкий стиль шрифта под заголовки
        ))
        layout.add_widget(self.toolbar) # Добавляем готовую шапку в самый верх главного макета настроек
        
        # 2. СКРОЛЛ КОНТЕНТА 
        # Создаем контейнер прокрутки, чтобы длинный список коэффициентов не обрезался на маленьких смартфонах
        scroll = ScrollView() 
        # Вертикальный бокс внутри скролла, автоматически подстраивающийся под высоту всех полей ввода
        container = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(12), padding=dp(20)) 
        
        # Текстовая метка подзаголовка для секции нормативных коэффициентов Ростехнадзора
        container.add_widget(MDLabel(text="КОЭФФИЦИЕНТЫ ЗАПАСА Ростехнадзора (Кз):", font_style="Caption", theme_text_color="Hint")) 
        
        # Поля ввода для Кз 
        # Генерируем компактные поля для ручной калибровки коэффициентов репрессии пласта по трем категориям опасности
        self.f_k_z1 = self.create_setting_field("Кз для 1 КАТЕГОРИИ (Газ/АВПД)", "1.10") # Коэффициент запаса для 1-й категории газоопасности
        self.f_k_z2 = self.create_setting_field("Кз для 2 КАТЕГОРИИ (Нефтяные)", "1.07") # Коэффициент запаса для 2-й категории
        self.f_k_z3 = self.create_setting_field("Кз для 3 КАТЕГОРИИ (Обычные)", "1.05") # Коэффициент запаса для 3-й категории (базовый)
        
        container.add_widget(self.f_k_z1) # Выводим поле Кз 1-й категории в контейнер скролла
        container.add_widget(self.f_k_z2) # Выводим поле Кз 2-й категории в контейнер скролла
        container.add_widget(self.f_k_z3) # Выводим поле Кз 3-й категории в контейнер скролла
        
        # Текстовая метка подзаголовка для секции аварийных технологических запасов и регламентных допусков КРС
        container.add_widget(MDLabel(text="ТЕХНОЛОГИЧЕСКИЕ РЕГЛАМЕНТЫ:", font_style="Caption", theme_text_color="Hint")) 
        
        # Поля ввода для нормативных запасов 
        self.f_v_zapas = self.create_setting_field("Аварийный запас объема на мерник (м³)", "3.0") # Сургутский нормативный избыток объема жидкости глушения
        self.f_v_crit_spo = self.create_setting_field("Допуск на перелив при СПО (м³)", "0.5") # Критический объем разности доливов мерников ЦС
        
        container.add_widget(self.f_v_zapas) # Выводим поле аварийного запаса объема в контейнер скролла
        container.add_widget(self.f_v_crit_spo) # Выводим поле критического допуска СПО в контейнер скролла

        
        # 3. СТАТУС И ОБНОВЛЕНИЕ БАЗ ДАННЫХ EXCEL С ТВОЕГО GITHUB 
        # Текстовый заголовок для секции управления таблицами сортамента НКТ и параметров ЦА-320
        container.add_widget(MDLabel(text="УПРАВЛЕНИЕ ИНЖЕНЕРНЫМИ БАЗАМИ ДАННЫХ:", font_style="Caption", theme_text_color="Hint")) 
        
        # Системная метка для вывода статуса проверки файлов Excel (актуальны, устарели, скачиваются)
        self.db_status_label = MDLabel(text="Базы данных: опрос...", theme_text_color="Secondary", font_style="Body2", size_hint_y=None, height=dp(25)) 
        container.add_widget(self.db_status_label) 
        
        # Якорный макет для фиксации и центрирования кнопки сетевого обновления баз
        anchor_db_update = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(48)) 
        # Кнопка для запуска парсера и скачивания свежих .xlsx файлов с удаленного репозитория GitHub
        self.btn_db_update = MDFillRoundFlatIconButton( 
            text="ОБНОВИТЬ ИНЖЕНЕРНЫЕ БАЗЫ С ГИТХАБ", 
            icon="database-refresh", 
            size_hint=(1, 1), 
            md_bg_color=(0.1, 0.4, 0.8, 1), 
            on_release=self.download_fresh_databases 
        ) 
        self.btn_db_update.size_hint_max_x = dp(2000) # Ограничитель максимальной ширины кнопки
        anchor_db_update.add_widget(self.btn_db_update) 
        container.add_widget(anchor_db_update) 

        # 4. ОБНОВЛЕНИЕ СИСТЕМЫ (ПОИСК НОВЫХ APK) 
        # Метка, выводящая текущую жестко зашитую версию сборки приложения
        container.add_widget(MDLabel(text=f"ВЕРСИЯ ПРИЛОЖЕНИЯ: v{self.current_version}", font_style="Caption", theme_text_color="Hint")) 
        
        # Центровщик для кнопки проверки обновлений самого приложения
        anchor_update = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(48)) 
        # Кнопка для опроса сервера на наличие новых скомпилированных версий APK-пакета
        self.btn_update_apk = MDFillRoundFlatIconButton( 
            text="ПРОВЕРИТЬ ОБНОВЛЕНИЯ", 
            icon="cloud-download-outline", 
            size_hint=(1, 1), 
            md_bg_color=(0.1, 0.4, 0.8, 1), 
            on_release=self.check_for_updates 
        ) 
        self.btn_update_apk.size_hint_max_x = dp(2000) 
        anchor_update.add_widget(self.btn_update_apk) 
        container.add_widget(anchor_update) 
        
        container.add_widget(MDBoxLayout(size_hint_y=None, height=dp(10))) # Пустая техническая распорка в 10 пикселей

        # ТЕХНОЛОГИЧЕСКИЙ ФИКС: Кнопка вызова справочника бригад 
        # Подзаголовок для административной секции персонала КРС
        container.add_widget(MDLabel(text="ПЕРСОНАЛ И СВЯЗЬ:", font_style="Caption", theme_text_color="Hint")) 
        
        # Центровщик для кнопки открытия базы данных бригад
        anchor_brigades = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(48)) 
        # Кнопка, вызывающая модальное диалоговое окно со списком бригад ТКРС и их контактами
        self.btn_brigades_book = MDFillRoundFlatIconButton( 
            text="ОТКРЫТЬ СПРАВОЧНИК БРИГАД", 
            icon="account-group-outline", 
            size_hint=(1, 1), 
            md_bg_color=(0.1, 0.4, 0.8, 1), 
            on_release=self.open_brigades_dialog 
        ) 
        self.btn_brigades_book.size_hint_max_x = dp(2000) 
        anchor_brigades.add_widget(self.btn_brigades_book) 
        container.add_widget(anchor_brigades) 

        # --- ТЕХНОЛОГИЧЕСКИЙ ИНТЕРФЕЙСНЫЙ ВЫБОР: МАСШТАБ КНОПОК И ШРИФТОВ ГЛАВНОЙ СТРАНИЦЫ ---
        container.add_widget(MDLabel(text="МАСШТАБ ИНТЕРФЕЙСА ГЛАВНОГО МЕНЮ:", font_style="Caption", theme_text_color="Hint"))
        
        # Горизонтальный ряд для трех кнопок выбора размеров
        scale_row = MDBoxLayout(orientation='horizontal', spacing=dp(6), size_hint_y=None, height=dp(40))
        
        # Функция связывается с ядром, чтобы узнать, какой масштаб сохранен сейчас в системе
        app = MDApp.get_running_app()
        current_scale = getattr(app, 'menu_scale', 1.0)
        
        # Вычисляем цвета кнопок: активный режим красим в оранжевый, остальные — в базовый синий
        c_small = (0.9, 0.4, 0.0, 1) if current_scale == 1.0 else (0.1, 0.4, 0.8, 1)
        c_med   = (0.9, 0.4, 0.0, 1) if current_scale == 1.5 else (0.1, 0.4, 0.8, 1)
        c_large = (0.9, 0.4, 0.0, 1) if current_scale == 2.0 else (0.1, 0.4, 0.8, 1)
        
        # Создаем три кнопки масштаба (Мелкий / Средний / Крупный) с автоматическим вызовом метода изменения
        self.btn_s_small = MDFillRoundFlatButton(text="МЕЛКИЙ", size_hint=(1, 1), md_bg_color=c_small, on_release=lambda x: self.change_local_scale(1.0))
        self.btn_s_med   = MDFillRoundFlatButton(text="СРЕДНИЙ", size_hint=(1, 1), md_bg_color=c_med, on_release=lambda x: self.change_local_scale(1.5))
        self.btn_s_large = MDFillRoundFlatButton(text="КРУПНЫЙ", size_hint=(1, 1), md_bg_color=c_large, on_release=lambda x: self.change_local_scale(2.0))
        
        # Вкладываем кнопки в горизонтальный ряд
        scale_row.add_widget(self.btn_s_small)
        scale_row.add_widget(self.btn_s_med)
        scale_row.add_widget(self.btn_s_large)
        container.add_widget(scale_row) # Выводим ряд масштабирования в контент настроек
        
        container.add_widget(MDBoxLayout(size_hint_y=None, height=dp(10))) #  зазор перед финальной кнопкой

        # 5. КНОПКА СОХРАНЕНИЯ (СТЕНА) 
        # Якорный макет для фиксации финишной кнопки сохранения параметров
        anchor_save = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(55)) 
        # Зеленая кнопка, собирающая все коэффициенты и записывающая их в постоянную память JSON
        self.btn_save_config = MDFillRoundFlatIconButton( 
            text="СОХРАНИТЬ КОНФИГУРАЦИЮ", 
            icon="content-save-outline", 
            size_hint=(1, 1), 
            md_bg_color=(0.1, 0.6, 0.2, 1), 
            on_release=self.save_settings 
        ) 
        self.btn_save_config.size_hint_max_x = dp(2000) 
        anchor_save.add_widget(self.btn_save_config) 
        container.add_widget(anchor_save) # Помещаем зеленую кнопку в самый низ списка

        scroll.add_widget(container) # Вкладываем заполненную вертикальную коробку параметров в скролл-область
        layout.add_widget(scroll) # Вкладываем скролл под верхний графический тулбар
        self.add_widget(layout) # Выводим полностью сформированный интерфейс настроек на экран дисплея

    def create_setting_field(self, hint, def_val): 
        """Вспомогательный метод: генерирует компактные текстовые поля для ввода коэффициентов"""
        f = MDTextField(
            hint_text=hint, # Текст-подсказка внутри рамки поля
            text=def_val, # Предустановленное дефолтное значение параметра
            mode="fill", 
            size_hint_y=None, 
            height=dp(34), # Фиксированная компактная высота поля в 34 пикселя
            input_filter='float' # Блокировка клавиатуры: разрешен ввод только цифр и точек (дробные float)
        )
        # Автоматическое выделение всего текста в поле за 0.1 секунды при клике для удобного переписывания данных
        f.bind(focus=lambda ins, val: Clock.schedule_once(lambda dt: ins.select_all(), 0.1) if val else None) 
        return f 

    def on_enter(self): 
        """Автоматическое событие Kivy: подгружает все сохраненные коэффициенты из JSON при открытии экрана настроек"""
        store = JsonStore(os.path.join(BASE_DIR, 'settings.json')) # Открываем локальный файл конфигурации JSON
        
        # Если в базе есть ветка инженерных коэффициентов, вычитываем их в текстовые поля интерфейса
        if store.exists('engineering_specs'): 
            data = store.get('engineering_specs') 
            self.f_k_z1.text = str(data.get('kz1', '1.10')) # Кз 1 категории
            self.f_k_z2.text = str(data.get('kz2', '1.07')) # Кз 2 категории
            self.f_k_z3.text = str(data.get('kz3', '1.05')) # Кз 3 категории
            self.f_v_zapas.text = str(data.get('v_zapas', '3.0')) # Аварийный избыток объема (+3 м3)
            self.f_v_crit_spo.text = str(data.get('v_crit_spo', '0.5')) # Критический перелив при СПО (0.5 м3)
            
        # --- ТЕХНОЛОГИЧЕСКИЙ ФИКС: СИНХРОНИЗАЦИЯ ПОДСВЕТКИ КНОПОК МАСШТАБА ШРИФТА ПРИ ВХОДЕ ---
        app = MDApp.get_running_app()
        # Вычитываем актуальный масштаб из ядра. Если его там еще нет — по умолчанию берем 1.0 (мелкий)
        current_scale = getattr(app, 'menu_scale', 1.0)
        # Динамически перекрашиваем кнопки: выбранный масштаб становится оранжевым, остальные — сине-голубыми
        if hasattr(self, 'btn_s_small') and self.btn_s_small:
            self.btn_s_small.md_bg_color = (0.9, 0.4, 0.0, 1) if current_scale == 1.0 else (0.1, 0.4, 0.8, 1)
            self.btn_s_med.md_bg_color   = (0.9, 0.4, 0.0, 1) if current_scale == 1.5 else (0.1, 0.4, 0.8, 1)
            self.btn_s_large.md_bg_color = (0.9, 0.4, 0.0, 1) if current_scale == 2.0 else (0.1, 0.4, 0.8, 1)

        # Опрашиваем физическое наличие всех трех файлов Excel в памяти устройства 
        nkt_ok = os.path.exists(EXCEL_PATH) # Проверяем файл базы труб НКТ
        vzd_ok = os.path.exists(VZD_PATH) # Проверяем каталог гидравлических двигателей ВЗД
        ca_ok = os.path.exists(os.path.join(BASE_DIR, 'data', 'ca320_stats.xlsx')) # Проверяем спецификации агрегата ЦА-320
        
        # Формируем итоговую текстовую строку статуса с цветовой разметкой BB-кодов (зеленый ОК / красный НЕТ)
        status_txt = ( 
            f"НКТ: {'[color=#33FF33]ОК[/color]' if nkt_ok else '[color=#FF3333]НЕТ[/color]'} | " 
            f"ВЗД: {'[color=#33FF33]ОК[/color]' if vzd_ok else '[color=#FF3333]НЕТ[/color]'} | " 
            f"ЦА-320: {'[color=#33FF33]ОК[/color]' if ca_ok else '[color=#FF3333]НЕТ[/color]'}" 
        ) 
        self.db_status_label.text = status_txt # Записываем строку статуса в интерфейсную метку
        self.db_status_label.markup = True # Разрешаем KivyMD обрабатывать цветные теги внутри строки

    def change_local_scale(self, scale_val):
        """Инженерный метод: обрабатывает клик по кнопкам шрифта, меняет масштаб в ядре и мгновенно перекрашивает кнопки"""
        app = MDApp.get_running_app()
        app.menu_scale = scale_val # Перезаписываем глобальный масштаб интерфейса главной страницы в ядре приложения
        
        # Сохраняем выбранный масштаб в локальный файл настроек JSON, чтобы телефон запомнил его навсегда
        app.store.put('menu_settings', scale=scale_val)
        
        # Мгновенно обновляем цвета кнопок прямо перед глазами пользователя на экране Настроек
        self.btn_s_small.md_bg_color = (0.9, 0.4, 0.0, 1) if scale_val == 1.0 else (0.1, 0.4, 0.8, 1)
        self.btn_s_med.md_bg_color   = (0.9, 0.4, 0.0, 1) if scale_val == 1.5 else (0.1, 0.4, 0.8, 1)
        self.btn_s_large.md_bg_color = (0.9, 0.4, 0.0, 1) if scale_val == 2.0 else (0.1, 0.4, 0.8, 1)

    def download_fresh_databases(self, *args): 
        """Инициирует процесс сетевого обновления баз данных сортамента с удаленного сервера"""
        # Фоновое скачивание, чтобы интерфейс приложения не зависал намертво во время сетевого запроса
        self.btn_db_update.text = "СКАЧИВАНИЕ ТАБЛИЦ С ГИТХАБ..." # Меняем надпись на кнопке, информируя инженера
        self.btn_db_update.disabled = True # Временно блокируем кнопку от повторных случайных кликов
        
        import threading # Импортируем модуль многопоточности Python
        # Запускаем тяжелый сетевой процесс скачивания в отдельном фоновом потоке, освобождая основной UI-поток графики
        threading.Thread(target=self._run_download).start() 

    def _run_download(self): 
        """Внутренний изолированный поток: поочередно скачивает три инженерные таблицы Excel по протоколу HTTP"""
        import urllib.request # Импортируем системную библиотеку выполнения сетевых URL-запросов
        try: 
            # 1. Загрузка базы данных марок и характеристик труб НКТ
            urllib.request.urlretrieve( 
                "https://githubusercontent.com", # Ссылка-источник на вашем удаленном GitHub
                EXCEL_PATH # Локальный путь сохранения файла на флеш-памяти телефона
            ) 
            # 2. Загрузка технического каталога винтовых забойных двигателей ВЗД
            urllib.request.urlretrieve( 
                "https://githubusercontent.com", 
                VZD_PATH 
            ) 
            # 3. Загрузка калибровочной статистики теоретических подач насоса ЦА-320
            ca_path = os.path.join(BASE_DIR, 'data', 'ca320_stats.xlsx') 
            urllib.request.urlretrieve( 
                "https://githubusercontent.com", 
                ca_path 
            ) 
            # При успешном завершении всех загрузок передаем управление в главный графический поток через Clock
            Clock.schedule_once(lambda dt: self._download_success(), 0) 
        except Exception as e: 
            # ФИКС NAMEERROR: жестко переводим ошибку в текст до того, как Python сотрет переменную 'e' при выходе из потока
            err_msg = str(e) 
            # Передаем текст ошибки в главный поток для вывода всплывающего уведомления
            Clock.schedule_once(lambda dt: self._download_error(err_msg), 0)

    def open_brigades_dialog(self, *args):
        """Инженерный метод: мгновенно переключает интерфейс на полноценный экран списка бригад"""
        self.manager.current = "open_brigades_screen" # Безопасный переход к новому экрану

        # Контейнер для отображения списка бригад
        self.dialog_box = MDBoxLayout(
            orientation="vertical", # Элементы внутри бокса (список, кнопка добавления) укладываются строго вертикально
            spacing=dp(10), # Вертикальный зазор между списком и кнопкой в 10 адаптивных пикселей
            adaptive_height=True, # Позволяет контейнеру автоматически растягиваться по высоте вложенного содержимого
            padding=[0, dp(10), 0, 0] # Отступ только сверху в 10 пикселей, чтобы текст не прилипал к шапке диалога
        )
        
        # Скролл, чтобы список не вылезал за границы экрана телефона
        list_scroll = ScrollView(
            size_hint_y=None, 
            height=dp(200) # Ограничиваем зону просмотра бригад фиксированной высотой в 200 пикселей для компактности в окне
        )
        
        # Внутренний вертикальный контейнер скролла, куда будут физически вставляться строки с персоналом
        self.brigades_list_container = MDBoxLayout(orientation="vertical", adaptive_height=True)
        
        # Вытаскиваем сохраненные бригады из SQLite
        conn = get_db_connection() # Устанавливаем соединение с локальной базой данных SQLite
        cursor = conn.cursor()
        # Выгружаем данные персонала (номер бригады, ФИО мастера, внутренний и мобильный телефоны) со сквозной сортировкой по номерам
        cursor.execute("SELECT number, master, phone_int, phone_mob FROM brigades ORDER BY number ASC")
        rows = cursor.fetchall() # Забираем все строки из таблицы brigades
        conn.close() # Обязательно закрываем сессию базы данных, освобождая файл
        
        if not rows:
            # Если в базе данных еще нет записей о персонале КРС, генерируем информирующую заглушку
            from kivymd.uix.label import MDLabel
            empty_lbl = MDLabel(
                text="Справочник пуст. Добавьте первую бригаду.", # Текст подсказки бурильщику
                halign="center", # Выравнивание надписи строго по центру окна
                theme_text_color="Hint", # Системный приглушенный серый цвет шрифта
                size_hint_y=None, 
                height=dp(40), # Фиксированная высота заглушки в 40 пикселей
                font_style="Body2" # Компактный аккуратный стиль шрифта
            )
            self.brigades_list_container.add_widget(empty_lbl) # Выводим заглушку на экран диалога
        else:
            # ТЕХНОЛОГИЧЕСКИЙ ФИКС: Корректно читаем кортеж базы по индексам [0], [1], [2], [3]
            # Запускаем цикл перебора строк и формирования карточек персонала для каждой бригады
            for row in rows:
                b_num = str(row[0]) # Номер бригады КРС/ТКРС (Индекс 0)
                b_master = str(row[1]) # ФИО мастера бригады (Индекс 1)
                
                # Формируем префиксы связи: если номера в ячейках SQLite существуют, подставляем условные маркеры промысла
                p_int = f"вн. {row[2]}" if row[2] else "" # Внутренний короткий АТС-номер (Индекс 2)
                p_mob = f"сот. {row[3]}" if row[3] else "" # Сотовый / спутниковый мобильный номер (Индекс 3)
                
                # Фильтруем пустые значения и склеиваем доступные контакты в одну красивую строку через разделитель «вертикальная черта»
                phones = " | ".join([p for p in [p_int, p_mob] if p])
                
                # Инициализируем двухстрочный графический элемент списка Material Design
                item = TwoLineListItem(
                    text=f"Бригада №{b_num} — мастер {b_master}", # Верхняя основная строка: номер и ФИО руководящего ИТР
                    secondary_text=phones if phones else "Контакты не указаны", # Нижняя строка: собранные номера связи
                    theme_text_color="Custom", # Разрешаем ручной выбор палитры для главного текста
                    text_color=(1, 1, 1, 1), # Чистый белый цвет букв для идеальной читаемости на темном фоне
                    secondary_theme_text_color="Custom", # Разрешаем ручной выбор палитры для подписи контактов
                    secondary_text_color=(0.9, 0.4, 0.0, 1) # Фирменный ярко-оранжевый цвет для выделения телефонов мастера
                )
                self.brigades_list_container.add_widget(item) # Добавляем готовую информационную строчку в общий список контейнера
                
        list_scroll.add_widget(self.brigades_list_container) # Вкладываем заполненный список внутрь зоны прокрутки ScrollView
        self.dialog_box.add_widget(list_scroll) # Помещаем скролл в главный вертикальный макет диалога
        
        # Широкая кнопка добавления новой бригады под списком
        # Создаем якорную разметку для фиксации геометрии нижней кнопки триггера 
        add_btn_anchor = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(45)) 
        
        # Сама кнопка, переводящая пользователя на экран создания нового персонала в базе данных SQLite
        add_btn = MDRaisedButton( 
            text="➕ ДОБАВИТЬ НОВУЮ БРИГАДУ", # Текст кнопки 
            md_bg_color=(0.1, 0.4, 0.8, 1), # Сине-голубой цвет кнопки, стандартный для сервисных действий WELL CONTROL
            
            # ТЕХНОЛОГИЧЕСКИЙ ФИКС: Сначала принудительно закрываем (dismiss) висящее модальное окно списка,
            # и только после этого плавно переключаем ScreenManager на наш новый чистый экран ввода!
            on_release=lambda x: (
                self.br_dialog.dismiss() if hasattr(self, 'br_dialog') and self.br_dialog else None,
                setattr(self.manager, 'current', 'create_brigade_screen')
            )
        ) 
        add_btn_anchor.add_widget(add_btn) # Вкладываем кнопку в центровщик-якорь 
        self.dialog_box.add_widget(add_btn_anchor) # Помещаем узел кнопки в самый низ вертикального макета диалогового окна


         # Открываем чистое окно просмотра 
        # Инициализируем кастомное диалоговое окно MDDialog для отображения выгруженного списка бригад
        self.br_dialog = MDDialog( 
            title="СПРАВОЧНИК БРИГАД", # Главный заголовок окна
            type="custom", # Включаем кастомный тип окна, позволяющий передать сложный макет content_cls
            content_cls=self.dialog_box, # Передаем собранный ранее вертикальный контейнер со скроллом и кнопкой
            size_hint_x=0.9, # Ширина всплывающего окна занимает ровно 90% от ширины экрана смартфона
            auto_dismiss=True, # Разрешаем автоматическое скрытие окна при клике пользователя мимо него
            buttons=[ 
                # Добавляем нижнюю кнопку «ЗАКРЫТЬ» для принудительного сворачивания диалога
                MDFlatButton(text="ЗАКРЫТЬ", on_release=lambda x: self.br_dialog.dismiss()) 
            ] 
        ) 
        self.br_dialog.open() # Физически выводим окно Справочника бригад поверх текущего интерфейса настроек



    def _download_success(self): 
        """Событие Kivy: срабатывает в главном потоке при успешном окончании скачивания всех таблиц Excel с GitHub"""
        self.btn_db_update.text = "БАЗЫ ДАННЫХ УСПЕШНО ОБНОВЛЕНЫ!" # Выводим информирующий статус-успех на кнопку
        self.btn_db_update.md_bg_color = (0.1, 0.6, 0.2, 1) # Перекрашиваем кнопку в сигнальный зеленый цвет успеха
        self.btn_db_update.disabled = False # Снимаем блокировку, возвращая кнопке кликабельность
        self.on_enter() # Принудительно перезапускаем метод on_enter, чтобы мгновенно обновить статус-маркеры (ОК) на экране настроек
        
        def restore_db_btn(dt):
            # Внутренняя функция-таймер для возвращения кнопке исходного текстового состояния
            self.btn_db_update.text = "ОБНОВИТЬ ИНЖЕНЕРНЫЕ БАЗЫ С ГИТХАБ"
            self.btn_db_update.md_bg_color = (0.1, 0.4, 0.8, 1) # Возвращаем стандартный сине-голубой цвет
            
        Clock.schedule_once(restore_db_btn, 3.0) # Запускаем отложенный таймер: кнопка вернет свой исходный вид ровно через 3 секунды

    def _download_error(self, err_msg): 
        """Событие Kivy: срабатывает в главном потоке, если сетевой поток скачивания файлов Excel завершился аварией"""
        self.btn_db_update.text = "ОШИБКА ЗАГРУЗКИ (ПРОВЕРЬ СВЯЗЬ)" # Выводим предупреждение о проблемах с интернетом на промысле
        self.btn_db_update.md_bg_color = (0.85, 0.15, 0.15, 1) # Окрашиваем кнопку в ярко-красный аварийный цвет ошибки
        self.btn_db_update.disabled = False # Возвращаем кнопке кликабельность для повторных попыток запроса
        
        def restore_db_btn(dt):
            # Внутренняя функция-таймер для сброса аварийного состояния кнопки
            self.btn_db_update.text = "ОБНОВИТЬ ИНЖЕНЕРНЫЕ БАЗЫ С ГИТХАБ"
            self.btn_db_update.md_bg_color = (0.1, 0.4, 0.8, 1)
            
        Clock.schedule_once(restore_db_btn, 4.0) # Запускаем таймер: аварийная надпись пропадет с экрана через 4 секунды
        print(f"Update error: {err_msg}") # Выводим детальный текст технического сбоя в системную консоль/терминал

    def check_for_updates(self, *args): 
        """Интерактивный метод: перенаправляет инженера на удаленный веб-репозиторий для проверки свежих версий APK"""
        self.btn_update_apk.text = "ПЕРЕХОД НА СТРАНИЦУ ЗАГРУЗКИ..." # Меняем надпись, давая мгновенный визуальный отклик на клик
        self.btn_update_apk.md_bg_color = (0.1, 0.5, 0.8, 1) # Включаем приятный синий цвет перехода
        
        def restore_update_btn(dt):
            # Функция возвращает кнопке обновлений стандартный вид
            self.btn_update_apk.text = "ПРОВЕРИТЬ ОБНОВЛЕНИЯ"
            self.btn_update_apk.md_bg_color = (0.1, 0.4, 0.8, 1)
            
        Clock.schedule_once(restore_update_btn, 3.0) # Запускаем таймер сброса состояния кнопки на 3 секунды
        
        try: 
            import webbrowser # Импортируем системный модуль для работы с веб-браузерами операционной системы
            webbrowser.open(self.update_url) # Инициируем команду ОС на открытие интернет-ссылки GitHub в стандартном браузере телефона
        except Exception as e: 
            print(f"Browser error: {e}") # Защита от сбоя, если на устройстве физически заблокированы или отсутствуют веб-браузеры

    def save_settings(self, *args): 
        """Инженерный метод: собирает веденные технологические коэффициенты из полей экрана и сохраняет их в постоянную память JSON"""
        store = JsonStore(os.path.join(BASE_DIR, 'settings.json')) # Открываем дескриптор локального файла постоянного хранения настроек
        
        # Намертво записываем в конфигурационный JSON-файл новые калибровочные коэффициенты Ростехнадзора и регламентные допуски
        store.put(
            'engineering_specs', 
            kz1=float(self.f_k_z1.text or 1.10), # Коэффициент запаса 1 категории
            kz2=float(self.f_k_z2.text or 1.07), # Коэффициент запаса 2 категории
            kz3=float(self.f_k_z3.text or 1.05), # Коэффициент запаса 3 категории
            v_zapas=float(self.f_v_zapas.text or 3.0), # Аварийный нормативный избыток жидкости глушения
            v_crit_spo=float(self.f_v_crit_spo.text or 0.5) # Критический перелив мерников при СПО
        ) 
        
        # СИНХРОНИЗАЦИЯ С КАЛЬКУЛЯТОРОМ: Находим расчетный экран с именем 'calc'
        calc_scr = self.manager.get_screen('calc') 
        # Если у калькулятора есть метод refresh_ui_elements, принудительно вызываем его для мгновенного применения новых Кз
        if hasattr(calc_scr, 'refresh_ui_elements'): 
            calc_scr.refresh_ui_elements() 
            
        # Интерактивный отклик на кнопке сохранения 
        self.btn_save_config.text = "КОНФИГУРАЦИЯ УСПЕШНО СОХРАНЕНА И ПРИМЕНЕНА!" # Информируем бурильщика об успешной записи параметров
        self.btn_save_config.md_bg_color = (0.1, 0.5, 0.8, 1) # Временно перекрашиваем кнопку сохранения в синий цвет фиксации


        def restore_btn(dt):
            # Внутренняя функция возвращает кнопке сохранения базовый зеленый цвет
            self.btn_save_config.text = "СОХРАНИТЬ КОНФИГУРАЦИЮ"
            self.btn_save_config.md_bg_color = (0.1, 0.6, 0.2, 1) # Возвращаем стандартный зеленый цвет ("Все сохранено")
            
        Clock.schedule_once(restore_btn, 2.5) # Запускаем отложенный таймер возврата внешнего вида кнопки на 2.5 секунды
        print("Конфигурация успешно сохранена и применена!") # Дублируем отладочную запись в консоль разработчика


class CementAcidScreen(BaseScreen):
    
    def __init__(self, **kw):
        super().__init__(**kw)
        self.v_pipe = 3.02 
        self.mode = "CEMENT"
        
        main_layout = MDBoxLayout(orientation='vertical', size_hint=(1, 1))

        # 1. ШАПКА
        self.toolbar = MDBoxLayout(
            orientation='horizontal', size_hint_y=None, height=dp(45),
            md_bg_color=(0.1, 0.4, 0.8, 1), padding=[dp(15), 0, dp(10), 0]
        )
        btn_back = MDIconButton(
            icon="arrow-left",
            theme_text_color="Custom",
            text_color=(1, 1, 1, 1),
            # Фикс: убрали 'x', теперь лямбда вызывается чисто и не ломает метод go_back
           on_release=lambda *args: setattr(self.manager, 'current', 'menu') if self.manager else None

        )
        self.toolbar.add_widget(btn_back)
        self.toolbar.add_widget(MDLabel(text="ЦМОСТ / ВАННЫ", theme_text_color="Custom", 
                                       text_color=(1,1,1,1), font_style="H6"))
        main_layout.add_widget(self.toolbar)

        # 2. ОСНОВНОЙ КОНТЕЙНЕР СО СКРОЛЛОМ
        scroll = ScrollView()
        self.container = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(10), padding=dp(15))

        # ТАБЫ (ОСТАЮТСЯ СВЕРХУ)
        tab_grid = MDGridLayout(cols=2, spacing=dp(5), size_hint_y=None, height=dp(45))
        self.b1 = MDFillRoundFlatButton(text="ЦЕМЕНТ", size_hint=(1,1), on_release=lambda x: self.switch_mode("CEMENT"))
        self.b2 = MDFillRoundFlatButton(text="ВАННА", size_hint=(1,1), on_release=lambda x: self.switch_mode("ACID"))
        tab_grid.add_widget(self.b1); tab_grid.add_widget(self.b2)
        self.container.add_widget(tab_grid)

        # ПОЛЯ ВВОДА
        self.inputs_box = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(8))

        def create_f(hint, text=""):
            f = MDTextField(hint_text=hint, text=text, mode="fill", font_size="15sp", 
                               size_hint_y=None, height=dp(34), input_filter='float')
            f.bind(focus=lambda ins, val: Clock.schedule_once(lambda dt: ins.select_all(), 0.1) if val else None)
            return f

        self.f_diam = create_f("Ø экспл. колонны (мм)", "168")
        self.f_depth = create_f("Глубина воронки (м)", "1500")
        self.f_target = create_f("Интервал установки (м)", "1550")
        self.f_vol = create_f("Объем смеси (м³)", "1.5")
        
        for f in [self.f_diam, self.f_depth, self.f_target, self.f_vol]: 
            self.inputs_box.add_widget(f)
        self.container.add_widget(self.inputs_box)

        # ОКНО РЕЗУЛЬТАТА
        self.res_label = MDLabel(
            text="Введите данные и выберите НКТ", 
            theme_text_color="Custom", text_color=(1,1,1,1), 
            markup=True, size_hint_y=None, line_height=1.3
        )
        self.res_label.bind(texture_size=lambda ins, sz: setattr(ins, 'height', sz[1] + dp(20)))
        self.container.add_widget(self.res_label)

        # 3. НИЖНИЙ БЛОК КНОПОК (ВЫБОР НКТ И РАССЧИТАТЬ)
        # Оборачиваем в AnchorLayout для принудительной растяжки на весь экран
        
        # КНОПКА ВЫБОРА НКТ
        anchor_nkt = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(52))
        self.btn_nkt = MDFillRoundFlatButton(
            text="ВЫБРАТЬ НКТ (л/м)", 
            size_hint=(1, 1), md_bg_color=(0.1, 0.3, 0.6, 1), 
            on_release=self.open_nkt_menu
        )
        self.btn_nkt.size_hint_max_x = dp(2000)
        anchor_nkt.add_widget(self.btn_nkt)
        self.container.add_widget(anchor_nkt)

        # КНОПКА РАСЧЕТА
        anchor_calc = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(55))
        self.btn_calc = MDFillRoundFlatIconButton(
            text="РАССЧИТАТЬ ТЕХНОЛОГИЮ", icon="calculator", 
            size_hint=(1, 1), md_bg_color=(0.1, 0.6, 0.2, 1), 
            on_release=self.do_calc
        )
        self.btn_calc.size_hint_max_x = dp(2000)
        anchor_calc.add_widget(self.btn_calc)
        self.container.add_widget(anchor_calc)

        scroll.add_widget(self.container)
        main_layout.add_widget(scroll)
        self.add_widget(main_layout)
        self.switch_mode("CEMENT")

    def switch_mode(self, mode):
        self.mode = mode
        active, inactive = (0.1, 0.4, 0.8, 1), (0.2, 0.2, 0.2, 1)
        self.b1.md_bg_color = active if mode == "CEMENT" else inactive
        self.b2.md_bg_color = active if mode == "ACID" else inactive
        self.f_vol.hint_text = "Объем цемента (м³)" if mode == "CEMENT" else "Объем кислоты (м³)"

    def open_nkt_menu(self, b):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True); ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
            items = [{"viewclass":"OneLineListItem","text":f"НКТ {r['Тип_НКТ']}","on_release":lambda x=r:self.set_nkt(x)} for r in rows]
            self.menu = MDDropdownMenu(caller=b, items=items, width_mult=4); self.menu.open()
        except: self.res_label.text = "[color=#FF5555]Ошибка базы Excel[/color]"

    def set_nkt(self, r):
        d_in = float(r.get('Внутр_d_мм', 62))
        self.v_pipe = (d_in**2) * 0.0007854 * 1000
        self.btn_nkt.text = f"НКТ {r['Тип_НКТ']} ({self.v_pipe:.2f} л/м)"
        self.menu.dismiss()

    def do_calc(self, *args):
        try:
            def val(f): return float(f.text or 0)
            D, Hv, H_tar, V_sm = val(self.f_diam), val(self.f_depth), val(self.f_target), val(self.f_vol)
            
            # Геометрия (л/м)
            v_col_litr = ((D-15)**2) * 0.0007854 * 1000
            # Объем НКТ (сколько влезет в трубы до низа)
            v_nkt_total = (self.v_pipe * Hv) / 1000
            # Объем продавки (сколько качать в трубы, чтобы вытолкнуть смесь ниже воронки)
            h_diff = max(0, H_tar - Hv)
            v_prodavka = (v_col_litr * h_diff) / 1000

            if self.mode == "CEMENT":
                bags = int((V_sm * 1000) / 36)
                res = (
                    f"[size=18sp][color=#33FF33]ИНСТРУКЦИЯ: ЦЕМЕНТНЫЙ МОСТ[/color][/size]\n\n"
                    f"1. [b]Закачка:[/b] Закачай в трубы [b]{V_sm} м³[/b] цемента.\n"
                    f"   (Это примерно [b]{bags} мешков[/b] по 50кг).\n"
                    f"2. [b]Продавка:[/b] Качай следом воду/раствор в объеме [b]{v_nkt_total:.2f} м³[/b].\n"
                    f"   [i]*Важно: Не больше и не меньше, чтобы цемент вышел из труб.*[/i]\n"
                    f"3. [b]Подъем:[/b] Сразу подними трубы выше [b]{H_tar - 50} м[/b].\n"
                    f"4. [b]Срезка:[/b] Качай раствор в затруб (обратка), чтобы вымыть лишний цемент."
                )
            else:
                res = (
                    f"[size=18sp][color=#33FF33]ИНСТРУКЦИЯ: КИСЛОТНАЯ ВАННА[/color][/size]\n\n"
                    f"1. [b]Закачка:[/b] Закачай в трубы [b]{V_sm} м³[/b] кислоты.\n"
                    f"2. [b]Доводка:[/b] Качай раствор в трубы [b]{v_nkt_total:.2f} м³[/b].\n"
                    f"   [b]ЗАТРУБ ДОЛЖЕН БЫТЬ ОТКРЫТ![/b] (Ждем выхода воздуха/жидкости).\n"
                    f"3. [b]Продавка:[/b] Как только докачал объем выше, [color=#FF5555][b]ЗАКРОЙ ЗАТРУБ![/b][/color]\n"
                    f"4. [b]Давим в пласт:[/b] Качай еще [b]{v_prodavka:.2f} м³[/b] в трубы.\n"
                    f"   (Кислота уйдет из под воронки прямо в зону перфорации).\n"
                    f"5. [b]Финиш:[/b] Закрой задвижки на трубах. Ожидание реакции 2-4 часа."
                )
            self.res_label.text = res
        except: 
            self.res_label.text = "[color=#FF5555]ОШИБКА: Проверь, все ли цифры ввел![/color]"
        def go_back(self, *args):
          if self.manager:
            self.manager.current = "menu"
class CalculatorHubScreen(BaseScreen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.selected_weight_m = 9.5
        self.pipe_name = "НКТ 73"
        self.tank_shape = "RECT"

        main_layout = MDBoxLayout(orientation='vertical', size_hint=(1, 1))
        
        # 1. ШАПКА
        self.toolbar = MDBoxLayout(
            orientation='horizontal', size_hint_y=None, height=dp(50),
            md_bg_color=(0.1, 0.4, 0.8, 1), padding=[dp(15), 0, dp(10), 0]
        )
        from kivymd.uix.button import MDIconButton
        btn_back = MDIconButton(
            icon="arrow-left", theme_text_color="Custom", text_color=(1, 1, 1, 1),
            on_release=lambda x: setattr(self.manager, 'current', 'menu')
        )
        self.toolbar.add_widget(btn_back)
        self.toolbar.add_widget(MDLabel(text="ИНЖЕНЕРНЫЕ РАСЧЕТЫ", theme_text_color="Custom", 
                                       text_color=(1,1,1,1), font_style="H6"))
        main_layout.add_widget(self.toolbar)

        scroll = ScrollView()
        self.container = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(15), padding=dp(20))

        # --- РАЗДЕЛ 1: РАСЧЕТ ВЕСА (ИВЭ) ---
        self.container.add_widget(MDLabel(text="ВЕС ИНСТРУМЕНТА С ПЛАВУЧЕСТЬЮ:", font_style="Button", theme_text_color="Hint"))
        
        # Кнопка выбора трубы на всю ширину
        anchor_pipe = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(50))
        self.btn_pipe = MDFillRoundFlatButton(text=f"ТРУБА: {self.pipe_name}", size_hint=(1, 1), md_bg_color=(0.1, 0.3, 0.6, 1), on_release=self.menu_pipes)
        self.btn_pipe.size_hint_max_x = dp(2000)
        anchor_pipe.add_widget(self.btn_pipe)
        self.container.add_widget(anchor_pipe)

        self.f_depth = self.create_small_f("Глубина спуска (м)")
        self.f_rho_fluid = self.create_small_f("Плотность раствора (г/см³)", "1.00")
        self.container.add_widget(self.f_depth)
        self.container.add_widget(self.f_rho_fluid)

        # ТАБЛО ИВЭ (Крупно, симметрично)
        self.ive_display = MDGridLayout(
            cols=2, size_hint_y=None, height=dp(85), 
            padding=[dp(10), dp(10)], spacing=dp(10), md_bg_color=(0, 0, 0, 1)
        )
        
        # Блок "В ВОЗДУХЕ"
        box_v = MDBoxLayout(orientation='vertical', spacing=dp(2))
        box_v.add_widget(MDLabel(text="В ВОЗДУХЕ", halign="center", font_style="Caption", theme_text_color="Hint"))
        self.res_air_val = MDLabel(text="0.0", halign="center", font_style="H5", theme_text_color="Custom", text_color=(1,1,1,1), bold=True)
        box_v.add_widget(self.res_air_val)
        
        # Блок "НА КРЮКЕ"
        box_k = MDBoxLayout(orientation='vertical', spacing=dp(2))
        box_k.add_widget(MDLabel(text="НА КРЮКЕ", halign="center", font_style="Caption", theme_text_color="Hint"))
        self.res_hook_val = MDLabel(text="0.0", halign="center", font_style="H4", theme_text_color="Custom", text_color=(0.2, 1, 0.2, 1), bold=True)
        box_k.add_widget(self.res_hook_val)
        
        self.ive_display.add_widget(box_v)
        self.ive_display.add_widget(box_k)
        self.container.add_widget(self.ive_display)
        
        # Кнопка расчета на всю ширину
        anchor_calc_w = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(52))
        btn_calc_w = MDFillRoundFlatIconButton(text="РАССЧИТАТЬ ВЕС", icon="weight-lifter", size_hint=(1, 1), on_release=self.calc_weight)
        btn_calc_w.size_hint_max_x = dp(2000)
        anchor_calc_w.add_widget(btn_calc_w)
        self.container.add_widget(anchor_calc_w)

        # --- РАЗДЕЛ 2: СМЕШИВАНИЕ ---
        self.container.add_widget(self.add_line())
        self.container.add_widget(MDLabel(text="СМЕШИВАНИЕ ДВУХ РАСТВОРОВ:", font_style="Button", theme_text_color="Hint"))
        
        self.f_v1 = self.create_small_f("Объем 1 (м³)"); self.f_r1 = self.create_small_f("Ро 1 (г/см³)")
        self.f_v2 = self.create_small_f("Объем 2 (м³)"); self.f_r2 = self.create_small_f("Ро 2 (г/см³)")
        self.container.add_widget(self.f_v1); self.container.add_widget(self.f_r1)
        self.container.add_widget(self.f_v2); self.container.add_widget(self.f_r2)
        
        anchor_mix = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(50))
        btn_mix = MDFillRoundFlatButton(text="РАССЧИТАТЬ СМЕСЬ", size_hint=(1, 1), md_bg_color=(0.1, 0.6, 0.2, 1), on_release=self.calc_mixing)
        btn_mix.size_hint_max_x = dp(2000)
        anchor_mix.add_widget(btn_mix)
        self.container.add_widget(anchor_mix)

        self.res_mix = MDLabel(text="Итог: введите данные", halign="left", theme_text_color="Custom", text_color=(1,1,1,1), markup=True)
        self.container.add_widget(self.res_mix)

        # --- РАЗДЕЛ 3: ОБЪЕМ ЕМКОСТЕЙ ---
        self.container.add_widget(self.add_line())
        self.container.add_widget(MDLabel(text="ОБЪЕМ ЖИДКОСТИ В ЕМКОСТИ:", font_style="Button", theme_text_color="Hint"))
        
        shape_grid = MDBoxLayout(orientation='horizontal', spacing=dp(5), size_hint_y=None, height=dp(45))
        self.btn_rect = MDFillRoundFlatButton(text="ПРЯМОУГОЛЬНАЯ", size_hint=(1,1), on_release=lambda x: self.set_shape("RECT"))
        self.btn_cyl = MDFillRoundFlatButton(text="ЦИЛИНДР (ГОР.)", size_hint=(1,1), on_release=lambda x: self.set_shape("CYL"))
        shape_grid.add_widget(self.btn_rect); shape_grid.add_widget(self.btn_cyl)
        self.container.add_widget(shape_grid)

        self.f_t_l = self.create_small_f("Длина емкости (м)")
        self.f_t_w_d = self.create_small_f("Ширина емкости (м)")
        self.f_t_h = self.create_small_f("Уровень взлива (м)")
        self.container.add_widget(self.f_t_l); self.container.add_widget(self.f_t_w_d); self.container.add_widget(self.f_t_h)
        
        anchor_tank = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(50))
        btn_tank = MDFillRoundFlatButton(text="ПОСЧИТАТЬ ОБЪЕМ", size_hint=(1, 1), md_bg_color=(0.1, 0.3, 0.6, 1), on_release=self.calc_tank_vol)
        btn_tank.size_hint_max_x = dp(2000)
        anchor_tank.add_widget(btn_tank)
        self.container.add_widget(anchor_tank)

        self.res_tank = MDLabel(text="Объем: --- м³", halign="left", bold=True)
        self.container.add_widget(self.res_tank)

        self.container.add_widget(MDBoxLayout(size_hint_y=None, height=dp(40))) 
        scroll.add_widget(self.container);
        main_layout.add_widget(scroll); 
        self.add_widget(main_layout)
        self.set_shape("RECT")

    # --- ВСПОМОГАТЕЛЬНЫЕ МЕТОДЫ ---
    def create_small_f(self, hint, text=""):
        f = MDTextField(hint_text=hint, text=text, mode="fill", size_hint_y=None, height=dp(40), input_filter='float')
        f.bind(focus=lambda ins, val: Clock.schedule_once(lambda dt: ins.select_all(), 0.1) if val else None)
        return f

    def add_line(self):
        return MDBoxLayout(size_hint_y=None, height=dp(1), md_bg_color=(0.3, 0.3, 0.3, 1))

    def set_shape(self, shape):
        self.tank_shape = shape
        active, inactive = (0.1, 0.4, 0.8, 1), (0.2, 0.2, 0.2, 1)
        self.btn_rect.md_bg_color = active if shape == "RECT" else inactive
        self.btn_cyl.md_bg_color = active if shape == "CYL" else inactive
        self.f_t_w_d.hint_text = "Ширина (м)" if shape == "RECT" else "Диаметр (м)"

    def menu_pipes(self, b):
        # Список кортежей: (Отображаемый текст, Чистый вес для расчетов)
        list_p = [
            ("НКТ 60 (7.1 кг/м)", 7.1), 
            ("НКТ 73 (9.5 кг/м)", 9.5), 
            ("НКТ 89 (13.5 кг/м)", 13.5), 
            ("СБТ 73 (14.5 кг/м)", 14.5)
        ]
        # Важно: передаем в lambda весь кортеж x
        items = [{
            "viewclass": "OneLineListItem", 
            "text": i[0], 
            "on_release": lambda x=i: self.set_pipe(x)
        } for i in list_p]
        
        self.m_p = MDDropdownMenu(caller=b, items=items, width_mult=4)
        self.m_p.open()

    def set_pipe(self, val):
        # val[0] - это текст (напр. "НКТ 73 (9.5 кг/м)")
        # val[1] - это число (9.5)
        self.pipe_name = val[0].split(" (")[0]
        self.selected_weight_m = val[1]
        
        self.btn_pipe.text = f"ТРУБА: {self.pipe_name}"
        self.m_p.dismiss()
        # Сразу пересчитываем, если глубина уже введена
        self.calc_weight()

    def calc_weight(self, *args):
        try:
            L = float(self.f_depth.text or 0)
            rho_f = float(self.f_rho_fluid.text or 1.0)
            w_air = (L * self.selected_weight_m) / 1000
            w_hook = w_air * (1 - rho_f / 7.85)
            self.res_air_val.text = f"{w_air:.1f} т"
            self.res_hook_val.text = f"{w_hook:.1f} т"
            # Цвет красный при превышении 60 тонн
            self.res_hook_val.text_color = (1, 0.2, 0.2, 1) if w_hook > 60 else (0.2, 1, 0.2, 1)
        except: pass

    def calc_mixing(self, *args):
        try:
            v1, r1 = float(self.f_v1.text or 0), float(self.f_r1.text or 0)
            v2, r2 = float(self.f_v2.text or 0), float(self.f_r2.text or 0)
            if (v1 + v2) > 0:
                res = (v1 * r1 + v2 * r2) / (v1 + v2)
                self.res_mix.text = f"• Плотность смеси: [b]{res:.2f} г/см³[/b]\n• Общий объем: [b]{v1+v2:.1f} м³[/b]"
        except: self.res_mix.text = "Ошибка данных"

    def calc_tank_vol(self, *args):
        try:
            L, WD, H = float(self.f_t_l.text or 0), float(self.f_t_w_d.text or 0), float(self.f_t_h.text or 0)
            if self.tank_shape == "RECT":
                vol = L * WD * H
            else:
                vol = (math.pi * (WD/2)**2 * L) * (H/WD)
            self.res_tank.text = f"Объем жидкости: [b]{vol:.2f} м³[/b]"
        except: pass

class GnvpScreen(BaseScreen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.method = "DRILLER"
        self.v_vnutr_nkt = 3.02 
        
        main_layout = MDBoxLayout(orientation='vertical', size_hint=(1, 1))
        
        # 1. ШАПКА (Оранжевый — цвет опасности)
        self.toolbar = MDBoxLayout(
            orientation='horizontal', size_hint_y=None, height=dp(50),
            md_bg_color=(0.8, 0.4, 0, 1), padding=[dp(15), 0, dp(10), 0]
        )
        from kivymd.uix.button import MDIconButton
        btn_back = MDIconButton(
            icon="arrow-left", theme_text_color="Custom", text_color=(1, 1, 1, 1),
            on_release=lambda x: setattr(self.manager, 'current', 'menu')
        )
        self.toolbar.add_widget(btn_back)
        self.toolbar.add_widget(MDLabel(text="ЛИСТ ГЛУШЕНИЯ (ГНВП)", theme_text_color="Custom", 
                                       text_color=(1,1,1,1), font_style="H6"))
        main_layout.add_widget(self.toolbar)

        scroll = ScrollView()
        self.container = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(10), padding=dp(20))

        # 2. ИНТЕРАКТИВНЫЕ ОКНА (ПОЛЯ ВВОДА)
        def f_make(hint, val=""):
            f = MDTextField(hint_text=hint, text=val, mode="fill", input_filter='float', size_hint_y=None, height=dp(34))
            f.bind(focus=lambda ins, v: Clock.schedule_once(lambda dt: ins.select_all(), 0.1) if v else None)
            return f

        self.f_d_col = f_make("Ø колонны (мм)", "168")
        self.f_ckod = f_make("ЦКОД / Иск. забой (м)", "2500")
        self.f_h_hvost = f_make("Длина хвостовика 114 (м)", "0")
        self.f_h_v = f_make("Глубина воронки (м)", "2000")
        self.f_h_perf = self.create_small_field("Интервал перфорации (кровля, м)", "2150")
        self.f_h_tvd = f_make("Глубина по вертикали (м)", "1800")
        self.f_sidpp = f_make("Р-трубное SIDPP (атм)", "0")
        self.f_sicp = f_make("Р-затрубное SICP (атм)", "15")
        self.f_rho_old = f_make("Ро раствора в скв. (г/см³)", "1.10")

        fields = [self.f_d_col, self.f_ckod, self.f_h_hvost, self.f_h_v, 
                  self.f_h_tvd, self.f_sidpp, self.f_sicp, self.f_rho_old]
        for f in fields: 
            self.container.add_widget(f)

        # 3. ВЫБОР НКТ (РАСПОЛОЖЕН ПОД ОКНАМИ)
        anchor_nkt = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(50))
        self.btn_nkt = MDFillRoundFlatButton(
            text="ВЫБРАТЬ ТИП НКТ ПОДВЕСКИ", 
            size_hint=(1, 1), md_bg_color=(0.1, 0.3, 0.6, 1), 
            on_release=self.open_nkt_menu
        )
        self.btn_nkt.size_hint_max_x = dp(2000)
        anchor_nkt.add_widget(self.btn_nkt)
        self.container.add_widget(anchor_nkt)

        # 4. ВЫБОР ТАКТИКИ (КНОПКИ НА ШИРИНУ ЭКРАНА)
        self.container.add_widget(MDLabel(text="ТАКТИКА ГЛУШЕНИЯ:", font_style="Caption", theme_text_color="Hint"))
        
        def create_wide_mode_btn(text, mode):
            anchor = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(45))
            btn = MDFillRoundFlatButton(text=text, size_hint=(1, 1), on_release=lambda x: self.set_m(mode))
            btn.size_hint_max_x = dp(2000)
            anchor.add_widget(btn)
            return anchor, btn

        self.anc1, self.m1 = create_wide_mode_btn("МЕТОД БУРИЛЬЩИКА (2 ЦИКЛА)", "DRILLER")
        self.anc2, self.m2 = create_wide_mode_btn("ОЖИДАНИЕ И УТЯЖЕЛЕНИЕ (1 ЦИКЛ)", "WAIT")
        self.anc3, self.m3 = create_wide_mode_btn("ЗАДАВКА (BULLHEADING)", "BULL")
        
        for a in [self.anc1, self.anc2, self.anc3]: self.container.add_widget(a)

        # 5. КНОПКА РАСЧЕТА (НА ВСЮ ШИРИНУ)
        anchor_calc = MDAnchorLayout(anchor_x='center', size_hint_y=None, height=dp(55))
        self.btn_calc = MDFillRoundFlatIconButton(
            text="РАССЧИТАТЬ ИНСТРУКЦИЮ", icon="alert-octagon", 
            size_hint=(1, 1), md_bg_color=(0.8, 0, 0, 1), 
            on_release=self.calculate
        )
        self.btn_calc.size_hint_max_x = dp(2000)
        anchor_calc.add_widget(self.btn_calc)
        self.container.add_widget(anchor_calc)

        # 6. ВЫВОД РЕЗУЛЬТАТОВ
        self.res = MDLabel(text="Введите данные замеров ГК", markup=True, halign="center", 
                          theme_text_color="Custom", text_color=(1,1,1,1), size_hint_y=None)
        self.res.bind(texture_size=lambda ins, sz: setattr(ins, 'height', sz[1] + dp(20)))
        self.container.add_widget(self.res)

        scroll.add_widget(self.container)
        main_layout.add_widget(scroll)
        self.add_widget(main_layout)
        self.set_m("DRILLER")

    def set_m(self, mode):
        self.method = mode
        active, inactive = (0.1, 0.4, 0.8, 1), (0.2, 0.2, 0.2, 1)
        self.m1.md_bg_color = active if mode == "DRILLER" else inactive
        self.m2.md_bg_color = active if mode == "WAIT" else inactive
        self.m3.md_bg_color = active if mode == "BULL" else inactive

    def open_nkt_menu(self, b):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True); ws = wb.active
            head = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            data = [dict(zip(head, r)) for r in ws.iter_rows(min_row=2, values_only=True) if any(r)]
            items = [{"viewclass": "OneLineListItem", "text": f"НКТ {r['Тип_НКТ']}", "on_release": lambda x=r: self.set_nkt(x)} for r in data]
            self.m_nkt = MDDropdownMenu(caller=b, items=items, width_mult=4, background_color=(0.1, 0.1, 0.15, 1))
            self.m_nkt.open()
        except: self.res.text = "[color=#FF5555]Ошибка базы данных[/color]"

    def set_nkt(self, r):
        d_in = float(r.get('Внутр_d_мм', 62))
        self.v_vnutr_nkt = (d_in**2) * 0.0007854 * 1000
        self.btn_nkt.text = f"НКТ {r['Тип_НКТ']} ({self.v_vnutr_nkt:.2f} л/м)"
        self.m_nkt.dismiss()

    def calculate(self, *args):
        try:
            def val(f): return float(f.text or 0)
            D, L_ckod, L_hvo, Hv = val(self.f_d_col), val(self.f_ckod), val(self.f_h_hvost), val(self.f_h_v)
            H_tvd, pt, pz, ro = val(self.f_h_tvd), val(self.f_sidpp), val(self.f_sicp), val(self.f_rho_old)
            
            h_ref = H_tvd if H_tvd > 0 else L_ckod
            if h_ref == 0: 
                self.res.text = "[color=#FFCC00]Укажите глубину![/color]"; return
            
            # 1. РАСЧЕТ ПЛОТНОСТИ (Запас +0.05 г/см3 по ПБ 08-624-03)
            # Если трубное давление (pt) равно 0, значит скважина стоит под давлением затруба (pz)
            p_izb = pt if pt > 0 else pz
            ro_k = round(ro + (p_izb * 10.197 / h_ref) + 0.05, 2)
            p_plas = int(h_ref * (ro + (pt * 10.197 / h_ref)) / 10.197)

            # 2. ГЕОМЕТРИЯ (литры на метр)
            v_vnutr_litr = self.v_vnutr_nkt  # Объем НКТ л/м
            v_col_litr = ((D - 15)**2) * 0.0007854 * 1000 # Объем колонны л/м
            
            # Объемы в м3
            v_nkt = round((v_vnutr_litr * Hv) / 1000, 2)
            v_annulus = round(((v_col_litr * Hv) / 1000) - (v_nkt * 0.35), 2) # Затруб
            v_pod = round((v_col_litr * (L_ckod - Hv)) / 1000, 2) # Под воронкой
            v_well_total = round(v_nkt + v_annulus + v_pod, 2)

            # --- ИНСТРУКЦИИ ПО РЕГЛАМЕНТУ ---
            if self.method == "DRILLER":
                tech = (
                    f"[color=#33FF33][b]МЕТОД БУРИЛЬЩИКА (2 ЦИКЛА):[/b][/color]\n"
                    f"• [b]1-й цикл:[/b] Вымыв пачки старым раствором.\n"
                    f"  Качай в затруб. Держи Р_затр = [b]{pz} атм[/b].\n"
                    f"  Пачка выйдет на устье через [color=#FFFF33][b]{v_nkt} м³[/b][/color] (объем НКТ).\n"
                    f"• [b]2-й цикл:[/b] Замещение на Ро = [b]{ro_k}[/b].\n"
                    f"  Качай в затруб. Давление Р_затр снижать до 0."
                )
            elif self.method == "WAIT":
                tech = (
                    f"[color=#33FF33][b]ОЖИДАНИЕ И УТЯЖЕЛЕНИЕ:[/b][/color]\n"
                    f"1. Приготовь [b]{v_well_total} м³[/b] раствора Ро = [b]{ro_k}[/b].\n"
                    f"2. Закачка в НКТ. Снижай Р_труб с [b]{pt}[/b] до [b]0[/b] атм\n"
                    f"   пропорционально объему [b]{v_nkt} м³[/b].\n"
                    f"3. При выходе из воронки держи Р_затр = [b]{pz} атм[/b]."
                )
            else: # BULLHEADING
                tech = (
                    f"[color=#FF5555][b]ЗАДАВКА (BULLHEADING):[/b][/color]\n"
                    f"1. [b]Глушение без циркуляции.[/b] Закрой трубное.\n"
                    f"2. Качай Ро = [b]{ro_k}[/b] в затрубное пространство.\n"
                    f"3. [b]ВНИМАНИЕ:[/b] Контроль Р_заб. Не превышать Р_погл.\n"
                    f"   При наличии МГРП: Р_max = [b]Р_откр - 20 атм[/b]."
                )

            self.res.text = (
                f"[size=20sp]ПЛОТНОСТЬ ГЛУШЕНИЯ: [b]{ro_k}[/b][/size]\n"
                f"Пластовое давление: [b]{p_plas} атм[/b]\n"
                f"Общий объем скважины: [b]{v_well_total} м³[/b]\n"
                f"--------------------------------\n"
                f"{tech}\n"
                f"--------------------------------\n"
                f"[i]Контроль КВД в течение 2-х часов после закачки.[/i]"
            )
        except Exception as e: 
            self.res.text = f"[color=#FF5555]Ошибка в расчетах[/color]"

class KillResultScreen(BaseScreen):  #Результаты глушения. Показывает итоговые цифры и пошаговую тактическую инструкцию.
    def __init__(self, **kw):
        super().__init__(**kw)
        layout = MDBoxLayout(orientation='vertical')
        
        # 1. ШАПКА
        self.toolbar = MDBoxLayout(
            orientation='horizontal', size_hint_y=None, height=dp(45),
            md_bg_color=(0.1, 0.4, 0.8, 1), padding=[dp(15), 0, dp(10), 0]
        )
        from kivymd.uix.button import MDIconButton
        btn_back = MDIconButton(
            icon="arrow-left", pos_hint={'center_y': .5},
            theme_text_color="Custom", text_color=(1, 1, 1, 1),
            on_release=lambda x: self.go_back()
        )
        self.toolbar.add_widget(btn_back)
        self.toolbar.add_widget(MDLabel(text="РЕЗУЛЬТАТЫ РАСЧЕТА", theme_text_color="Custom", 
                                       text_color=(1,1,1,1), font_style="H6"))
        layout.add_widget(self.toolbar)

        # 2. ОБЛАСТЬ ВЫВОДА
        scroll = ScrollView()
        self.result_box = MDBoxLayout(orientation='vertical', adaptive_height=True, padding=dp(20), spacing=dp(15))
        
        # КАРТОЧКА С ОСНОВНЫМИ ПАРАМЕТРАМИ
        self.res_card = MDCard(
            orientation='vertical', padding=dp(15), spacing=dp(5),
            size_hint=(1, None), adaptive_height=True,
            md_bg_color=(0.15, 0.2, 0.3, 1), radius=[dp(10)]
        )
        self.result_text = MDLabel(
            text="Данные рассчитываются...", theme_text_color="Custom", text_color=(1, 1, 1, 1),
            font_style="Subtitle1", size_hint_y=None, markup=True, line_height=1.4
        )
        self.result_text.bind(texture_size=lambda ins, sz: setattr(ins, 'height', sz[1]))
        self.res_card.add_widget(self.result_text)
        self.result_box.add_widget(self.res_card)

        # ИНСТРУКЦИЯ (Тактика действий)
        self.result_box.add_widget(MDLabel(text="ПОРЯДОК ДЕЙСТВИЙ:", font_style="Caption", theme_text_color="Hint"))
        
        self.instruction_text = MDLabel(
            text="Инструкция формируется...", theme_text_color="Primary",
            font_style="Body2", size_hint_y=None, markup=True, line_height=1.3
        )
        self.instruction_text.bind(texture_size=lambda ins, sz: setattr(ins, 'height', sz[1] + dp(10)))
        self.result_box.add_widget(self.instruction_text)

        # 3. КНОПКА ПЕРЕХОДА К СИМУЛЯЦИИ
        self.btn_view_scheme = MDFillRoundFlatIconButton(
            text="ВИЗУАЛИЗАЦИЯ СКВАЖИНЫ", icon="eye-outline",
            size_hint_x=None, width=dp(280), pos_hint={'center_x': 0.5},
            md_bg_color=(0.1, 0.6, 0.2, 1), on_release=self.go_to_scheme
        )
        self.result_box.add_widget(self.btn_view_scheme)

        scroll.add_widget(self.result_box)
        layout.add_widget(scroll)
        self.add_widget(layout)

    def go_to_scheme(self, x):
        self.manager.current = 'well_view'

    def go_back(self):
        # Возвращаемся в калькулятор для правки данных
        self.manager.current = 'calc'

class MenuCard(MDCard):
    def __init__(self, text, icon, color, screen_name, **kwargs):
        super().__init__(**kwargs)
        self.size_hint = (None, None)
        self.radius = [dp(12)]
        
        # Строка строго горизонтальная
        self.orientation = "horizontal"
        self.md_bg_color = color
        self.ripple_behavior = True
        self.target = screen_name
        
        # Компактные отступы: слева под иконку, сверху/снизу минимум для узкой строки
        self.padding = [dp(10), dp(4), dp(16), dp(4)]
        
        # 1. ИКОНКА СЛЕВА (Компактная, центрированная)
        self.icon_widget = MDIconButton(
            icon=icon,
            theme_icon_color="Custom",
            icon_color=(1, 1, 1, 1),
            icon_size=dp(24), # Узкий аккуратный размер
            size_hint=(None, 1),
            width=dp(40),
            pos_hint={"center_y": .5}
        )
        self.icon_widget.disabled = True
        self.add_widget(self.icon_widget)

        # 2. ТЕКСТ (Идёт сразу за иконкой, выровнен по левому краю)
        self.label_widget = MDLabel(
            text=text.replace('\n', ' '), # Убираем переносы, пишем в одну строчку
            halign="left",
            valign="center",
            theme_text_color="Custom",
            text_color=(1, 1, 1, 1),
            bold=True,
            font_style="Body2", # Собранный шрифт для узкой строки
            size_hint=(1, 1)
        )
        self.add_widget(self.label_widget)

    def on_release(self):
        app = MDApp.get_running_app()
        if hasattr(app, 'sm'):
            app.sm.current = self.target

class MainMenuScreen(MDScreen): # Класс главного экрана приложения, унаследованный от MDScreen (основа навигации)
    def __init__(self, **kw):
        super().__init__(**kw) # Инициализируем базовые графические свойства родительского класса Kivy
        self.root_ui = FloatLayout() # Создаем корневой слой со свободной (относительной) привязкой виджетов по координатам
        
        # 1. ФОН (Картинка вышки) 
        # Добавляем фоновую картинку вышки, растянутую по всему экрану (FitImage) с приглушенной матовой прозрачностью 45%
        self.root_ui.add_widget(FitImage(source=BACKGROUND_PATH, opacity=0.45)) 
        
        # 2. ЗАГОЛОВОК И СЛОГАН (ЦЕНТРИРОВАНИЕ) 
        # Крупное текстовое название приложения WELL CONTROL с черной обводкой букв толщиной 2 пикселя для контраста на любом фоне
        self.title_label = Label( 
            text="WELL CONTROL", # Название программы
            bold=True, # Жирный шрифт
            color=(1, 1, 1, 1), # Белый цвет букв
            outline_color=(0, 0, 0, 1), # Черный контур обводки
            outline_width=2, # Толщина обводки
            size_hint=(None, None), # Отключаем автоматическое растягивание границ текстового блока
            halign='center', # Горизонтальное центрирование текста внутри блока
            pos_hint={'center_x': 0.5, 'top': 0.97} # Размещаем ровно по центру экрана по оси X и прижимаем к самому верху (97% высоты)
        ) 
        
        # Инженерный слоган приложения, оформленный с легкой черной обводкой в 1 пиксель
        self.subtitle_label = Label( 
            text="«Твой расчет — твоя безопасность»", # Текст слогана
            color=(1, 1, 1, 1), 
            outline_color=(0, 0, 0, 1), 
            outline_width=1, 
            size_hint=(None, None), 
            halign='center', 
            pos_hint={'center_x': 0.5, 'top': 0.88} # Смещаем под главный заголовок (на уровень 88% высоты экрана)
        ) 
        
        # 3. ЛОГОТИП ПОЛНОСТЬЮ УДАЛЕН ИЗ ВЕРСТКИ ДЛЯ КРАСОТЫ 
        
        # 4. ВЕРТИКАЛЬНЫЙ СКРОЛЛ ДЛЯ КНОПОК-СТРОК (Поднят выше, так как место освободилось) 
        # Создаем контейнер прокрутки, позволяющий плавно листать список кнопок-модулей пальцем
        self.scroll = ScrollView( 
            size_hint=(1, None), # Занимает 100% ширины экрана, высота задается жестко или настраивается динамически
            pos_hint={'center_x': 0.5, 'top': 0.82}, # Скролл красиво поджат к слогану (начинается на уровне 82% высоты экрана)
            do_scroll_x=False, # Полностью блокируем горизонтальную прокрутку, чтобы меню не уезжало влево-вправо
            do_scroll_y=True # Разрешаем вертикальную прокрутку (вверх и вниз)
        ) 
        
        # Вертикальный бокс для кнопок 
        # Контейнер внутри скролла, который автоматически увеличивает свою общую высоту по мере добавления новых кнопок
        self.menu_box = MDBoxLayout( 
            orientation='vertical', # Элементы укладываются строго друг под друга по вертикали
            spacing=dp(8), # Зазор (вертикальный интервал) между соседними кнопками меню в 8 пикселей
            padding=[dp(20), dp(5), dp(20), dp(5)], # Боковые отступы меню: слева и справа по 20dp, сверху и снизу по 5dp
            adaptive_height=True # Разрешаем контейнеру динамически менять высоту по суммарному объему вложенных виджетов
        ) 
        
        # Определение цветовой палитры для дизайна главного меню (цвета с альфа-каналом прозрачности 55%)
        c_blue = (0.1, 0.4, 0.8, 0.55) # Профессиональный синий полупрозрачный оттенок для основных технологических расчетов
        c_grey = (0.2, 0.25, 0.3, 0.55) # Спокойный графитово-серый полупрозрачный оттенок для справочных и сервисных модулей
        
        # Ровно 8 рабочих кнопок 
        # Массив данных (Спецификация), содержащий название, иконку Material Design, цвет и имя экрана для перехода
        grid_items = [ 
            ("ГЛУШЕНИЕ", "water-alert", c_blue, "calc"), # Модуль технологических расчетов глушения
            ("КОНТРОЛЬ ДОЛИВА", "water-pump", c_blue, "spo"), # Расчеты объемов СПО и доливов скважины
            ("ЦМОСТ / ВАННЫ", "format-color-fill", c_blue, "cmoist"), # Калькулятор цементных мостов и нефтяных/кислотных ванн
            ("ЛОВИЛЬНЫЕ РАБОТЫ", "hook", c_blue, "fishing"), # Инструментальный справочник и расчеты ловильных работ в КРС
            ("ИНЖЕНЕРНЫЙ ХАБ", "calculator", c_blue, "calc_hub"), # Узел дополнительных гидравлических калькуляторов
            ("ЖУРНАЛ СКВАЖИН", "notebook", c_blue, "notebook_main_menu"), # Блокнот (база данных SQLite) сохраненных скважин бригад
            ("ИНСТРУМЕНТ (СПРАВКА)", "book-open-variant", c_grey, "threads"), # Каталог-справочник эксплуатационных характеристик резьб НКТ
            ("НАСТРОЙКИ ПРИЛОЖЕНИЯ", "cog", c_grey, "settings") # Экран внутренних настроек интерфейса и профиля
        ] 
        # Цикл генерации: перебираем массив спецификаций кнопок и создаем для каждой объект MenuCard
        for t, i, c, s in grid_items: 
            self.menu_box.add_widget(MenuCard(t, i, c, s)) # Физически вкладываем созданную кнопку в вертикальный бокс
            
        self.scroll.add_widget(self.menu_box) # Помещаем вертикальный бокс с кнопками внутрь зоны прокрутки ScrollView
        
        # Собираем интерфейс на главном экране (чистый заголовок и кнопки) 
        self.root_ui.add_widget(self.title_label) # Выводим на экран верхний главный текстовый логотип OIL MATE
        self.root_ui.add_widget(self.subtitle_label) # Выводим на экран текстовый инженерный слоган безопасности
        self.root_ui.add_widget(self.scroll) # Выводим на экран адаптивную зону прокрутки со списком кнопок-модулей
        self.add_widget(self.root_ui) # Добавляем полностью собранную FloatLayout-верстку в качестве корня экрана
        
        # Кроссплатформенный фикс размеров: привязываем событие изменения размеров окна (on_resize) к методу update_ui_sizes
        Window.bind(on_resize=self.update_ui_sizes) 
        self.update_ui_sizes() # Первичный вызов метода для принудительного расчета геометрии при самом первом запуске

    def on_enter(self):
        """Автоматическое событие Kivy: срабатывает каждый раз, когда инженер возвращается на главный экран (например, из Настроек)"""
        self.update_ui_sizes() # Принудительно перерисовываем размеры элементов под актуальный масштаб из настроек

    def update_ui_sizes(self, *args): 
        """Автоматический метод адаптивности: пересчитывает размеры текстов и кнопок с учетом выбранного в Настройках масштаба"""
        w, h = Window.width, Window.height # Фиксируем текущие физические ширину (w) и высоту (h) экрана смартфона или окна ПК
        app = MDApp.get_running_app() # Получаем оперативный доступ к центральному ядру приложения KrsBookApp
        
        # ДИНАМИЧЕСКИЙ ФИКС МАСШТАБА: Безопасно вытаскиваем масштаб из ядра. Если переменной еще нет в JSON — берем базовый 1.0 (мелкий)
        current_scale = getattr(app, 'menu_scale', 1.0)
        
        # Рассчитываем размер шрифта главного логотипа как 11% от ширины экрана, ограничивая его сверху лимитом в 42 пикселя
        self.title_label.font_size = min(w * 0.11, dp(42)) 
        # Размер самого текстового блока по ширине равен экрану, по высоте — в полтора раза больше размера букв
        self.title_label.size = (w, self.title_label.font_size * 1.5) 
        
        # Рассчитываем размер шрифта слогана как 3.8% от ширины экрана, ограничивая его сверху лимитом в 14 пикселей
        self.subtitle_label.font_size = min(w * 0.05, dp(20)) # Шрифт заголовка
        # Размер блока слогана по ширине равен экрану, по высоте — в два раза больше размера его букв
        self.subtitle_label.size = (w, self.subtitle_label.font_size * 2) 
        
        # Выделяем максимум места под список прокрутки 
        self.scroll.height = h * 0.72 # Зона прокрутки ScrollView жестко занимает 72% от всей доступной вертикали экрана
        
        # МАСШТАБИРУЕМАЯ ГЕОМЕТРИЯ КНОПОК МЕНЮ
        card_w = w - dp(40) # Ширина каждой кнопки-карточки рассчитывается как полная ширина экрана минус боковые отступы в 40dp
        
        # ИСПРАВЛЕНО: Умножаем базовую высоту плашки (46dp) на коэффициент масштаба (1.0, 1.5 или 2.0)
        card_h = dp(46 * current_scale) 
        
        # Цикл калибровки: пробегаемся по всем вложенным элементам внутри вертикального контейнера меню
        for card in self.menu_box.children: 
            if isinstance(card, MenuCard): # Отбираем только объекты кастомных кнопок-плашек MenuCard
                card.size_hint = (None, None) # Отключаем автоматическое пропорциональное растяжение Kivy по осям X и Y
                card.size = (card_w, card_h) # Принудительно задаем вычисленные ширину и новую адаптивную высоту кнопки
                card.radius = [dp(10)] # Настраиваем скругление углов плашки на 10 адаптивных пикселей для дизайна M3
                
                # Проверяем, содержит ли карточка внутренний текстовый виджет label_widget во избежание сбоя AttributeError
                if hasattr(card, 'label_widget') and card.label_widget: 
                    # ИСПРАВЛЕНО: Умножаем базовый размер шрифта (13sp) на коэффициент масштаба
                    card.label_widget.font_size = sp(13 * current_scale) 



# --- ГЛАВНОЕ ЯДРО --- 
class KrsBookApp(MDApp): 
    # Ядро. Запускает всё приложение, регистрирует экраны, настраивает тему и управляет системной кнопкой «Назад». 
    
    # Глобальная переменная для текстуры фона (чтобы не грузить её на каждом экране заново) 
    global_texture = ObjectProperty(None) # Специальное Kivy-свойство для кэширования графической подложки в ОЗУ

    def build(self): 
        # 1. Название приложения (должно совпадать с заголовком в buildozer.spec) 
        self.title = "WELL CONTROL" # ИСПРАВЛЕНО: Новый официальный тайтл приложения для релиза в RuStore

        
        # 2. ИСПРАВЛЕНИЕ БАГА ИКОНКИ: Жестко привязываем значок на первой строчке до старта графического окна
        if os.path.exists(ICON_PATH): 
            self.icon = ICON_PATH # Напрямую передаем путь к icon.png в дескриптор окна Windows
            
        # 3. Настройка темы KivyMD 
        self.theme_cls.theme_style = "Dark" # Включаем глобальную темную тему Material Design
        self.theme_cls.primary_palette = "Blue" # Задаем основной сине-голубой цвет по умолчанию для элементов палитры
        self.theme_cls.material_style = "M3" # Активируем современный дизайн-код Material You (M3) от Google
        
        # Настройка клавиатуры: при вводе текста экран сдвигается вверх 
        Window.softinput_mode = "pan" # Важнейший Android-фикс для работы нашего умного фикса в BaseScreen
        
        # --- ТЕХНОЛОГИЧЕСКИЙ ФИКС: ЗАГРУЗКА МАСШТАБА ШРИФТА ГЛАВНОГО МЕНЮ ИЗ JSON ---
        # Подключаемся к локальному файлу хранения настроек
        self.store = JsonStore(os.path.join(BASE_DIR, 'settings.json')) 
        # Проверяем, сохранял ли ранее бурильщик масштаб меню. Если да — считываем, если нет — ставим дефолтный 1.0 (мелкий)
        if self.store.exists('menu_settings'):
            self.menu_scale = float(self.store.get('menu_settings').get('scale', 1.0))
        else:
            self.menu_scale = 1.0 # Базовый масштаб по умолчанию

        # 4. МГНОВЕННЫЙ СТАРТ: Создаем только пустой менеджер и один экран загрузки
        self.sm = ScreenManager(transition=FadeTransition()) # Создаем корневой менеджер с эффектом плавного растворения окон
        
        # Мы создаем ТОЛЬКО экран загрузки. Он легкий и появится перед глазами инженера мгновенно без подвисаний
        self.sm.add_widget(LoadingScreen(name='loading'))
        
        return self.sm # Возвращаем менеджер, на экране сразу же загорается заставка с ниточкой прогресса


    def show_disclaimer(self): 
        """Окно предупреждения (Отказ от ответственности)""" 
        # Используем JsonStore для сохранения факта принятия условий 
        # Инициализируем локальное хранилище в файле settings.json в корне проекта
        self.store = JsonStore(os.path.join(BASE_DIR, 'settings.json')) 
        
        # Проверяем: если инженер уже запускал приложение ранее и принял условия, то окно больше не показываем
        if self.store.exists('disclaimer') and self.store.get('disclaimer')['accepted']: 
            return # Досрочно выходим из функции
            
        # Текст юридического и технологического предупреждения для инженеров и бурильщиков КРС
        disclaimer_text = ( 
            "Данное приложение является вспомогательным инженерным инструментом. " 
            "Все расчеты носят справочный характер.\n\n" 
            "Разработчик не несет ответственности за любые последствия (аварии, " 
            "материальный ущерб), возникшие в результате использования данных.\n\n" 
            "Сверяйте результаты с утвержденным Планом работ и РД." 
        ) 
        
        # Генерируем модальное окно MDDialog типа confirmation (подтверждение)
        self.dialog = MDDialog( 
            title="ОТКАЗ ОТ ОТВЕТСТВЕННОСТИ", # Крупный заголовок предупреждения
            text=disclaimer_text, # Основной текст юридической защиты
            type="confirmation", 
            auto_dismiss=False, # Блокировка: пользователь не может закрыть окно кликом мимо него, обязан нажать кнопку
            size_hint_x=0.85, # Ширина окна занимает 85% от ширины экрана смартфона
            buttons=[ 
                MDFlatButton( 
                    text="Я ПРИНИМАЮ", # Текст на управляющей кнопке
                    theme_text_color="Custom", 
                    text_color=self.theme_cls.primary_color, # Окрашиваем кнопку в основной синий цвет темы
                    on_release=self.close_disclaimer # При нажатии вызываем метод фиксации согласия
                ), 
            ], 
        ) 
        self.dialog.open() # Выводим окно дисклеймера на экран смартфона

    def close_disclaimer(self, *args): 
        """Записывает флаг согласия инженера в файл настроек JSON и закрывает модальное окно"""
        self.store.put('disclaimer', accepted=True) # Намертво сохраняем ключ 'disclaimer' в settings.json
        self.dialog.dismiss() # Скрываем окно предупреждения с экрана

    def on_start(self): 
        """События при старте приложения""" 
        # ИСПРАВЛЕНИЕ БАГА ИКОНКИ: Принудительно обновляем значок в момент, когда окно уже физически создано в Windows 
        if os.path.exists(ICON_PATH): 
            self.icon = ICON_PATH # Передаем картинку в кэш открытого дескриптора заголовка 
            
        # Привязываем физическую или системную кнопку 'Назад' 
        from kivy.base import EventLoop # Импортируем системный цикл обработки событий Kivy 
        # Перехватываем нажатия кнопок на клавиатуре ПК или системной сенсорной кнопки «Назад» на Android 
        EventLoop.window.bind(on_keyboard=self.hook_keyboard)


    def hook_keyboard(self, window, key, *args):
        # Код клавиши 'Назад' (ESC на ПК / Назад на Android) = 27
        if key == 27:
            current = self.sm.current  # Фиксируем имя экрана, на котором сейчас находится пользователь

            # Если мы на загрузке или в главном меню — закрываем приложение
            if current in ['loading', 'menu']:
                return False  # Возвращаем False, чтобы Kivy передал событие ОС Android и закрыл программу

            # НАВИГАЦИОННЫЙ ФИКС БЛОКНОТА, СПРАВОЧНИКОВ И ПЕРСОНАЛА:
            elif current == 'create_brigade_screen':
                self.sm.current = 'open_brigades_screen'  # Из добавления новой бригады возвращаем в список персонала
                return True

            elif current == 'open_brigades_screen':
                self.sm.current = 'settings'  # Из списка персонала возвращаем назад на экран Настроек приложения
                return True

            elif current == 'edit_well_screen':
                self.sm.current = 'view_well_screen'  # Из режима редактирования скважины возвращаем на экран просмотра её параметров
                return True  # Возвращаем True, чтобы остановить дальнейшую обработку клавиши (экранируем от закрытия)

            elif current in ['view_well_screen', 'create_well_screen']:
                self.sm.current = 'open_well_screen'  # Из просмотра параметров или создания скважины возвращаем в общий Журнал списка
                return True

            elif current == 'open_well_screen':
                self.sm.current = 'notebook_main_menu'  # Из Журнала списка скважин возвращаем в главное меню Блокнота
                return True

            elif current in ['notebook_main_menu', 'threads', 'settings', 'calc', 'spo', 'vzd', 'cmoist', 'fishing', 'calc_hub']:
                self.sm.current = 'menu'  # Из всех основных калькуляторов, справочников и меню Блокнота возвращаем на главный экран приложения
                return True

            elif current == 'kill_res':
                self.sm.current = 'calc'  # С экрана результатов расчетов глушения возвращаем обратно в калькулятор параметров
                return True

            else:
                self.sm.current = 'menu'  # Аварийный сценарий: с любого непредусмотренного экрана сбрасываем в главное меню
                return True

        return True  # ИСПРАВЛЕНО: Если нажата любая другая кнопка (буква/цифра) — просто пропускаем её без вылетов!

Builder.load_string('''
<NotebookMainMenu>:
    name: "notebook_main_menu"
    MDBoxLayout:
        orientation: 'vertical'
        md_bg_color: 0.05, 0.05, 0.05, 1
        MDTopAppBar:
            title: "Журнал скважин"
            anchor_title: "left"
            elevation: 4
            md_bg_color: 0.9, 0.4, 0.0, 1
            left_action_items: [["arrow-left", lambda x: root.manager.nav_to('menu') if hasattr(root.manager, 'nav_to') else setattr(root.manager, 'current', 'menu')]]

        MDBoxLayout:
            orientation: 'vertical'
            padding: ["16dp", "24dp", "16dp", "16dp"]
            spacing: "16dp"
            size_hint_y: None
            height: self.minimum_height
            MDRaisedButton:
                md_bg_color: 0.9, 0.4, 0.0, 1
                size_hint_x: 1
                height: "54dp"
                elevation: 2
                on_release: root.manager.current = "create_well_screen"
                MDLabel:
                    text: "СОЗДАТЬ СКВАЖИНУ"
                    halign: "center"
                    valign: "middle"
                    font_size: self.width * 0.045
                    theme_text_color: "Custom"
                    text_color: 1, 1, 1, 1
                    bold: True
            MDRaisedButton:
                md_bg_color: 0.15, 0.15, 0.15, 1
                size_hint_x: 1
                height: "54dp"
                elevation: 1
                on_release: root.manager.current = "open_well_screen"
                MDLabel:
                    text: "ОТКРЫТЬ СУЩЕСТВУЮЩУЮ"
                    halign: "center"
                    valign: "middle"
                    font_size: self.width * 0.045
                    theme_text_color: "Custom"
                    text_color: 0.9, 0.4, 0.0, 1
                    bold: True
        Widget:
            size_hint_y: 1

<OpenWellScreen>:
    name: "open_well_screen"
    MDBoxLayout:
        orientation: 'vertical'
        md_bg_color: 0.05, 0.05, 0.05, 1
        MDTopAppBar:
            title: "Выбор скважины"
            anchor_title: "left"
            md_bg_color: 0.9, 0.4, 0.0, 1
            left_action_items: [["arrow-left", lambda x: root.go_back()]]
        ScrollView:
            do_scroll_x: False
            do_scroll_y: True
            MDBoxLayout:
                id: container_list
                orientation: 'vertical'
                adaptive_height: True
                padding: "16dp"
                spacing: "12dp"

<WellCard>:
    size_hint_y: None
    height: "72dp"
    canvas.before:
        Color:
            rgba: 0.12, 0.12, 0.12, 1
        RoundedRectangle:
            pos: self.pos
            size: self.size
            radius: [8, ]
        Color:
            rgba: 0.9, 0.4, 0.0, 1
        RoundedRectangle:
            pos: self.pos
            size: [6, self.height]
            radius: [8, ]
''')
Builder.load_string('''
<CreateWellScreen>:
    name: "create_well_screen"
    MDBoxLayout:
        orientation: 'vertical'
        md_bg_color: 0.05, 0.05, 0.05, 1
        MDTopAppBar:
            title: "Внесение данных из плана"
            anchor_title: "left"
            md_bg_color: 0.9, 0.4, 0.0, 1
            left_action_items: [["arrow-left", lambda x: root.go_back()]]
        ScrollView:
            MDBoxLayout:
                orientation: 'vertical'
                adaptive_height: True
                padding: "20dp"
                spacing: "14dp"
                
                MDLabel:
                    text: "ОБЩАЯ ИНФОРМАЦИЯ"
                    theme_text_color: "Custom"
                    text_color: 0.9, 0.4, 0.0, 1
                    font_style: "Subtitle2"
                    bold: True
                MDTextField:
                    id: f_field
                    hint_text: "Месторождение"
                    mode: "rectangle"
                    line_color_normal: 0.9, 0.4, 0.0, 1
                    text_color_normal: 1, 1, 1, 1
                MDTextField:
                    id: f_pad
                    hint_text: "Куст"
                    mode: "rectangle"
                    line_color_normal: 0.9, 0.4, 0.0, 1
                    text_color_normal: 1, 1, 1, 1
                MDTextField:
                    id: f_well
                    hint_text: "Скважина"
                    mode: "rectangle"
                    line_color_normal: 0.9, 0.4, 0.0, 1
                    text_color_normal: 1, 1, 1, 1
                    
                MDSeparator:
                    color: 0.15, 0.15, 0.15, 1
                MDLabel:
                    text: "ТЕКУЩИЕ ДАВЛЕНИЯ (атм)"
                    theme_text_color: "Custom"
                    text_color: 0.9, 0.4, 0.0, 1
                    font_style: "Subtitle2"
                    bold: True
                MDBoxLayout:
                    orientation: 'horizontal'
                    spacing: "10dp"
                    adaptive_height: True
                    MDTextField:
                        id: f_p_buf
                        hint_text: "Р труб"
                        mode: "rectangle"
                        line_color_normal: 0.9, 0.4, 0.0, 1
                        text_color_normal: 1, 1, 1, 1
                    MDTextField:
                        id: f_p_zatr
                        hint_text: "Р затр"
                        mode: "rectangle"
                        line_color_normal: 0.9, 0.4, 0.0, 1
                        text_color_normal: 1, 1, 1, 1
                    MDTextField:
                        id: f_p_pl
                        hint_text: "Р пл"
                        mode: "rectangle"
                        line_color_normal: 0.9, 0.4, 0.0, 1
                        text_color_normal: 1, 1, 1, 1

                # ТЕХНОЛОГИЧЕСКИЙ ФИКС: Вынесли плотность отдельной широкой строкой строго под давления
                MDTextField:
                    id: f_rho_c
                    hint_text: "Текущая плотность раствора в скважине, г/см³"
                    mode: "rectangle"
                    line_color_normal: 0.9, 0.4, 0.0, 1
                    text_color_normal: 1, 1, 1, 1

                MDSeparator:
                    color: 0.15, 0.15, 0.15, 1

                MDLabel:
                    text: "ГЕОМЕТРИЯ И КОНСТРУКЦИЯ СКВАЖИНЫ"
                    theme_text_color: "Custom"
                    text_color: 0.9, 0.4, 0.0, 1
                    font_style: "Subtitle2"
                    bold: True
                MDTextField:
                    id: f_depth
                    hint_text: "Искусственный забой, м"
                    mode: "rectangle"
                    line_color_normal: 0.9, 0.4, 0.0, 1
                    text_color_normal: 1, 1, 1, 1
                 # ТЕХНОЛОГИЧЕСКИЙ ФИКС: Новое поле длины хвостовика отдельной широкой строкой
                MDTextField:
                    id: f_l_hvost
                    hint_text: "Длина хвостовика 114, м"
                    mode: "rectangle"
                    line_color_normal: 0.9, 0.4, 0.0, 1
                    text_color_normal: 1, 1, 1, 1
               
                MDLabel:
                    text: "Интервал перфорации, м"
                    theme_text_color: "Secondary"
                    font_style: "Caption"
                MDBoxLayout:
                    orientation: 'horizontal'
                    spacing: "10dp"
                    adaptive_height: True
                    MDTextField:
                        id: f_perf_top
                        hint_text: "Верх (кровля)"
                        mode: "rectangle"
                        line_color_normal: 0.9, 0.4, 0.0, 1
                        text_color_normal: 1, 1, 1, 1
                    MDTextField:
                        id: f_perf_bot
                        hint_text: "Низ (подошва)"
                        mode: "rectangle"
                        line_color_normal: 0.9, 0.4, 0.0, 1
                        text_color_normal: 1, 1, 1, 1
                        
                MDTextField:
                    id: f_d_ek
                    hint_text: "Диаметр ЭК, мм"
                    mode: "rectangle"
                    line_color_normal: 0.9, 0.4, 0.0, 1
                    text_color_normal: 1, 1, 1, 1
                    
                MDBoxLayout:
                    orientation: 'horizontal'
                    spacing: "10dp"
                    adaptive_height: True
                    MDLabel:
                        text: "Тип НКТ:"
                        theme_text_color: "Secondary"
                        size_hint_x: 0.3
                        pos_hint: {"center_y": .5}
                    MDDropDownItem:
                        id: f_nkt_drop
                        text: "Выбрать НКТ"
                        size_hint_x: 0.7
                        on_release: root.open_nkt_menu()
                        
                MDTextField:
                    id: f_h_gno
                    hint_text: "Глубина спуска ГНО / воронки, м"
                    mode: "rectangle"
                    line_color_normal: 0.9, 0.4, 0.0, 1
                    text_color_normal: 1, 1, 1, 1
                    
                MDSeparator:
                    color: 0.15, 0.15, 0.15, 1
                MDLabel:
                    text: "ТЕКУЩИЙ СТАТУС"
                    theme_text_color: "Custom"
                    text_color: 0.9, 0.4, 0.0, 1
                    font_style: "Subtitle2"
                    bold: True
                MDTextField:
                    id: f_brigade
                    hint_text: "Бригада"
                    mode: "rectangle"
                    line_color_normal: 0.9, 0.4, 0.0, 1
                    text_color_normal: 1, 1, 1, 1
                MDTextField:
                    id: f_operation
                    hint_text: "Текущая операция"
                    mode: "rectangle"
                    line_color_normal: 0.9, 0.4, 0.0, 1
                    text_color_normal: 1, 1, 1, 1
                MDTextField:
                    id: f_notes
                    hint_text: "Осложнения, примечания..."
                    mode: "rectangle"
                    line_color_normal: 0.9, 0.4, 0.0, 1
                    text_color_normal: 1, 1, 1, 1
                    multiline: True
                    max_height: "100dp"
                    
                MDFloatLayout:
                    size_hint_y: None
                    height: "54dp"
                    canvas.before:
                        Color:
                            rgba: 0.9, 0.4, 0.0, 1
                        RoundedRectangle:
                            pos: self.pos
                            size: self.size
                            radius: [8, ]
                    Button:
                        background_color: 0, 0, 0, 0
                        size_hint: 1, 1
                        pos_hint: {"center_x": .5, "center_y": .5}
                        on_release: root.save_well()
                    MDLabel:
                        text: "СОХРАНИТЬ В ЖУРНАЛ"
                        halign: "center"
                        valign: "middle"
                        size_hint: 0.9, 0.9
                        pos_hint: {"center_x": .5, "center_y": .5}
                        font_size: self.width * 0.05
                        theme_text_color: "Custom"
                        text_color: 1, 1, 1, 1
                        bold: True
''')
Builder.load_string('''
<EditWellScreen>:
    name: "edit_well_screen"
    MDBoxLayout:
        orientation: 'vertical'
        md_bg_color: 0.05, 0.05, 0.05, 1
        MDTopAppBar:
            title: "Параметры и действия"
            anchor_title: "left"
            md_bg_color: 0.9, 0.4, 0.0, 1
            left_action_items: [["arrow-left", lambda x: root.go_back()]]
        ScrollView:
            MDBoxLayout:
                orientation: 'vertical'
                adaptive_height: True
                padding: "20dp"
                spacing: "12dp"
                MDTextField:
                    id: e_field
                    hint_text: "Месторождение"
                    mode: "rectangle"
                MDTextField:
                    id: e_pad
                    hint_text: "Куст"
                    mode: "rectangle"
                MDTextField:
                    id: e_well
                    hint_text: "Скважина"
                    mode: "rectangle"
                MDSeparator:
                    color: 0.15, 0.15, 0.15, 1
                MDLabel:
                    text: "ДАВЛЕНИЯ (атм)"
                    theme_text_color: "Custom"
                    text_color: 0.9, 0.4, 0.0, 1
                    bold: True
                MDBoxLayout:
                    orientation: 'horizontal'
                    spacing: "8dp"
                    adaptive_height: True
                    MDTextField:
                        id: e_p_buf
                        hint_text: "Р труб"
                        mode: "rectangle"
                    MDTextField:
                        id: e_p_zatr
                        hint_text: "Р затр"
                        mode: "rectangle"
                    MDTextField:
                        id: e_p_pl
                        hint_text: "Р пл"
                        mode: "rectangle"
                MDSeparator:
                    color: 0.15, 0.15, 0.15, 1
                MDLabel:
                    text: "ГЕОМЕТРИЯ И КОНСТРУКЦИЯ"
                    theme_text_color: "Custom"
                    text_color: 0.9, 0.4, 0.0, 1
                    bold: True
                MDTextField:
                    id: e_depth
                    hint_text: "Искусственный забой, м"
                    mode: "rectangle"
                    
                # ТЕХНОЛОГИЧЕСКИЙ ФИКС: Длина хвостовика на экране редактирования параметров
                MDTextField:
                    id: e_l_hvost
                    hint_text: "Длина хвостовика 114, м"
                    mode: "rectangle"

                MDBoxLayout:
                    orientation: 'horizontal'
                    spacing: "10dp"
                    adaptive_height: True
                    MDTextField:
                        id: e_perf_top
                        hint_text: "Перф. верх (кровля)"
                        mode: "rectangle"
                    MDTextField:
                        id: e_perf_bot
                        hint_text: "Перф. низ (подошва)"
                        mode: "rectangle"
                MDTextField:
                    id: e_d_ek
                    hint_text: "Диаметр ЭК, мм"
                    mode: "rectangle"
                MDBoxLayout:
                    orientation: 'horizontal'
                    spacing: "10dp"
                    adaptive_height: True
                    MDLabel:
                        text: "Тип НКТ:"
                        theme_text_color: "Secondary"
                        size_hint_x: 0.3
                        pos_hint: {"center_y": .5}
                    MDDropDownItem:
                        id: e_nkt_drop
                        text: "Выбрать НКТ"
                        size_hint_x: 0.7
                        on_release: root.open_nkt_menu()
                MDTextField:
                    id: e_h_gno
                    hint_text: "Глубина спуска ГНО / воронки, м"
                    mode: "rectangle"
                MDSeparator:
                    color: 0.15, 0.15, 0.15, 1
                MDTextField:
                    id: e_brigade
                    hint_text: "Бригада"
                    mode: "rectangle"
                MDTextField:
                    id: e_operation
                    hint_text: "Текущая операция"
                    mode: "rectangle"
                MDTextField:
                    id: e_notes
                    hint_text: "Заметки..."
                    mode: "rectangle"
                    multiline: True
                    max_height: "100dp"
                MDSeparator:
                    color: 0.15, 0.15, 0.15, 1
                MDBoxLayout:
                    orientation: 'horizontal'
                    spacing: "12dp"
                    adaptive_height: True
                MDRaisedButton:
                    text: "💾   СОХРАНИТЬ ИЗМЕНЕНИЯ"
                    md_bg_color: 0.9, 0.4, 0.0, 1
                    size_hint_x: 1
                    height: "50dp"
                    on_release: root.update_well()
''')

def get_db_connection():
    import sqlite3
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'oil_mate_notebook_test.db')
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # 1. Обновленная таблица журнала скважин с полем l_hvost (длина хвостовика)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS wells (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            field TEXT, pad TEXT, well TEXT, 
            p_buf TEXT, p_zatr TEXT, p_pl TEXT, rho_c TEXT,
            depth TEXT, l_hvost TEXT, perf_top TEXT, perf_bot TEXT, 
            d_ek TEXT, nkt_type TEXT, h_gno TEXT,
            brigade TEXT, operation TEXT, notes TEXT
        )
    ''')
    
    # 2. Наша рабочая таблица справочника бригад
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS brigades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            number TEXT UNIQUE,
            master TEXT,
            phone_int TEXT,
            phone_mob TEXT
        )
    ''')
    conn.commit()
    return conn


class WellCard(MDFloatLayout):
    pass

class NotebookMainMenu(MDScreen):
    pass

class CreateWellScreen(MDScreen):
    def go_back(self):
        self.manager.current = "notebook_main_menu"
        self.clear_fields()

    def on_enter(self):
        # Железно включаем отслеживание ввода текста при каждом входе на экран
        self.ids.f_brigade.bind(text=self.auto_check_brigade)

    def open_nkt_menu(self):
        from kivymd.uix.menu import MDDropdownMenu
        import os

        if hasattr(self, 'menu') and self.menu:
            self.menu.open()
            return

        nkt_items = []
        excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'nkt_base.xlsx')
        
        if os.path.exists(excel_path):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(excel_path, read_only=True)
                sheet = wb.active
                for row in range(2, sheet.max_row + 1):
                    val = sheet.cell(row=row, column=1).value
                    if val is not None:
                        nkt_items.append(f"НКТ-{int(float(val))}")
                wb.close()
            except Exception:
                nkt_items = []

        if not nkt_items:
            nkt_items = ["НКТ-60", "НКТ-73", "НКТ-89"]

        menu_items = [
            {
                "viewclass": "OneLineListItem",
                "text": item,
                "on_release": lambda x=item: self.set_nkt_value(x),
            } for item in nkt_items
        ]
        self.menu = MDDropdownMenu(
            caller=self.ids.f_nkt_drop,
            items=menu_items,
            width=200,
        )
        self.menu.open()

    def set_nkt_value(self, value):
        self.ids.f_nkt_drop.text = value
        if hasattr(self, 'menu') and self.menu:
            self.menu.dismiss()

    def clear_fields(self):
        self.ids.f_field.text = ""
        self.ids.f_pad.text = ""
        self.ids.f_well.text = ""
        self.ids.f_p_buf.text = ""
        self.ids.f_p_zatr.text = ""
        self.ids.f_p_pl.text = ""
        self.ids.f_rho_c.text = "" # Чистим поле текущей плотности
        self.ids.f_depth.text = ""
        self.ids.f_perf_top.text = ""
        self.ids.f_perf_bot.text = ""
        self.ids.f_d_ek.text = ""
        self.ids.f_nkt_drop.text = "Выбрать НКТ"
        self.ids.f_h_gno.text = ""
        self.ids.f_brigade.text = ""
        self.ids.f_operation.text = ""
        self.ids.f_notes.text = ""

    def save_well(self):
        if not self.ids.f_well.text.strip():
            return
        conn = get_db_connection()
        cursor = conn.cursor()
        # ИСПРАВЛЕНО: Добавлен параметр rho_c строго на 7-е место (всего 16 полей базы)
        cursor.execute(
            "INSERT INTO wells (field, pad, well, p_buf, p_zatr, p_pl, rho_c, depth, perf_top, perf_bot, d_ek, nkt_type, h_gno, brigade, operation, notes) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (self.ids.f_field.text, self.ids.f_pad.text, self.ids.f_well.text, 
             self.ids.f_p_buf.text, self.ids.f_p_zatr.text, self.ids.f_p_pl.text, self.ids.f_rho_c.text,
             self.ids.f_depth.text, self.ids.f_perf_top.text, self.ids.f_perf_bot.text,
             self.ids.f_d_ek.text, self.ids.f_nkt_drop.text, self.ids.f_h_gno.text,
             self.ids.f_brigade.text, self.ids.f_operation.text, self.ids.f_notes.text)
        )
        conn.commit()
        conn.close()
        self.go_back()

    def auto_check_brigade(self, instance, text):
        val = text.strip()
        # ЗАЩИТА ОТ ЦИКЛА: если в поле уже длинная подстановка со словом "мастер", выходим
        if not val or "мастер" in val.lower():
            return

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT master, phone_int, phone_mob FROM brigades WHERE number=?", (val,))
        b = cursor.fetchone()
        conn.close()

        if b:
            master_name = b[0]
            p_int = f", вн. {b[1]}" if b[1] else ""
            p_mob = f", сот. {b[2]}" if b[2] else ""
            
            # Временно снимаем слежение, чтобы не вызвать повторный триггер на изменение текста
            self.ids.f_brigade.unbind(text=self.auto_check_brigade)
            self.ids.f_brigade.text = f"Бригада №{val} (мастер {master_name}{p_int}{p_mob})"
            self.ids.f_brigade.bind(text=self.auto_check_brigade)

class OpenWellScreen(MDScreen): # Класс экрана, отображающего весь список сохраненных в Блокноте скважин
    def go_back(self):
        """Метод навигации: возвращает пользователя на главный экран меню Блокнота"""
        self.manager.current = "notebook_main_menu"

    def on_enter(self):
        """Автоматическое событие Kivy: срабатывает при открытии списка и динамически перерисовывает карточки из БД"""
        from kivymd.uix.label import MDLabel # Локальный импорт текстовой метки Material Design
        from kivy.uix.button import Button # Локальный импорт базовой прозрачной кнопки для перехвата кликов
        
        # Гарантированно очищаем контейнер
        self.ids.container_list.clear_widgets() # Стираем все старые карточки из списка, чтобы избежать дублирования при повторном входе
        
        conn = get_db_connection() # Открываем сессию соединения с локальной базой данных SQLite
        cursor = conn.cursor()
        # Выгружаем сокращенный список параметров (ID, месторождение, куст, скважина) с сортировкой от новых записей к старым
        cursor.execute("SELECT id, field, pad, well FROM wells ORDER BY id DESC")
        rows = cursor.fetchall() # Забираем все найденные строки из базы данных
        conn.close() # Закрываем соединение с файлом SQLite
        
        if not rows:
            # ТЕХНОЛОГИЧЕСКИЙ ФИКС: Стабильный MDLabel вместо капризного OneLineListItem
            # Если база данных пустая, выводим информирующую заглушку по центру контентной зоны
            lbl = MDLabel(
                text="Журнал скважин пуст. Создайте запись.", # Текст-подсказка
                halign="center", # Центрирование текста по горизонтали
                valign="middle", # Центрирование текста по вертикали
                theme_text_color="Hint", # Серый цвет шрифта, принятый для системных подсказок
                size_hint_y=None, 
                height="48dp" # Компактная фиксированная высота текстовой заглушки
            )
            self.ids.container_list.add_widget(lbl) # Физически выводим заглушку на экран Журнала
        else:
            # Если записи обнаружены, запускаем цикл генерации графических карточек для каждой строки БД
            for row in rows:
                w_id = int(row[0]) # Извлекаем уникальный целочисленный идентификатор скважины
                w_field = str(row[1]) # Извлекаем текстовое название месторождения
                w_pad = str(row[2]) # Извлекаем номер кустовой площадки
                w_well = str(row[3]) # Извлекаем номер скважины
                
                card = WellCard() # Создаем экземпляр кастомного графического виджета карточки WellCard
                
                # Создаем верхний текстовый блок карточки с номером скважины (крупный белый текст)
                lbl_title = MDLabel(
                    text=f"Скважина № {w_well}",
                    font_style="Subtitle1",
                    bold=True, # Жирный акцент на номере
                    theme_text_color="Custom",
                    text_color=[1, 1, 1, 1], # Чистый белый цвет букв
                    pos_hint={"x": 0.05, "top": 0.95}, # Позиционирование с отступом 5% слева и вверху карточки
                    size_hint=[0.9, 0.4] # Адаптивные габариты текстового блока
                )
                
                # Создаем нижний текстовый блок карточки с кустом и месторождением (оранжевый компактный текст)
                lbl_sub = MDLabel(
                    text=f"Куст {w_pad} | {w_field} м-р",
                    font_style="Caption", # Мелкий системный шрифт подписей
                    theme_text_color="Custom",
                    text_color=[0.9, 0.4, 0.0, 1], # Корпоративный оранжевый цвет Oil Mate
                    pos_hint={"x": 0.05, "top": 0.55}, # Смещение под верхний заголовок
                    size_hint=[0.9, 0.4]
                )
                
                # Создаем невидимую (прозрачную) кнопку поверх всей карточки для обработки нажатия инженером
                btn = Button(
                    background_color=[0, 0, 0, 0], # Полностью прозрачный фон (альфа-канал равен 0)
                    size_hint=[1, 1], # Кнопка растягивается на 100% ширины и высоты родительской карточки
                    pos_hint={"center_x": 0.5, "center_y": 0.5} # Центрирование триггера клика
                )
                
                # Привязываем событие отпускания кнопки к методу перехода, жестко зафиксировав текущий w_id через lambda-замыкание
                btn.bind(on_release=lambda x, id_val=w_id: self.open_view_screen(id_val))
                
                # Укладываем все сформированные элементы внутрь нашей карточки WellCard
                card.add_widget(lbl_title)
                card.add_widget(lbl_sub)
                card.add_widget(btn)
                
                # Помещаем готовую карточку в общий вертикальный список container_list на экране Журнала
                self.ids.container_list.add_widget(card)

    def open_view_screen(self, well_id):
        """Инженерный метод: принимает ID кликнутой скважины, передает его экрану просмотра и осуществляет переход"""
        # ИСПРАВЛЕНО ИМЯ: Теперь имя метода строго совпадает с вызовом выше
        view_screen = self.manager.get_screen("view_well_screen") # Находим целевой объект экрана просмотра
        view_screen.current_well_id = well_id # Передаем ID скважины, чтобы экран просмотра выгрузил именно её параметры
        self.manager.current = "view_well_screen" # Совершаем физическое переключение интерфейса на просмотр параметров



class EditWellScreen(MDScreen): # Класс экрана редактирования параметров существующей скважины, наследуемый от MDScreen
    current_well_id = None # Глобальный идентификатор (ID) текущей редактируемой скважины в базе данных SQLite
    
    def go_back(self):
        """Метод навигации: возвращает пользователя обратно на экран детального просмотра параметров скважины"""
        self.manager.current = "view_well_screen" 
        
    def open_nkt_menu(self):
        """Инженерный метод: динамически считывает типы труб НКТ из Excel и открывает выпадающее меню выбора"""
        from kivymd.uix.menu import MDDropdownMenu # Локальный импорт компонента выпадающего меню Material Design
        import os # Локальный импорт модуля для проверки существования файлов на диске
        
        # Защита: если меню уже открыто, повторно объекты не создаем
        if hasattr(self, 'menu') and self.menu: 
            self.menu.open() # Если объект меню уже существует в памяти, просто открываем его визуально
            return # Досрочно прерываем функцию, чтобы не перегружать файл Excel заново
            
        nkt_items = [] # Инициализируем пустой массив для хранения прочитанных марок труб НКТ
        excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'nkt_base.xlsx') # Формируем абсолютный путь к инженерной базе НКТ
        
        # ТЕХНОЛОГИЧЕСКИЙ ФИКС: Читаем чистые числа из Excel, а не имена файлов папки
        if os.path.exists(excel_path): # Проверяем, физически ли присутствует файл таблицы в папке data внутри APK/директории
            try:
                import openpyxl # Локальный импорт библиотеки для парсинга документов Excel
                wb = openpyxl.load_workbook(excel_path, read_only=True) # Открываем книгу Excel в режиме «только для чтения» для экономии ОЗУ смартфона
                sheet = wb.active # Выбираем первый активный лист таблицы со спецификациями труб
                
                # Цикл парсинга: начинаем со 2-й строки (пропуская шапку) и идем до самого конца заполненных ячеек
                for row in range(2, sheet.max_row + 1): 
                    val = sheet.cell(row=row, column=1).value # Извлекаем значение из первой колонки (Типоразмер/Диаметр)
                    if val is not None: # Если ячейка не пустая
                        nkt_items.append(f"НКТ-{int(float(val))}") # Конвертируем дробь в целое число и собираем стандартное обозначение (н-р, НКТ-73)
                wb.close() # Обязательно закрываем файл Excel, освобождая дескрипторы операционной системы
            except Exception:
                nkt_items = [] # В случае любого сбоя (битый файл, ошибка чтения) очищаем массив для предотвращения падения
                
        if not nkt_items: # Если файл Excel отсутствовал или заблокировался, включаем аварийный хардкод-резерв
            nkt_items = ["НКТ-60", "НКТ-73", "НКТ-89"] # Стандартный сортамент труб, используемый на промыслах РФ
            
        # Формируем структуру элементов выпадающего меню KivyMD. Каждая строка при нажатии передает свой текст в set_nkt_value
        menu_items = [ 
            {
                "viewclass": "OneLineListItem",
                "text": item,
                "on_release": lambda x=item: self.set_nkt_value(x),
            }
            for item in nkt_items
        ]
        
        # Инициализируем графический объект выпадающего меню
        self.menu = MDDropdownMenu(
            caller=self.ids.e_nkt_drop, # Привязываем выпадающий список к графическому элементу (кнопке/текстовому полю) с ID 'e_nkt_drop'
            items=menu_items, # Передаем сформированный массив строк с трубами НКТ
            width=200, # Задаем фиксированную адаптивную ширину всплывающего окна в 200 пикселей
        )
        self.menu.open() # Физически выводим меню выбора НКТ на экран поверх остальных элементов верстки
    def set_nkt_value(self, value):
        """Метод-обработчик: записывает выбранную марку НКТ из выпадающего меню в текстовое поле интерфейса"""
        self.ids.e_nkt_drop.text = value # Заменяем текст в поле-кнопке на выбранный вариант (например, 'НКТ-73')
        if hasattr(self, 'menu') and self.menu:
            self.menu.dismiss() # Закрываем всплывающее окно выбора, если оно существует в памяти

    def on_enter(self):
        """Автоматическое событие Kivy: безопасно подгружает параметры скважины в поля редактирования"""
        if self.current_well_id is None:
            return # Защита: если ID скважины не передан, прерываем функцию во избежание сбоя в БД
            
        conn = get_db_connection() # Открываем соединение с локальной базой данных SQLite
        cursor = conn.cursor()
        
        # Выгружаем строго все 17 столбцов, включая l_hvost на 9-м месте для полной синхронизации Tuple Mapping
        cursor.execute("SELECT field, pad, well, p_buf, p_zatr, p_pl, rho_c, depth, l_hvost, perf_top, perf_bot, d_ek, nkt_type, h_gno, brigade, operation, notes FROM wells WHERE id=?", (self.current_well_id,))
        w = cursor.fetchone() # Извлекаем полученную строку с технологическими данными скважины
        conn.close() # Обязательно закрываем соединение, освобождая файл базы данных
        
        if w:
            # ТЕХНОЛОГИЧЕСКИЙ ФИКС: Заворачиваем каждое поле в проверку 'in self.ids', полностью исключая краш AttributeError.
            # Если какого-то ID временно нет в вашей .kv разметке, Python просто пропустит его, но приложение больше никогда не вылетит.
            if 'e_field' in self.ids: self.ids.e_field.text = str(w[0]) if w[0] is not None else "" # Месторождение
            if 'e_pad' in self.ids: self.ids.e_pad.text = str(w[1]) if w[1] is not None else "" # Кустовая площадка (Куст)
            if 'e_well' in self.ids: self.ids.e_well.text = str(w[2]) if w[2] is not None else "" # Номер скважины
            if 'e_p_buf' in self.ids: self.ids.e_p_buf.text = str(w[3]) if w[3] is not None else "" # Давление (трубное)
            if 'e_p_zatr' in self.ids: self.ids.e_p_zatr.text = str(w[4]) if w[4] is not None else "" # Давление затрубное
            if 'e_p_pl' in self.ids: self.ids.e_p_pl.text = str(w[5]) if w[5] is not None else "" # Давление пластовое
            if 'e_rho_c' in self.ids: self.ids.e_rho_c.text = str(w[6]) if w[6] is not None else "" # ИСПРАВЛЕНО: Поле плотности раствора больше не уронит систему!
            if 'e_depth' in self.ids: self.ids.e_depth.text = str(w[7]) if w[7] is not None else "" # Искусственный забой / Глубина
            if 'e_l_hvost' in self.ids: self.ids.e_l_hvost.text = str(w[8]) if w[8] is not None else "" # Длина хвостовика (Tuple Mapping индекс 9)
            if 'e_perf_top' in self.ids: self.ids.e_perf_top.text = str(w[9]) if w[9] is not None else "" # Интервал перфорации (кровля)
            if 'e_perf_bot' in self.ids: self.ids.e_perf_bot.text = str(w[10]) if w[10] is not None else "" # Интервал перфорации (подошва)
            if 'e_d_ek' in self.ids: self.ids.e_d_ek.text = str(w[11]) if w[11] is not None else "" # Диаметр эксплуатационной колонны
            if 'e_nkt_drop' in self.ids: self.ids.e_nkt_drop.text = str(w[12]) if w[12] else "Выбрать НКТ" # Тип труб НКТ из выпадающего меню
            if 'e_h_gno' in self.ids: self.ids.e_h_gno.text = str(w[13]) if w[13] is not None else "" # Глубина спуска ГНО / воронки
            if 'e_brigade' in self.ids: self.ids.e_brigade.text = str(w[14]) if w[14] is not None else "" # Номер бригады КРС, выполняющей работы
            if 'e_operation' in self.ids: self.ids.e_operation.text = str(w[15]) if w[15] is not None else "" # Текущая технологическая операция
            if 'e_notes' in self.ids: self.ids.e_notes.text = str(w[16]) if w[16] is not None else "" # Примечания и дополнительные заметки инженера

    def update_well(self):
        """Инженерный метод: собирает измененные пользователем данные из полей ввода и сохраняет их обратно в SQLite"""
        if self.current_well_id is None:
            return # Защита от перезаписи, если ID скважины потерян
            
        conn = get_db_connection() # Подключаемся к локальной базе данных SQLite
        cursor = conn.cursor()
        
        # ТЕХНОЛОГИЧЕСКИЙ ФИКС: Безопасный сборщик текста. Полностью исключает краш AttributeError при сохранении,
        # если какого-то ID (например, e_rho_c) физически нет в вашей текущей .kv разметке экрана редактирования.
        f_field     = self.ids.e_field.text     if 'e_field'     in self.ids else "" # Месторождение
        f_pad       = self.ids.e_pad.text       if 'e_pad'       in self.ids else "" # Кустовая площадка (Куст)
        f_well      = self.ids.e_well.text      if 'e_well'      in self.ids else "" # Номер скважины
        f_p_buf     = self.ids.e_p_buf.text     if 'e_p_buf'     in self.ids else "" # Давление  (трубное)
        f_p_zatr    = self.ids.e_p_zatr.text    if 'e_p_zatr'    in self.ids else "" # Давление затрубное
        f_p_pl      = self.ids.e_p_pl.text      if 'e_p_pl'      in self.ids else "" # Давление пластовое
        f_rho_c     = self.ids.e_rho_c.text     if 'e_rho_c'     in self.ids else "" # Плотность раствора (защищено от краша)
        f_depth     = self.ids.e_depth.text     if 'e_depth'     in self.ids else "" # Искусственный забой / Глубина
        f_l_hvost   = self.ids.e_l_hvost.text   if 'e_l_hvost'   in self.ids else "" # Длина хвостовика (защищено от краша)
        f_perf_top  = self.ids.e_perf_top.text  if 'e_perf_top'  in self.ids else "" # Интервал перфорации (кровля)
        f_perf_bot  = self.ids.e_perf_bot.text  if 'e_perf_bot'  in self.ids else "" # Интервал перфорации (подошва)
        f_d_ek      = self.ids.e_d_ek.text      if 'e_d_ek'      in self.ids else "" # Диаметр эксплуатационной колонны
        f_nkt_type  = self.ids.e_nkt_drop.text  if 'e_nkt_drop'  in self.ids else "Выбрать НКТ" # Тип труб НКТ
        f_h_gno     = self.ids.e_h_gno.text     if 'e_h_gno'     in self.ids else "" # Глубина спуска ГНО / воронки
        f_brigade   = self.ids.e_brigade.text   if 'e_brigade'   in self.ids else "" # Номер бригады КРС
        f_operation = self.ids.e_operation.text if 'e_operation' in self.ids else "" # Текущая технологическая операция
        f_notes     = self.ids.e_notes.text     if 'e_notes'     in self.ids else "" # Примечания и заметки инженера

        # СТРОГИЙ ТЕХНОЛОГИЧЕСКИЙ ФИКС: Ровно 17 полей и ровно 17 знаков '?' для сохранения в базу данных
        # Выполняем SQL-запрос обновления (UPDATE) на основе собранных безопасных переменных
        cursor.execute(
            "UPDATE wells SET field=?, pad=?, well=?, p_buf=?, p_zatr=?, p_pl=?, rho_c=?, depth=?, l_hvost=?, perf_top=?, perf_bot=?, d_ek=?, nkt_type=?, h_gno=?, brigade=?, operation=?, notes=? WHERE id=?",
            (
                f_field, 
                f_pad, 
                f_well, 
                f_p_buf, 
                f_p_zatr, 
                f_p_pl, 
                f_rho_c, 
                f_depth, 
                f_l_hvost, 
                f_perf_top, 
                f_perf_bot, 
                f_d_ek, 
                f_nkt_type, 
                f_h_gno, 
                f_brigade, 
                f_operation, 
                f_notes, 
                self.current_well_id # Передаем ID текущей скважины в качестве условия WHERE
            )
        )
        conn.commit() # Жестко фиксируем (сохраняем) транзакцию изменений в файле базы данных
        conn.close() # Закрываем сессию соединения с локальной БД
        self.go_back() # Вызываем метод навигации для автоматического возврата на экран просмотра параметров

    def share_well(self):
        """Инженерный метод: собирает все технологические параметры скважины в текстовый отчет и отправляет в мессенджеры"""
        # Формируем интервал перфорации: если подошва указана, склеиваем через дефис (кровля-подошва), иначе берем только кровлю
        perf_summary = f"{self.ids.e_perf_top.text}-{self.ids.e_perf_bot.text}" if self.ids.e_perf_bot.text else self.ids.e_perf_top.text 
        
        # Собираем структурированный текстовый рапорт с использованием эмодзи для наглядности в мессенджерах (WhatsApp, Telegram)
        report = ( 
            f" ПЛАН РАБОТ: Скв. №{self.ids.e_well.text} (К-{self.ids.e_pad.text})\n" # Номер скважины и кустовая площадка
            f" {self.ids.e_field.text} м-р\n" # Наименование месторождения
            f" Давления: Ртруб={self.ids.e_p_buf.text}, Рзатр={self.ids.e_p_zatr.text}, Рпл={self.ids.e_p_pl.text} атм\n" # Устьевые и пластовые давления
            f" Забой: {self.ids.e_depth.text} м | Перфорация: {perf_summary} м\n" # Искусственный забой и интервалы вскрытия пласта
            f" Конструкция: Диаметр ЭК={self.ids.e_d_ek.text} мм\n" # Диаметр эксплуатационной колонны
            f" Подвеска: {self.ids.e_nkt_drop.text} на гл. {self.ids.e_h_gno.text} м\n" # Тип НКТ и глубина посадки воронки/насоса
            f" Бригада: {self.ids.e_brigade.text} | Операция: {self.ids.e_operation.text}\n" # Номер бригады КРС и текущие работы
            f" Заметки: {self.ids.e_notes.text}" # Текстовые примечания инженера
        ) 
        
        # Кроссплатформенная отправка: проверяем, запущено ли приложение на операционной системе Android
        if platform == 'android': 
            from jnius import autoclass # Импортируем мост Pyjnius для вызова родных Java-классов Android напрямую из Python
            
            # Подгружаем системные Java-классы операционной системы Android
            PythonActivity = autoclass('org.kivy.android.PythonActivity') # Главная активность запущенного Kivy-приложения
            Intent = autoclass('android.content.Intent') # Класс намерений Android для передачи данных между программами
            String = autoclass('java.lang.String') # Нативный строковый класс Java
            
            # Создаем системное намерение ACTION_SEND (отправка данных) с типом контента "обычный текст" (text/plain)
            shareIntent = Intent(Intent.ACTION_SEND).setType("text/plain") 
            # Вкладываем сформированный текст рапорта в тело отправления
            shareIntent.putExtra(Intent.EXTRA_TEXT, String(report)) 
            # Запускаем системное Android-меню «Поделиться» (Chooser), предлагающее пользователю выбор мессенджера или почты
            PythonActivity.mActivity.startActivity(Intent.createChooser(shareIntent, String("Поделиться"))) 
            
        else:
            # Сценарий для Windows/Linux/Mac (Режим тестирования на компьютере разработчика)
            from kivy.core.clipboard import Clipboard # Импортируем системный буфер обмена Kivy
            Clipboard.copy(report) # Копируем готовый текстовый рапорт в буфер обмена операционной системы ПК
            
            # Формируем всплывающее диалоговое окно MDDialog с уведомлением инженера о копировании
            dialog = MDDialog( 
                title="Тест ПК", # Заголовок тестового окна
                text="Текст инженерного отчета скопирован в буфер обмена.", # Сообщение-инструкция
                # Добавляем оранжевую кнопку «ОК» для закрытия этого всплывающего окна
                buttons=[MDRaisedButton(text="ОК", md_bg_color=[0.9, 0.4, 0.0, 1], on_release=lambda x: dialog.dismiss())] 
            ) 
            dialog.open() # Выводим диалог уведомления на экран компьютера

Builder.load_string(''' # Декларативная разметка интерфейса Kivy (KV Lang)
<ViewWellScreen>: # Шаблон графических элементов для экрана детального просмотра параметров скважины
    name: "view_well_screen" # Уникальное текстовое имя экрана для переключения на него менеджером ScreenManager
    
    MDBoxLayout: # Корневой вертикальный контейнер Material Design для шапки и контента
        orientation: 'vertical' # Элементы внутри укладываются друг под друга строго вертикально
        md_bg_color: 0.05, 0.05, 0.05, 1 # Глубокий темно-серый фоновый цвет интерфейса
        
        MDTopAppBar: # Верхняя панель навигации (Тулбар экрана просмотра)
            title: "Просмотр параметров" # Текст главного заголовка на панели
            anchor_title: "left" # Выравнивание заголовка по левому краю тулбара
            md_bg_color: 0.9, 0.4, 0.0, 1 # Фирменный оранжевый цвет верхней панели Oil Mate
            left_action_items: [["arrow-left", lambda x: root.go_back()]] # Левая иконка-стрелка с привязанной функцией возврата назад
            
        ScrollView: # Зона прокрутки, позволяющая скроллить длинный список параметров скважины
            MDBoxLayout: # Внутренний контейнер скролла, сжимающийся по высоте
                orientation: 'vertical'
                adaptive_height: True # Позволяет контейнеру автоматически менять высоту в зависимости от количества полей
                padding: "20dp" # Внешний отступ контента от краев дисплея смартфона
                spacing: "12dp" # Пошаговые вертикальные расстояния (зазоры) между соседними текстовыми полями
                
                MDTextField: # Поле для отображения названия месторождения
                    id: v_field # Идентификатор виджета для автоматического связывания данных из Python-кода
                    hint_text: "Месторождение" # Всплывающая подсказка над рамкой поля
                    mode: "rectangle" # Стиль отображения поля со сплошным прямоугольным контуром
                    readonly: True # Блокировка ввода: поле закрыто только на чтение данных инженером
                    
                MDTextField: # Поле для отображения кустовой площадки (Куста)
                    id: v_pad
                    hint_text: "Куст"
                    mode: "rectangle"
                    readonly: True
                    
                MDTextField: # Поле для отображения номера скважины
                    id: v_well
                    hint_text: "Скважина"
                    mode: "rectangle"
                    readonly: True
                    
                MDSeparator: # Компонент тонкой визуальной полосы-разделителя логических блоков
                    color: 0.15, 0.15, 0.15, 1 # Серый цвет линии разделения
                    
                MDLabel: # Текстовый блок-подзаголовок для секции давлений
                    text: "ДАВЛЕНИЯ (атм)" # Название группы технологических параметров
                    theme_text_color: "Custom" # Активация кастомной палитры цвета для надписи
                    text_color: 0.9, 0.4, 0.0, 1 # Выделение оранжевым цветом
                    bold: True # Использование жирного шрифта
                    
                MDBoxLayout: # Горизонтальный ряд для размещения трех полей давлений в одну строчку
                    orientation: 'horizontal'
                    spacing: "8dp" # Горизонтальный зазор между текстовыми ячейками
                    adaptive_height: True # Минимизация высоты ряда под вертикальный габарит полей
                    
                    MDTextField: # Отображение устьевого буферного давления
                        id: v_p_buf
                        hint_text: "Р труб"
                        mode: "rectangle"
                        readonly: True
                        
                    MDTextField: # Отображение затрубного давления скважины
                        id: v_p_zatr
                        hint_text: "Р затр"
                        mode: "rectangle"
                        readonly: True
                        
                    MDTextField: # Отображение текущего пластового давления скважины КРС
                        id: v_p_pl
                        hint_text: "Р пл"
                        mode: "rectangle"
                        readonly: True
                        
                # ТЕХНОЛОГИЧЕСКИЙ ФИКС: Добавили плотность отдельной закрытой строкой на экран просмотра
                MDTextField: # Отображение текущего удельного веса раствора в стволе
                    id: v_rho_c
                    hint_text: "Текущая плотность раствора в скважине, г/см³"
                    mode: "rectangle"
                    readonly: True
                    
                MDSeparator: # Разделительная черта перед геометрическими параметрами искусственного забоя
                    color: 0.15, 0.15, 0.15, 1
                    
                MDTextField: # Поле для вывода глубины искусственного забоя (цементного кольца)
                    id: v_depth
                    hint_text: "Искусственный забой, м"
                    mode: "rectangle"
                    readonly: True
                    
                # ТЕХНОЛОГИЧЕСКИЙ ФИКС: Новое закрытое поле длины хвостовика на экране просмотра
                MDTextField: # Отображение длины спущенного хвостовика (Tuple Mapping индекс 9)
                    id: v_l_hvost
                    hint_text: "Длина хвостовика 114, м"
                    mode: "rectangle"
                    readonly: True
                    
                MDBoxLayout: # Горизонтальный ряд для интервалов перфорации
                    orientation: 'horizontal'
                    spacing: "10dp"
                    adaptive_height: True
                    
                    MDTextField: # Отображение верхней границы вскрытого пласта (кровли перфорации)
                        id: v_perf_top
                        hint_text: "Перф. верх"
                        mode: "rectangle"
                        readonly: True
                        
                    MDTextField: # Отображение нижней границы вскрытого пласта (подошвы перфорации)
                        id: v_perf_bot
                        hint_text: "Перф. низ"
                        mode: "rectangle"
                        readonly: True
                        
                MDTextField: # Отображение диаметра эксплуатационной колонны скважины
                    id: v_d_ek
                    hint_text: "Диаметр ЭК, мм"
                    mode: "rectangle"
                    readonly: True
                    
                MDTextField: # Отображение маркировки и типа труб НКТ, выбранных из Excel
                    id: v_nkt_type
                    hint_text: "Тип НКТ"
                    mode: "rectangle"
                    readonly: True
                    
                MDTextField: # Отображение текущей глубины спуска воронки НКТ или пакера
                    id: v_h_gno
                    hint_text: "Глубина спуска ГНО / воронки, м"
                    mode: "rectangle"
                    readonly: True
                    
                MDSeparator: # Разделительная черта перед административным блоком бригад КРС
                    color: 0.15, 0.15, 0.15, 1
                    
                MDTextField: # Отображение номера бригады ТКРС/КРС, выполняющей работы на объекте
                    id: v_brigade
                    hint_text: "Бригада "
                    mode: "rectangle"
                    readonly: True
                    
                MDTextField: # Отображение текущей технологической операции по плану работ
                    id: v_operation
                    hint_text: "Текущая операция"
                    mode: "rectangle"
                    readonly: True
                    
                MDTextField: # Отображение текстовых заметок и примечаний инженера
                    id: v_notes
                    hint_text: "Заметки..."
                    mode: "rectangle"
                    multiline: True # Разрешаем перенос текста на новые строки для больших примечаний
                    max_height: "100dp" # Ограничиваем максимальную высоту раскрытия текстовой области
                    readonly: True
                    
                MDSeparator:
                    color: 0.15, 0.15, 0.15, 1
                MDBoxLayout:
                    orientation: 'horizontal'
                    spacing: "12dp"
                    adaptive_height: True
                    MDRaisedButton:
                        text: "ПРАВКА"
                        md_bg_color: 0.9, 0.4, 0.0, 1
                        size_hint_x: 0.5
                        height: "48dp"
                        on_release: root.open_edit_screen()
                    MDRaisedButton:
                        text: "ПОДЕЛИТЬСЯ"
                        md_bg_color: 0.15, 0.45, 0.15, 1
                        size_hint_x: 0.5
                        height: "48dp"
                        on_release: root.share_well()
                MDBoxLayout:
                    orientation: 'horizontal'
                    spacing: "12dp"
                    adaptive_height: True
                    MDRaisedButton:
                        text: "ГЛУШЕНИЕ"
                        md_bg_color: 0.1, 0.5, 0.8, 1
                        size_hint_x: 0.5
                        height: "48dp"
                        on_release: root.open_in_glushenie()
                    MDRaisedButton:
                        text: "УДАЛИТЬ"
                        md_bg_color: 0.85, 0.15, 0.15, 1
                        size_hint_x: 0.5
                        height: "48dp"
                        on_release: root.confirm_delete_well()

''')
class ViewWellScreen(MDScreen):
    current_well_id = None # Глобальный идентификатор выбранной скважины

    def go_back(self):
        """Метод навигации: возвращает пользователя на экран списка скважин"""
        self.manager.current = "open_well_screen"

    def on_enter(self):
        """Автоматическое событие Kivy: подгружает полные данные скважины при открытии экрана"""
        if self.current_well_id is None:
            return
            
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # ТЕХНОЛОГИЧЕСКИЙ ФИКС БАГА: Выгружаем строго все 17 полей, включая l_hvost на 9-м месте для синхронизации Tuple Mapping
        cursor.execute("SELECT field, pad, well, p_buf, p_zatr, p_pl, rho_c, depth, l_hvost, perf_top, perf_bot, d_ek, nkt_type, h_gno, brigade, operation, notes FROM wells WHERE id=?", (self.current_well_id,))
        w = cursor.fetchone()
        conn.close()
        
        if w:
            # Распределяем данные по идентификаторам разметки (IDs) согласно эталонным индексам паспорта
            self.ids.v_field.text = str(w[0]) if w[0] is not None else ""
            self.ids.v_pad.text = str(w[1]) if w[1] is not None else ""
            self.ids.v_well.text = str(w[2]) if w[2] is not None else ""
            self.ids.v_p_buf.text = str(w[3]) if w[3] is not None else ""
            self.ids.v_p_zatr.text = str(w[4]) if w[4] is not None else ""
            self.ids.v_p_pl.text = str(w[5]) if w[5] is not None else ""
            self.ids.v_rho_c.text = str(w[6]) if w[6] is not None else "" # Текущая плотность раствора
            self.ids.v_depth.text = str(w[7]) if w[7] is not None else "" # Искусственный забой
            self.ids.v_l_hvost.text = str(w[8]) if w[8] is not None else "" # ИСПРАВЛЕНО: Теперь длина хвостовика успешно заполняется!
            self.ids.v_perf_top.text = str(w[9]) if w[9] is not None else "" # Индексы корректно сместились на +1
            self.ids.v_perf_bot.text = str(w[10]) if w[10] is not None else ""
            self.ids.v_d_ek.text = str(w[11]) if w[11] is not None else ""
            self.ids.v_nkt_type.text = str(w[12]) if w[12] is not None else ""
            self.ids.v_h_gno.text = str(w[13]) if w[13] is not None else ""
            self.ids.v_brigade.text = str(w[14]) if w[14] is not None else ""
            self.ids.v_operation.text = str(w[15]) if w[15] is not None else ""
            self.ids.v_notes.text = str(w[16]) if w[16] is not None else ""

    def open_edit_screen(self):
        """Переходит на экран редактирования, передавая ID скважины"""
        edit_screen = self.manager.get_screen("edit_well_screen")
        edit_screen.current_well_id = self.current_well_id
        self.manager.current = "edit_well_screen"

    def confirm_delete_well(self):
        """Инженерный метод: вызывает модальное диалоговое окно для подтверждения удаления скважины"""
        from kivymd.uix.button import MDFlatButton
        from kivymd.uix.dialog import MDDialog
        
        well_num = self.ids.v_well.text # Достаем номер текущей скважины для вывода в предупреждении
        
        self.delete_dialog = MDDialog(
            title="Предупреждение",
            text=f"Вы действительно хотите удалить скважину №{well_num}? Данное действие полностью сотрет её из базы данных.",
            buttons=[
                MDFlatButton(
                    text="ОТМЕНА",
                    on_release=lambda x: self.delete_dialog.dismiss()
                ),
                MDFlatButton(
                    text="УДАЛИТЬ",
                    text_color=(1, 0, 0, 1), # Сигнальный красный цвет для кнопки подтверждения
                    on_release=lambda x: self.execute_delete_well()
                ),
            ],
        )
        self.delete_dialog.open()

    def execute_delete_well(self):
        """Бэкенд-метод: производит физическое удаление строки из SQLite и обновляет Журнал"""
        self.delete_dialog.dismiss() # Закрываем диалоговое окно
        
        if self.current_well_id is not None:
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                # Удаляем скважину по её уникальному первичному ключу ID
                cursor.execute("DELETE FROM wells WHERE id=?", (self.current_well_id,))
                conn.commit()
                conn.close()
                
                # Автоматически обновляем список скважин на главном экране Журнала, чтобы карточка исчезла
                open_screen = self.manager.get_screen("open_well_screen")
                if hasattr(open_screen, 'on_enter'):
                    open_screen.on_enter() # Вызываем перезапуск чтения базы данных на экране списка
                    
                # Возвращаем пользователя назад в Журнал
                self.manager.current = "open_well_screen"
            except Exception as e:
                print(f"Ошибка при удалении скважины из SQLite: {e}")

    def share_well(self):
        """Инженерный метод: собирает параметры в текстовый рапорт для отправки в мессенджеры"""
        perf_summary = f"{self.ids.v_perf_top.text}-{self.ids.v_perf_bot.text}" if self.ids.v_perf_bot.text else self.ids.v_perf_top.text
        report = (
            f" ПЛАН РАБОТ: Скв. №{self.ids.v_well.text} (К-{self.ids.v_pad.text})\n"
            f" {self.ids.v_field.text} м-р\n"
            f" Давления: Ртруб={self.ids.v_p_buf.text}, Рзатр={self.ids.v_p_zatr.text}, Рпл={self.ids.v_p_pl.text} атм\n"
            f" Текущая плотность раствора: {self.ids.v_rho_c.text} г/см³\n"
            f" Забой: {self.ids.v_depth.text} м | Перфорация: {perf_summary} м\n"
            f" Конструкция: Диаметр ЭК={self.ids.v_d_ek.text} мм\n"
            f" Подвеска: {self.ids.v_nkt_type.text} на гл. {self.ids.v_h_gno.text} м\n"
            f" Бригада: {self.ids.v_brigade.text} | Операция: {self.ids.v_operation.text}\n"
            f" Заметки: {self.ids.v_notes.text}"
        )
        if platform == 'android':
            from jnius import autoclass
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            Intent = autoclass('android.content.Intent')
            String = autoclass('java.lang.String')
            shareIntent = Intent(Intent.ACTION_SEND).setType("text/plain")
            shareIntent.putExtra(Intent.EXTRA_TEXT, String(report))
            PythonActivity.mActivity.startActivity(Intent.createChooser(shareIntent, String("Поделиться планом работ")))
        else:
            from kivy.core.clipboard import Clipboard
            Clipboard.copy(report)

    def open_in_glushenie(self):
        """Технологический метод: экспортирует геометрические параметры скважины в калькулятор ЦА-320"""
        calc_screen = self.manager.get_screen("calc")
        calc_screen.f_p_pl.text = self.ids.v_p_pl.text
        calc_screen.f_ckod.text = self.ids.v_depth.text
        calc_screen.f_h_perf.text = self.ids.v_perf_top.text
        calc_screen.f_d_col.text = self.ids.v_d_ek.text
        calc_screen.f_rho.text = self.ids.v_rho_c.text
        
        if hasattr(calc_screen, 'f_l_hvost') and calc_screen.f_l_hvost:
            calc_screen.f_l_hvost.text = self.ids.v_l_hvost.text
            
        if hasattr(calc_screen, 'btn_nkt') and calc_screen.btn_nkt:
            calc_screen.btn_nkt.text = self.ids.v_nkt_type.text
        if hasattr(calc_screen, 'selected_nkt'):
            calc_screen.selected_nkt = self.ids.v_nkt_type.text
            
        calc_screen.refresh_ui_elements()
        if hasattr(calc_screen, 'gno_fields_box') and calc_screen.gno_fields_box:
            from kivymd.uix.textfield import MDTextField
            for child in calc_screen.gno_fields_box.children:
                if isinstance(child, MDTextField):
                    hint = getattr(child, 'hint_text', '').lower()
                    if 'глубина' in hint or 'спуск' in hint or 'длина' in hint:
                        child.text = self.ids.v_h_gno.text
                        break
        self.manager.current = "calc"


# --- ТОЧКА ВХОДА ---
if __name__ == "__main__":
    KrsBookApp().run()
